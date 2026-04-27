import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import numpy as np
import os
import base64
import requests
import re
import calendar
from pathlib import Path
from datetime import datetime
from github import Github
from github import Auth
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from st_aggrid import AgGrid, GridOptionsBuilder
from st_aggrid import JsCode
import time
from st_aggrid import GridUpdateMode
import uuid


st.markdown("""
<style>

.loading-container{
    display:flex;
    flex-direction:column;
    align-items:center;
    justify-content:center;
    height:70vh;
}

.loading-logo{
    width:280px;
    margin-bottom:20px;
}

.loading-title{
    font-size:40px;
    font-weight:700;
}

.loading-sub{
    font-size:40px;
    color:#64748b;
}

</style>
""", unsafe_allow_html=True)


# ===============================
# SISTEM NOTIFIKASI LOADING
# ===============================
def add_notification(msg):
    
    if "loading_notifications" not in st.session_state:
        st.session_state.loading_notifications = []

    # supaya tidak dobel
    if msg not in st.session_state.loading_notifications:
        st.session_state.loading_notifications.append(msg)

@st.cache_data(show_spinner=False)
def build_base_grid_options(columns):
    from st_aggrid import GridOptionsBuilder
    
    gb = GridOptionsBuilder()
    
    for col in columns:
        gb.configure_column(col)
    
    return gb

@st.cache_data(show_spinner=False)
def _detect_numeric_cols(col_values_tuple):
    """Cache deteksi kolom numerik berdasarkan sample data."""
    exclude_cols = {
        "__rowNum__", "Kode Satker", "KODE SATKER", "SATKER",
        "Nama Satker", "NAMA SATKER", "Uraian Satker-RINGKAS"
    }
    numeric_cols = set()
    for col, sample_tuple in col_values_tuple:
        if col in exclude_cols:
            continue
        sample = pd.Series(list(sample_tuple)).astype(str).dropna()
        if len(sample) == 0:
            continue
        if sample.str.contains(r"\d").mean() > 0.7:
            numeric_cols.add(col)
    return numeric_cols

def render_table_pin_satker(df):
    df = df.copy()

    if "__rowNum__" in df.columns:
        df = df.drop(columns="__rowNum__")

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df.insert(0, "__rowNum__", range(1, len(df) + 1))

    def calc_grid_height(df, row_height=45, header_height=40, max_height=600):
        min_rows = 5

        total_rows = max(len(df), min_rows)

        height = header_height + total_rows * row_height

        return min(height, max_height)

    gb = build_base_grid_options(tuple(df.columns))
    
    # =====================================================
    # ALIGNMENT OTOMATIS (CACHED)
    # =====================================================
    exclude_cols = {
        "__rowNum__",
        "Kode Satker",
        "KODE SATKER",
        "SATKER",
        "Nama Satker",
        "NAMA SATKER",
        "Uraian Satker-RINGKAS"
    }

    # Buat tuple hashable untuk cache: (col, sample_values[:30])
    col_samples = tuple(
        (col, tuple(df[col].astype(str).dropna().head(30).tolist()))
        for col in df.columns
        if col not in exclude_cols
    )
    numeric_cols = _detect_numeric_cols(col_samples)

    for col in df.columns:
        if col in exclude_cols:
            gb.configure_column(col, cellStyle={"textAlign": "left"})
        elif col in numeric_cols:
            gb.configure_column(col, cellStyle={"textAlign": "right"})
    
    # =====================================================
    # SEMBUNYIKAN KOLOM INTERNAL (JIKA ADA)
    # =====================================================
    if "Nilai Total" in df.columns:
        gb.configure_column("Nilai Total", hide=True)

    # =====================================================
    # CELL POPUP RENDERER (KLIK = POPUP DI TABEL)
    # =====================================================
    cell_popup_renderer = JsCode("""
    class CellPopupRenderer {
        init(params) {
            this.eGui = document.createElement('span');
            this.eGui.innerText = params.value;
            this.eGui.style.cursor = 'pointer';
            this.eGui.style.fontWeight = '600';

            const popupMap = {

            "Kualitas Perencanaan Anggaran": `
                <b>Kualitas Perencanaan Anggaran</b><br/><br/>

                <span style="color:#d1d5db">
                Aspek ini mengukur seberapa baik Satker dalam merencanakan anggaran.
                Penilaian dilakukan terhadap kesesuaian pelaksanaan anggaran dengan yang
                direncanakan dalam DIPA. Semakin sedikit revisi dan semakin sesuai realisasi
                dengan rencana, semakin tinggi nilainya.
                </span>

                <br/><br/>
                <b>Bobot:</b><br/>
                <span style="color:#e5e7eb">
                ● Revisi DIPA: <b>10%</b><br/>
                ● Deviasi Halaman III DIPA: <b>15%</b><br/>
                </span>

                <br/>
                <b>Total: 25%</b>
            `,

            "Kualitas Pelaksanaan Anggaran": `
                <b>Kualitas Pelaksanaan Anggaran</b><br/><br/>

                <span style="color:#d1d5db">
                Aspek ini mengukur kemampuan Satker dalam merealisasikan anggaran yang telah
                ditetapkan. Mencakup kecepatan penyerapan anggaran, kelengkapan kontrak,
                ketepatan pembayaran, dan pengelolaan uang persediaan.
                </span>

                <br/><br/>
                <b>Bobot:</b><br/>
                <span style="color:#e5e7eb">
                ● Penyerapan Anggaran: <b>20%</b><br/>
                ● Belanja Kontraktual: <b>10%</b><br/>
                ● Penyelesaian Tagihan: <b>10%</b><br/>
                ● Pengelolaan UP/TUP: <b>10%</b><br/>
                </span>

                <br/>
                <b>Total: 50%</b>
            `,

            "Kualitas Hasil Pelaksanaan Anggaran": `
                <b>Kualitas Hasil Pelaksanaan Anggaran</b><br/><br/>
                <span style="color:#d1d5db">
                Aspek ini mengukur kemampuan Satker dalam mencapai output atau target kegiatan
                yang telah ditetapkan dalam DIPA. Penilaian didasarkan pada ketepatan waktu
                pelaporan serta tingkat ketercapaian volume output kegiatan.
                </span>
                <br/><br/>
                <b>Bobot:</b><br/>
                <span style="color:#e5e7eb">
                25% (Capaian Output)
                </span>
            `,

            "Dispensasi SPM (Pengurang)": `
                <b>Dispensasi SPM (Pengurang Nilai)</b><br/><br/>
                <span style="color:#d1d5db">
                Indikator pengurang yang diberlakukan apabila Satker mengajukan SPM
                melebihi batas waktu di akhir tahun anggaran.
                Semakin banyak dispensasi, semakin besar pengurangan nilai IKPA.
                </span>
                <br/><br/>
                <b>Pengurangan Nilai:</b>
                <table style="width:100%;margin-top:6px;border-collapse:collapse;font-size:11px">
                <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Rasio (‰)</th>
                    <th style="padding:4px;border:1px solid #4b5563">Kategori</th>
                    <th style="padding:4px;border:1px solid #4b5563">Pengurangan</th>
                </tr>
                <tr><td style="padding:4px;border:1px solid #4b5563">0</td><td style="padding:4px;border:1px solid #4b5563">Tidak ada</td><td style="padding:4px;border:1px solid #4b5563">0</td></tr>
                <tr><td style="padding:4px;border:1px solid #4b5563">0,01 – 0,099</td><td style="padding:4px;border:1px solid #4b5563">Kategori 2</td><td style="padding:4px;border:1px solid #4b5563">0,25</td></tr>
                <tr><td style="padding:4px;border:1px solid #4b5563">0,1 – 0,99</td><td style="padding:4px;border:1px solid #4b5563">Kategori 3</td><td style="padding:4px;border:1px solid #4b5563">0,50</td></tr>
                <tr><td style="padding:4px;border:1px solid #4b5563">1 – 4,99</td><td style="padding:4px;border:1px solid #4b5563">Kategori 4</td><td style="padding:4px;border:1px solid #4b5563">0,75</td></tr>
                <tr><td style="padding:4px;border:1px solid #4b5563">≥ 5,00</td><td style="padding:4px;border:1px solid #4b5563">Kategori 5</td><td style="padding:4px;border:1px solid #4b5563">1,00</td></tr>
                </table>
                <br/>
                <small style="color:#9ca3af">
                Rasio = (Jumlah SPM Dispensasi / Jumlah SPM Triwulan IV) × 1.000
                </small>
            `,
            
            "Revisi DIPA": `
                <b>Revisi DIPA</b><br/><br/>

                <span style="color:#d1d5db">
                Mengukur frekuensi revisi DIPA dalam satu semester. Revisi yang dihitung
                adalah revisi dengan pagu tetap yang menjadi kewenangan Kementerian Keuangan.
                </span>

                <br/><br/>
                <b>Nilai Berdasarkan Frekuensi:</b>

                <table style="width:100%;margin-top:6px;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Jumlah Revisi / Semester</th>
                    <th style="padding:4px;border:1px solid #4b5563">Nilai</th>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">0 – 1</td>
                    <td style="padding:4px;border:1px solid #4b5563">110</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">2</td>
                    <td style="padding:4px;border:1px solid #4b5563">100</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">≥ 3</td>
                    <td style="padding:4px;border:1px solid #4b5563">50</td>
                    </tr>
                </table>

                <br/>
                <small style="color:#9ca3af">
                IKPA Revisi DIPA = (50% × Nilai Semester I) + (50% × Nilai Semester II)
                </small>
            `,
            
            "Deviasi Halaman III DIPA": `
                <b>Deviasi Halaman III DIPA</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur kesesuaian realisasi anggaran dengan Rencana Penarikan Dana (RPD)
                bulanan. Deviasi dihitung dari selisih realisasi terhadap RPD pada setiap
                jenis belanja.
                </span>

                <br/><br/>
                <b>Nilai Optimum:</b><br/>
                <span style="color:#e5e7eb">
                Deviasi ≤ 5% → <b>Nilai 100</b>
                </span>

                <br/><br/>
                <small style="color:#9ca3af">
                Deviasi = |Realisasi – RPD| / RPD × 100%<br/>
                IKPA = 100 – Rata-rata Deviasi Tertimbang (maksimum 100)
                </small>
                `,

                "Penyerapan Anggaran": `
                <b>Penyerapan Anggaran</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur kesesuaian realisasi anggaran dengan <b>TARGET MINIMUM</b><span style="color:#d1d5db"> triwulanan untuk
                masing-masing jenis belanja.
                </span>

                <br/><br/>
                <b>Target Penyerapan Minimal per Triwulan:</b>

                <table style="width:100%;margin-top:6px;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Jenis Belanja</th>
                    <th style="padding:4px;border:1px solid #4b5563">Tw I</th>
                    <th style="padding:4px;border:1px solid #4b5563">Tw II</th>
                    <th style="padding:4px;border:1px solid #4b5563">Tw III</th>
                    <th style="padding:4px;border:1px solid #4b5563">Tw IV</th>
                    </tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Belanja Pegawai</td><td style="padding:4px;border:1px solid #4b5563">20%</td><td style="padding:4px;border:1px solid #4b5563">50%</td><td style="padding:4px;border:1px solid #4b5563">75%</td><td style="padding:4px;border:1px solid #4b5563">95%</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Belanja Barang</td><td style="padding:4px;border:1px solid #4b5563">15%</td><td style="padding:4px;border:1px solid #4b5563">50%</td><td style="padding:4px;border:1px solid #4b5563">70%</td><td style="padding:4px;border:1px solid #4b5563">90%</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Belanja Modal</td><td style="padding:4px;border:1px solid #4b5563">10%</td><td style="padding:4px;border:1px solid #4b5563">40%</td><td style="padding:4px;border:1px solid #4b5563">70%</td><td style="padding:4px;border:1px solid #4b5563">90%</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Belanja Bansos</td><td style="padding:4px;border:1px solid #4b5563">25%</td><td style="padding:4px;border:1px solid #4b5563">50%</td><td style="padding:4px;border:1px solid #4b5563">75%</td><td style="padding:4px;border:1px solid #4b5563">95%</td></tr>
                </table>

                <br/>
                <small style="color:#9ca3af">
                Nilai = (Realisasi / Target) × 100 (maksimum 100)<br/>
                IKPA = Rata-rata Tertimbang seluruh Jenis Belanja
                </small>
            `,

                "Belanja Kontraktual": `
                <b>Belanja Kontraktual</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur upaya akselerasi pengadaan barang/jasa melalui kontrak dini,
                penyelesaian belanja modal, dan distribusi kontrak di semester I.
                </span>

                <br/><br/>
                <table style="width:100%;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Komponen</th>
                    <th style="padding:4px;border:1px solid #4b5563">Bobot</th>
                    <th style="padding:4px;border:1px solid #4b5563">Keterangan</th>
                    </tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Akselerasi Kontrak Dini</td><td style="padding:4px;border:1px solid #4b5563">40%</td><td style="padding:4px;border:1px solid #4b5563">Sebelum 1 Jan = 120; Jan–Mar = 110</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Akselerasi Belanja Modal</td><td style="padding:4px;border:1px solid #4b5563">40%</td><td style="padding:4px;border:1px solid #4b5563">Tw I=100, Tw II=90, Tw III=80, Tw IV=70</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Distribusi Kontrak</td><td style="padding:4px;border:1px solid #4b5563">20%</td><td style="padding:4px;border:1px solid #4b5563">&gt;75%=100, 50–75%=80, 25–50%=60, 0–25%=50</td></tr>
                </table>
            `,

                "Penyelesaian Tagihan": `
                <b>Penyelesaian Tagihan</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur ketepatan waktu penyelesaian tagihan kontraktual (SPM-LS) dari
                tanggal BAST/BAPP sampai diterima KPPN. Batas waktu tepat maksimal
                7 hari kerja.
                </span>

                <br/><br/>
                <small style="color:#9ca3af">
                IKPA = (SPM-LS Tepat Waktu / Total SPM-LS) × 100
                </small>
            `,

                "Pengelolaan UP dan TUP": `
                <b>Pengelolaan UP dan TUP</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur ketepatan waktu pertanggungjawaban UP/TUP, efisiensi besaran UP,
                serta penggunaan Kartu Kredit Pemerintah (KKP).
                </span>

                <br/><br/>
                <b>Komponen UP/TUP Tunai (90%):</b>
                <table style="width:100%;border-collapse:collapse;font-size:11px;margin-top:6px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Sub Komponen</th>
                    <th style="padding:4px;border:1px solid #4b5563">Bobot</th>
                    <th style="padding:4px;border:1px solid #4b5563">Keterangan</th>
                    </tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Ketepatan Waktu</td><td style="padding:4px;border:1px solid #4b5563">50%</td><td style="padding:4px;border:1px solid #4b5563">Tepat waktu=100; Terlambat=0</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Persentase GUP</td><td style="padding:4px;border:1px solid #4b5563">25%</td><td style="padding:4px;border:1px solid #4b5563">GUP disetarakan bulanan</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Setoran TUP</td><td style="padding:4px;border:1px solid #4b5563">25%</td><td style="padding:4px;border:1px solid #4b5563">Rasio setoran ke kas negara</td></tr>
                </table>

                <br/>
                <b>Komponen KKP (10%):</b><br/>
                <span style="color:#e5e7eb">
                Mencapai target = 110<br/>
                Belum mencapai target = 100
                </span>
            `,

                "Capaian Output": `
                <b>Capaian Output</b><br/><br/>
                <span style="color:#d1d5db">
                Mengukur ketepatan waktu pelaporan dan tingkat ketercapaian target
                Rincian Output (RO).
                </span>

                <br/><br/>
                <table style="width:100%;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Komponen</th>
                    <th style="padding:4px;border:1px solid #4b5563">Bobot</th>
                    <th style="padding:4px;border:1px solid #4b5563">Keterangan</th>
                    </tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Ketepatan Waktu</td><td style="padding:4px;border:1px solid #4b5563">30%</td><td style="padding:4px;border:1px solid #4b5563">≤5 hari kerja = 100</td></tr>
                    <tr><td style="padding:4px;border:1px solid #4b5563">Capaian RO</td><td style="padding:4px;border:1px solid #4b5563">70%</td><td style="padding:4px;border:1px solid #4b5563">Realisasi / Target × 100</td></tr>
                </table>
            `,
            
            "nilai_akhir_aspek": `
                <b>Nilai Akhir (Konversi Bobot)</b><br/><br/>

                <span style="color:#d1d5db">
                Nilai akhir IKPA dihitung berdasarkan pengelompokan aspek utama
                pengelolaan anggaran. Nilai ini mencerminkan kinerja Satker secara
                komprehensif berdasarkan perencanaan, pelaksanaan, dan hasil
                pelaksanaan anggaran.
                </span>

                <br/><br/>
                <b>Formula:</b><br/>
                <small style="color:#9ca3af">
                Nilai Akhir = Σ(Nilai Aspek × Bobot Aspek)
                </small>

                <br/><br/>
                <b>Bobot Aspek:</b>

                <table style="width:100%;margin-top:6px;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Aspek</th>
                    <th style="padding:4px;border:1px solid #4b5563">Bobot</th>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">Kualitas Perencanaan Anggaran</td>
                    <td style="padding:4px;border:1px solid #4b5563">25%</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">Kualitas Pelaksanaan Anggaran</td>
                    <td style="padding:4px;border:1px solid #4b5563">50%</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">Kualitas Hasil Pelaksanaan Anggaran</td>
                    <td style="padding:4px;border:1px solid #4b5563">25%</td>
                    </tr>
                </table>
            `,

            "nilai_akhir_komponen": `
                <b>Nilai Akhir (Total / Konversi)</b><br/><br/>

                <span style="color:#d1d5db">
                Perhitungan akhir yang menyimpulkan kinerja Satker secara keseluruhan.
                Nilai akhir diperoleh dari agregasi seluruh indikator komponen yang
                dibobotkan dan dikonversi, kemudian dikurangi nilai Dispensasi SPM.
                </span>

                <br/><br/>
                <b>Kategori Nilai IKPA:</b>

                <table style="width:100%;margin-top:6px;border-collapse:collapse;font-size:11px">
                    <tr style="background:#374151">
                    <th style="padding:4px;border:1px solid #4b5563">Nilai IKPA</th>
                    <th style="padding:4px;border:1px solid #4b5563">Predikat</th>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">≥ 95</td>
                    <td style="padding:4px;border:1px solid #4b5563">Sangat Baik</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">89 – &lt; 95</td>
                    <td style="padding:4px;border:1px solid #4b5563">Baik</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">70 – &lt; 89</td>
                    <td style="padding:4px;border:1px solid #4b5563">Cukup</td>
                    </tr>
                    <tr>
                    <td style="padding:4px;border:1px solid #4b5563">&lt; 70</td>
                    <td style="padding:4px;border:1px solid #4b5563">Kurang</td>
                    </tr>
                </table>

                <br/>
                <b>Rumus Final:</b><br/>
                <small style="color:#9ca3af">
                Nilai Akhir = [Σ(Indikator × Bobot) × Konversi Bobot] − Dispensasi SPM
                </small>
            `          
        };

            this.eGui.addEventListener('click', (e) => {
                e.stopPropagation();

                window.top.document
                    .querySelectorAll('.cell-popup')
                    .forEach(el => el.remove());
                const popup = document.createElement('div');
                popup.className = 'cell-popup';
                popup.style.position = 'fixed';
                popup.style.background = '#1f2937';
                popup.style.color = '#ffffff';
                popup.style.padding = '10px';
                popup.style.borderRadius = '8px';
                popup.style.fontSize = '12px';
                popup.style.zIndex = 2147483647; // MAX
                popup.style.boxShadow = '0 10px 30px rgba(0,0,0,0.5)';
                popup.style.maxWidth = '380px';
                const header = params.colDef.headerName;
                const field  = params.colDef.field;

                let popupKey = field || header;

                if (header === "Nilai Akhir (Nilai Total/Konversi Bobot)") {

                    // DETEKSI BERDASARKAN STRUKTUR TABEL
                    const allFields = params.api.getColumnDefs().map(c => c.field);

                    // kalau ada kolom ASPEK → popup aspek
                    if (allFields.includes("Kualitas Perencanaan Anggaran")) {
                        popupKey = "nilai_akhir_aspek";
                    } 
                    // kalau ada kolom KOMPONEN → popup komponen
                    else if (allFields.includes("Revisi DIPA")) {
                        popupKey = "nilai_akhir_komponen";
                    }
                }


                popup.innerHTML = popupMap[popupKey] || "Tidak ada keterangan";


                //  TEMBUS IFRAME STREAMLIT
                window.top.document.body.appendChild(popup);

                // HITUNG POSISI AMAN LAYAR
                const offset = 14;
                const vw = window.top.innerWidth;
                const vh = window.top.innerHeight;

                let left = e.clientX + offset;
                let top  = e.clientY - popup.offsetHeight - offset;

                if (left + popup.offsetWidth > vw - 8) {
                    left = e.clientX - popup.offsetWidth - offset;
                }
                if (top < 8) {
                    top = e.clientY + offset;
                }
                if (top + popup.offsetHeight > vh - 8) {
                    top = vh - popup.offsetHeight - 8;
                }
                if (left < 8) {
                    left = 8;
                }

                popup.style.left = left + 'px';
                popup.style.top  = top  + 'px';

                window.top.document.addEventListener(
                    'click',
                    () => popup.remove(),
                    { once: true }
                );
            });
        }

        getGui() {
            return this.eGui;
        }
    }
    """)


    # =====================================================
    # PASANG POPUP KE KOLOM NILAI
    # =====================================================
    POPUP_COLUMNS = [
        "Kualitas Perencanaan Anggaran",
        "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",
        "Revisi DIPA",
        "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran",
        "Belanja Kontraktual",
        "Penyelesaian Tagihan",
        "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Dispensasi SPM (Pengurangan)",
        "Dispensasi SPM (Pengurang)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)"
    ]

    for col in POPUP_COLUMNS:
        if col in df.columns:
            gb.configure_column(col, cellRenderer=cell_popup_renderer)

    # =====================================================
    # KOLOM NOMOR
    # =====================================================
    gb.configure_column(
        "__rowNum__",
        headerName="No",
        pinned="left",
        width=60,
        sortable=False,
        filter=False,
        cellStyle={"textAlign": "center"}
    )

    # =====================================================
    # DEFAULT COLUMN (🔥 FIX UTAMA)
    # =====================================================
    gb.configure_default_column(
        resizable=True,
        sortable=True,
        filter=True,
        minWidth=130   
    )

    # =====================================================
    # KOLOM BULAN (KECIL)
    # =====================================================
    bulan_cols = [
        "Jan","Feb","Mar","Apr","Mei","Jun",
        "Jul","Agu","Sep","Okt","Nov","Des"
    ]

    for col in bulan_cols:
        if col in df.columns:
            gb.configure_column(
                col,
                width=70,   # 🔥 FIX (pakai width, bukan min/max)
                cellStyle={"textAlign": "right"}
            )

    # =====================================================
    # KOLOM KECIL (PERINGKAT & BA)
    # =====================================================
    small_cols = [
        "Peringkat",
        "Kode BA",
        "Kode_BA",
        "BA"
    ]

    for col in small_cols:
        if col in df.columns:
            gb.configure_column(
                col,
                width=110,
                cellStyle={"textAlign": "center"}
            )

    # =====================================================
    # KOLOM TEKS
    # =====================================================
    text_columns = [
        "Kode Satker",
        "Uraian Satker-RINGKAS",
        "SATKER"
    ]

    for col in text_columns:
        if col in df.columns:
            gb.configure_column(
                col,
                cellStyle={"textAlign": "left"}
            )

    # =====================================================
    # KOLOM PINNED
    # =====================================================
    if "Uraian Satker-RINGKAS" in df.columns:
        gb.configure_column(
            "Uraian Satker-RINGKAS",
            headerName="Nama Satker",
            pinned="left",
            width=220   # 🔥 lebih lega
        )

    if "Kode Satker" in df.columns:
        gb.configure_column(
            "Kode Satker",
            pinned="left",
            width=100
        )

    # =====================================================
    # ZEBRA STYLE
    # =====================================================
    zebra_dark = JsCode("""
    function(params) {
        return {
            backgroundColor: params.node.rowIndex % 2 === 0 ? '#3D3D3D' : '#050505',
            color: '#FFFFFF'
        };
    }
    """)

    # =====================================================
    # GRID OPTIONS 
    # =====================================================
    gb.configure_grid_options(
        domLayout="normal",
        alwaysShowHorizontalScroll=True,
        suppressHorizontalScroll=False,
        getRowStyle=zebra_dark,
        headerHeight=40
    )

    # =====================================================
    # GRID
    # =====================================================
    grid_response = AgGrid(
        df,
        gridOptions=gb.build(),
        height=max(450, calc_grid_height(df)),
        width="100%",
        theme="streamlit",
        allow_unsafe_jscode=True,
        data_return_mode="FILTERED_AND_SORTED",
        update_mode="MODEL_CHANGED",
    )

    # ===== AMBIL DATA HASIL FILTER =====
    filtered_df = pd.DataFrame(grid_response["data"])

    if "__rowNum__" in filtered_df.columns:
        filtered_df = filtered_df.drop(columns="__rowNum__")

    # ===== STYLE MINI BUTTON =====
    st.markdown("""
        <style>
        div.stDownloadButton {
            text-align: right;
        }

        div.stDownloadButton > button {
            background: #4F46E5 !important;
            color: white !important;
            border: none !important;
            border-radius: 20px !important;

            font-size: 11px !important;
            padding: 6px 18px !important;
            height: 36px !important;
            font-weight: 500 !important;
        }

        div.stDownloadButton > button:hover {
            background: #4338CA !important;
        }
        </style>
        """, unsafe_allow_html=True)

    # ===== EXPORT BUTTON =====
    st.download_button(
        "Export Excel",
        data=to_excel_bytes(filtered_df),
        file_name="Data_Satker.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# AMBIL PASSWORD DARI SECRETS
# =========================
ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "")
if not ADMIN_PASSWORD:
    st.error("ADMIN_PASSWORD belum diset di Streamlit Secrets")
    st.stop()


# define month order map
MONTH_ORDER = {
    'JANUARI': 1,
    'FEBRUARI': 2, 'PEBRUARI': 2, 'PEBRUARY': 2,
    'MARET': 3, 'MAR': 3, 'MRT': 3,
    'APRIL': 4,
    'MEI': 5,
    'JUNI': 6,
    'JULI': 7,
    'AGUSTUS': 8, 'AGUSTUSS': 8,
    'SEPTEMBER': 9, 'SEPT': 9, 'SEP': 9,
    'OKTOBER': 10,
    'NOVEMBER': 11, 'NOPEMBER': 11,
    'DESEMBER': 12
}

# Path ke file template (akan diatur di session state)
TEMPLATE_PATH = r"C:\Users\KEMENKEU\Desktop\INDIKATOR PELAKSANAAN ANGGARAN.xlsx"

def detect_ikpa_type(df_raw):
    
    text = " ".join(
        df_raw.astype(str)
        .iloc[:5]
        .values
        .flatten()
    ).upper()

    if "KODE SATKER" in text:
        return "SATKER"

    elif "KPPN" in text:
        return "KPPN"

    else:
        return "UNKNOWN"

# Normalize kode satker
def normalize_kode_satker(k, width=6):
    if pd.isna(k):
        return ''

    digits = re.findall(r'\d+', str(k))
    if not digits:
        return ''

    # ambil angka pertama (format standar satker)
    kode = digits[0]

    return kode.zfill(width)



@st.cache_data(show_spinner=False)
def load_reference_satker():
    """
    Load referensi nama satker ringkas.
    WAJIB punya kolom:
    - Kode Satker
    - Uraian Satker-SINGKAT
    """
    try:
        url = (
            "https://raw.githubusercontent.com/ameernoor/kinerjakeuangansatkerbta/main/templates/Template_Data_Referensi.xlsx"
        )

        ref = pd.read_excel(url, dtype=str)

        # ===============================
        # NORMALISASI WAJIB
        # ===============================
        ref["Kode Satker"] = (
            ref["Kode Satker"]
            .apply(normalize_kode_satker)
            .astype(str)
            .str.strip()
        )

        ref["Uraian Satker-SINGKAT"] = (
            ref["Uraian Satker-SINGKAT"]
            .astype(str)
            .str.strip()
        )

        # ===============================
        # BUANG DATA TIDAK VALID
        # ===============================
        ref = ref[
            (ref["Kode Satker"] != "") &
            (ref["Kode Satker"].notna()) &
            (ref["Uraian Satker-SINGKAT"] != "") &
            (ref["Uraian Satker-SINGKAT"].notna())
        ].copy()

        return ref

    except Exception as e:
        st.error(f"❌ Referensi satker gagal dimuat: {e}")
        return pd.DataFrame(columns=["Kode Satker", "Uraian Satker-SINGKAT"])


# ================================
# INIT SESSION STATE 
# ================================
# Storage utama IKPA
if "data_storage" not in st.session_state:
    st.session_state.data_storage = {}

# Storage IKPA per KPPN
if "data_storage_kppn" not in st.session_state:
    st.session_state.data_storage_kppn = {}

# Storage DIPA per tahun
if "DATA_DIPA_by_year" not in st.session_state:
    st.session_state.DATA_DIPA_by_year = {}

# Flag merge IKPA–DIPA
if "ikpa_dipa_merged" not in st.session_state:
    st.session_state.ikpa_dipa_merged = False
    
if "data_storage_kkp" not in st.session_state:
    st.session_state.data_storage_kkp = {}
    
if "data_storage_digipay" not in st.session_state:
    st.session_state.data_storage_digipay = {}

# Log aktivitas
if "activity_log" not in st.session_state:
    st.session_state.activity_log = []

def log_activity(menu, action, detail=""):
    st.session_state.activity_log.insert(
        0,
        {
            "Waktu": datetime.now().strftime("%H:%M:%S"),
            "Menu": menu,
            "Aktivitas": action,
            "Detail": detail
        }
    )


# reference
if "_reference_loaded" not in st.session_state:
    st.session_state.reference_df = load_reference_satker()
    st.session_state["_reference_loaded"] = True

# ===============================
# 🔥 CLEAN NUMERIC (ANTI NAN)
# ===============================
def clean_numeric(val):
    if pd.isna(val):
        return 0
    # Jika sudah numerik (int/float), langsung return — jangan diproses lagi
    if isinstance(val, (int, float)):
        return float(val)
    val = str(val).strip().replace("%", "")
    # Format Indonesia: koma = desimal, titik = ribuan -> '88,98' atau '1.234,56'
    if "," in val:
        val = val.replace(".", "").replace(",", ".")
    # Format sudah Python float: '88.98' — jangan hapus titik
    val = re.sub(r"[^\d\.\-]", "", val)
    try:
        return float(val)
    except:
        return 0
    
def safe_float(val, default=0):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "."))
    except:
        return default

def safe_upper(val):
    if pd.isna(val):
        return ""
    return str(val).upper().strip()


MONTH_MAP = {
    "1": "JANUARI", "2": "FEBRUARI", "3": "MARET",
    "4": "APRIL", "5": "MEI", "6": "JUNI",
    "7": "JULI", "8": "AGUSTUS", "9": "SEPTEMBER",
    "10": "OKTOBER", "11": "NOVEMBER", "12": "DESEMBER"
}

def normalize_month(val):
    val = str(val).strip()
    return MONTH_MAP.get(val, safe_upper(val))


def fix_ikpa_header(df_raw):
    
    import pandas as pd

    df = None  # 🔥 ANTI ERROR WAJIB

    try:
        for i in range(min(10, len(df_raw))):
            row = df_raw.iloc[i].astype(str).str.upper()
            row_text = " ".join(row.values)

            if ("KODE" in row_text and "SATKER" in row_text):

                header_row = i

                header1 = df_raw.iloc[header_row]
                header2 = df_raw.iloc[header_row + 1]

                # 🔥 DETEKSI HEADER 3 BARIS (KHUSUS MARET)
                row_next = " ".join(
                    df_raw.iloc[header_row + 2].astype(str).str.upper().values
                )

                if any(x in row_next for x in ["REVISI", "DEVIASI", "PENYERAPAN"]):
                    header2 = df_raw.iloc[header_row + 2]
                    data_start = header_row + 3
                else:
                    data_start = header_row + 2

                cols = []

                for h1, h2 in zip(header1, header2):

                    h1 = str(h1).strip()
                    h2 = str(h2).strip()

                    col = ""

                    if "KODE SATKER" in h1.upper():
                        col = "Kode Satker"
                    elif "URAIAN SATKER" in h1.upper():
                        col = "Uraian Satker"
                    elif "KODE KPPN" in h1.upper():
                        col = "Kode KPPN"
                    elif "KODE BA" in h1.upper():
                        col = "Kode BA"
                    elif "PERIODE" in h1.upper():
                        col = "Periode"
                    elif "KETERANGAN" in h1.upper():
                        col = "Keterangan"
                    elif "NO" in h1.upper():
                        col = "No"

                    elif "REVISI" in h2.upper():
                        col = "Revisi DIPA"
                    elif "DEVIASI" in h2.upper():
                        col = "Deviasi Halaman III DIPA"
                    elif "PENYERAPAN" in h2.upper():
                        col = "Penyerapan Anggaran"
                    elif "BELANJA" in h2.upper():
                        col = "Belanja Kontraktual"
                    elif "TAGIHAN" in h2.upper():
                        col = "Penyelesaian Tagihan"
                    elif "OUTPUT" in h2.upper():
                        col = "Capaian Output"

                    elif "NILAI AKHIR" in h1.upper():
                        col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
                    elif "NILAI TOTAL" in h1.upper():
                        col = "Nilai Total"
                    elif "BOBOT" in h1.upper():
                        col = "Konversi Bobot"
                    elif "DISPENSASI" in h1.upper():
                        col = "Dispensasi SPM (Pengurangan)"
                    elif "NILAI ASPEK" in h2.upper():
                        col = "Nilai Aspek"
                    else:
                        col = f"IGNORE_{len(cols)}"

                    cols.append(col)

                df = df_raw.iloc[data_start:].copy()
                df.columns = cols
                df = df.reset_index(drop=True)

                # ===============================
                # FIX DUPLIKAT KOLOM
                # ===============================
                seen = {}
                new_cols = []

                for col in df.columns:
                    if col in seen:
                        seen[col] += 1
                        new_cols.append(f"{col}_{seen[col]}")
                    else:
                        seen[col] = 0
                        new_cols.append(col)

                df.columns = new_cols

                # ===============================
                # MAP NILAI ASPEK
                # ===============================
                aspek_cols = [c for c in df.columns if "Nilai Aspek" in c]

                for idx, col in enumerate(aspek_cols):
                    if idx == 0:
                        df.rename(columns={col: "Nilai Aspek Perencanaan"}, inplace=True)
                    elif idx == 1:
                        df.rename(columns={col: "Nilai Aspek Pelaksanaan"}, inplace=True)
                    elif idx == 2:
                        df.rename(columns={col: "Nilai Aspek Hasil"}, inplace=True)

                # ===============================
                # DROP KOLOM SAMPAH
                # ===============================
                df = df[[c for c in df.columns if not c.startswith("IGNORE")]]

                # ===============================
                # 🔥 FIX KHUSUS FILE MARET
                # ===============================
                if "Keterangan" in df.columns:
                    mask = df["Keterangan"].astype(str).str.upper().str.contains("NILAI", na=False)

                # hanya filter kalau memang ada
                if mask.any():
                    df = df[mask]

                return df

    except Exception as e:
        st.error(f"❌ ERROR parsing IKPA: {e}")
        return None

    return None

def detect_columns(df_raw):
    
    header = df_raw.iloc[2].astype(str).str.upper()

    col_map = {}

    for i, col in enumerate(header):

        if "REVISI" in col:
            col_map["revisi"] = i
        elif "DEVIASI" in col:
            col_map["deviasi"] = i
        elif "PENYERAPAN" in col:
            col_map["serapan"] = i
        elif "KONTRAK" in col:
            col_map["kontrak"] = i
        elif "TAGIHAN" in col:
            col_map["tagihan"] = i
        elif "OUTPUT" in col:
            col_map["output"] = i

    return col_map


def safe_get(row, idx):
    try:
        return float(str(row[idx]).replace(",", "."))
    except:
        return 0


def normalize_ikpa_columns(df):

    rename_map = {
        # ASPEK (dari format lama yang pakai nama berbeda)
        "Nilai Aspek Perencanaan": "Kualitas Perencanaan Anggaran",
        "Nilai Aspek Pelaksanaan": "Kualitas Pelaksanaan Anggaran",
        "Nilai Aspek Hasil": "Kualitas Hasil Pelaksanaan Anggaran",

        # URAIAN — semua variannya dipetakan ke Uraian Satker-RINGKAS
        "Uraian Satker": "Uraian Satker-RINGKAS",
        "Uraian Satker Final": "Uraian Satker-RINGKAS",

        # DISPENSASI
        "Dispensasi SPM (Pengurang)":   "Dispensasi SPM (Pengurangan)",
        "Dispensasi SPM Pengurang":     "Dispensasi SPM (Pengurangan)",
        "Dispensasi SPM Pengurangan":   "Dispensasi SPM (Pengurangan)",
    }

    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    # Jika Uraian Satker-RINGKAS masih belum ada tapi ada kolom lain yang bisa dipakai
    if "Uraian Satker-RINGKAS" not in df.columns:
        for fallback in ["SATKER", "Satker", "Nama Satker"]:
            if fallback in df.columns:
                df["Uraian Satker-RINGKAS"] = df[fallback].astype(str)
                break

    return df


def ensure_ikpa_columns(df):

    import numpy as np

    required_cols = [
        "Kode Satker",
        "Uraian Satker",
        "Uraian Satker-RINGKAS",

        "Kualitas Perencanaan Anggaran",
        "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",

        "Revisi DIPA",
        "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran",
        "Belanja Kontraktual",
        "Penyelesaian Tagihan",
        "Pengelolaan UP dan TUP",
        "Capaian Output",

        "Nilai Total",
        "Konversi Bobot",
        "Dispensasi SPM (Pengurangan)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)"
    ]

    for col in required_cols:
        if col not in df.columns:
            df[col] = np.nan

    return df
    

def fix_dipa_header(df_raw):
    
    import pandas as pd

    # cari header di 15 baris pertama
    for i in range(min(15, len(df_raw))):
        row = df_raw.iloc[i].astype(str).str.upper()

        if (
            row.str.contains("SATKER").any()
            and row.str.contains("PAGU").any()
        ):
            df = df_raw.iloc[i+1:].copy()
            df.columns = df_raw.iloc[i]
            return df.reset_index(drop=True)

    # fallback keras (biar tidak hancur)
    st.error("❌ HEADER DIPA TIDAK DITEMUKAN")
    st.stop()


def standardize_dipa(df_raw):
    
    import re
    from datetime import datetime

    df = df_raw.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # ===============================
    # 🔥 FIND COL (VERSI KUAT)
    # ===============================
    def find_col(possible_names):
        for c in df.columns:
            c_norm = re.sub(r'[^A-Z]', '', str(c).upper())

            for p in possible_names:
                p_norm = re.sub(r'[^A-Z]', '', p.upper())

                if p_norm in c_norm:
                    return c
        return None

    # ===============================
    # DETEKSI KOLOM
    # ===============================
    col_kode = find_col(["Kode Satker", "Satker"])
    col_nama = find_col(["Nama Satker", "Uraian Satker", "Satker"])

    col_pagu = find_col([
        "Total Pagu",
        "Pagu",
        "PAGU",
        "Pagu Belanja",
        "Jumlah",
        "TOTAL",
        "PAGU_RUPIAH"
    ])

    # 🔥 BACKUP DETEKSI PAGU
    if col_pagu is None:
        for c in df.columns:
            if "PAGU" in str(c).upper():
                col_pagu = c
                break


    col_tanggal_revisi = find_col([
        "Tanggal Posting Revisi",
        "Tanggal Revisi",
        "Tgl Revisi",
        "Tgl Posting",
        "Posting Date"
    ])

    col_revisi_ke = find_col(["Revisi Terakhir", "Revisi ke"])
    col_no = find_col(["No"])
    col_kementerian = find_col(["Kementerian", "BA", "K/L"])
    col_dipa = find_col(["No Dipa", "Nomor DIPA"])
    col_tanggal_dipa = find_col(["Tanggal Dipa"])
    col_owner = find_col(["Owner"])
    col_stamp = find_col(["Digital Stamp"])
    col_status_history = find_col(["Kode Status History"])
    col_jenis_revisi = find_col(["Jenis Revisi"])

    # ===============================
    # BUILD OUTPUT
    # ===============================
    out = pd.DataFrame()

    # KODE SATKER
    if col_kode:
        out["Kode Satker"] = df[col_kode].astype(str).str.extract(r"(\d{6})")[0]
        out["Kode Satker"] = out["Kode Satker"].apply(normalize_kode_satker)
    else:
        out["Kode Satker"] = None

    # NAMA
    if col_nama:
        out["Satker"] = df[col_nama].astype(str)
    else:
        out["Satker"] = ""

    # ===============================
    # PAGU (ANTI GAGAL TOTAL)
    # ===============================
    if col_pagu:
        out["Total Pagu"] = (
            df[col_pagu]
            .astype(str)
            .str.replace(r"[^\d]", "", regex=True)
            .replace("", "0")
            .astype(float)
        )
    else:
        st.error("❌ Kolom PAGU tidak ditemukan di file DIPA")
        st.stop()

    # ===============================
    # TANGGAL REVISI
    # ===============================
    if col_tanggal_revisi:
        out["Tanggal Posting Revisi"] = pd.to_datetime(
            df[col_tanggal_revisi],
            errors="coerce"
        )
    else:
        out["Tanggal Posting Revisi"] = pd.NaT

    out["Tanggal Posting Revisi"] = out["Tanggal Posting Revisi"].fillna(
        pd.Timestamp(f"{datetime.now().year}-12-31")
    )

    # TAHUN
    out["Tahun"] = out["Tanggal Posting Revisi"].dt.year.astype(int)

    # NO
    out["NO"] = df[col_no] if col_no else range(1, len(df) + 1)

    # KEMENTERIAN
    if col_kementerian:
        out["Kementerian"] = df[col_kementerian].astype(str)
    else:
        out["Kementerian"] = ""

    # REVISI
    if col_revisi_ke:
        out["Revisi ke-"] = (
            df[col_revisi_ke]
            .astype(str)
            .str.extract(r"(\d+)")
            .fillna(0)
            .astype(int)
        )
    else:
        out["Revisi ke-"] = 0

    # LAINNYA
    out["No Dipa"] = df[col_dipa].astype(str) if col_dipa else ""
    out["Tanggal Dipa"] = pd.to_datetime(df[col_tanggal_dipa], errors="coerce") if col_tanggal_dipa else pd.NaT
    out["Owner"] = df[col_owner].astype(str) if col_owner else ""
    out["Digital Stamp"] = df[col_stamp].astype(str) if col_stamp else ""
    out["Jenis Satker"] = ""
    out["Kode Status History"] = df[col_status_history].astype(str) if col_status_history else ""
    out["Jenis Revisi"] = df[col_jenis_revisi].astype(str) if col_jenis_revisi else ""

    # ===============================
    # FINAL CLEAN
    # ===============================
    out = out.dropna(subset=["Kode Satker"])
    out["Kode Satker"] = out["Kode Satker"].astype(str).str.zfill(6)

    # ===============================
    # SAFETY FINAL
    # ===============================
    out["Total Pagu"] = pd.to_numeric(out["Total Pagu"], errors="coerce").fillna(0)

    return out


#Normalisasi kode BA
def parse_dipa(df_raw):
    import pandas as pd
    import re
    from datetime import datetime

    # ====== 1. Hapus baris kosong ======
    df = df_raw.dropna(how="all").reset_index(drop=True)

    # ====== 2. Cari baris header yang BENAR ======
    header_row = None
    for i in range(min(15, len(df))):
        row_str = " ".join(df.iloc[i].astype(str).str.upper().tolist())
        if (
            "SATKER" in row_str
            and "DIPA" in row_str
            and ("PAGU" in row_str or "BELANJA" in row_str)
        ):
            header_row = i
            break

    if header_row is None:
        raise ValueError("Header DIPA tidak ditemukan (SPAN 2022–2023)")


    # ====== 3. Set header ======
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # ====== 4. Normalisasi kolom ======
    def get(names):
        for c in df.columns:
            cc = str(c).upper().replace(".", "").strip()
            for n in names:
                if n in cc:
                    return c
        return None

    col_no     = get(["NO"])
    col_satker = get(["SATKER"])
    col_nama   = get(["NAMA SATKER"])
    col_dipa   = get(["NO DIPA"])
    col_pagu   = get(["PAGU", "TOTAL PAGU"])
    col_tgl    = get(["TANGGAL POSTING", "TANGGAL REVISI"])
    col_rev    = get(["REVISI TERAKHIR", "REVISI KE"])
    col_tgl_dipa = get(["TANGGAL DIPA"])
    col_owner  = get(["OWNER"])
    col_stamp  = get(["STAMP"])
    col_status = get(["STATUS", "HISTORY"])

    out = pd.DataFrame()

    # NO
    out["NO"] = df[col_no] if col_no else range(1, len(df)+1)

    # Kode Satker
    out["Kode Satker"] = df[col_satker].astype(str).str.extract(r"(\d{6})")[0]

    # Nama Satker
    if col_nama:
        out["Satker"] = df[col_nama].astype(str)
    else:
        out["Satker"] = df[col_satker].astype(str).str.replace(r"^\d{6}\s*-?\s*", "", regex=True)

    # Total Pagu
    out["Total Pagu"] = (
        df[col_pagu].astype(str)
        .str.replace(r"[^\d\-\.]", "", regex=True)
        .replace("", "0")
        .astype(float)
    ) if col_pagu else 0

    # No DIPA
    out["No Dipa"] = df[col_dipa].astype(str)

    # Kementerian (BA)
    out["Kementerian"] = out["No Dipa"].str.extract(r"DIPA-(\d{3})")[0].fillna("")

    # Kode Status History -> BXX
    out["Kode Status History"] = (
        "B" + out["No Dipa"].str.extract(r"DIPA-\d{3}\.(\d{2})")[0].fillna("00")
    )

    # Revisi ke
    if col_rev:
        r = df[col_rev].astype(str).str.extract(r"(\d+)")[0].fillna(0).astype(int)
        out["Revisi ke-"] = r
        out["Jenis Revisi"] = r.apply(lambda x: "DIPA_REVISI" if x > 0 else "ANGKA_DASAR")
    else:
        out["Revisi ke-"] = 0
        out["Jenis Revisi"] = "ANGKA_DASAR"

    # Tanggal DIPA
    out["Tanggal Dipa"] = (
        pd.to_datetime(df[col_tgl_dipa], errors="coerce")
        if col_tgl_dipa else pd.NaT
    )

    # Tanggal Posting Revisi -> format dd-mm-yyyy
    out["Tanggal Posting Revisi"] = (
        pd.to_datetime(df[col_tgl], format="%d-%m-%Y", errors="coerce")
        if col_tgl else pd.NaT
    )

    # Tahun
    out["Tahun"] = (
        out["Tanggal Posting Revisi"].dt.year
            .fillna(datetime.now().year)
            .astype(int)
    )

    # Owner (default untuk 2022–2024)
    out["Owner"] = (
        df[col_owner].astype(str)
        if col_owner else "UNIT"
    )

    # Digital Stamp (default untuk 2022–2024)
    out["Digital Stamp"] = (
        df[col_stamp].astype(str)
        if col_stamp else "0000000000000000"
    )

    # Jenis Satker TIDAK ditentukan di parser
    out["Jenis Satker"] = None

    out = out.dropna(subset=["Kode Satker"])
    out["Kode Satker"] = out["Kode Satker"].astype(str).str.zfill(6)


    # ======================================================
    # 🔑 PERBAIKAN NAMA SATKER (KHUSUS SPAN 2022–2023)
    # ======================================================
    ref = st.session_state.reference_df[
        ["Kode Satker", "Uraian Satker-SINGKAT"]
    ].copy()

    ref["Kode Satker"] = ref["Kode Satker"].astype(str).str.strip()
    out["Kode Satker"] = out["Kode Satker"].astype(str).str.strip()

    out = out.merge(
        ref,
        on="Kode Satker",
        how="left"
    )

    # isi Satker dari referensi JIKA kosong
    out["Satker"] = (
        out["Satker"]
        .replace(["", "nan", "None"], pd.NA)
        .fillna(out["Uraian Satker-SINGKAT"])
    )

    out.drop(columns=["Uraian Satker-SINGKAT"], inplace=True)
 
    return out


# ============================================================
# FUNGSI HELPER: Load Data DIPA dari GitHub
# ============================================================
def auto_process_dipa(df_raw):
    
    # 🔥 FIX HEADER DULU
    df_raw = fix_dipa_header(df_raw)

    # 🔥 LANGSUNG STANDARDIZE (JANGAN PAKAI YANG LAIN)
    df = standardize_dipa(df_raw)

    # safety
    if "Tanggal Posting Revisi" not in df.columns:
        df["Tanggal Posting Revisi"] = pd.Timestamp("2026-12-31")

    df["Tanggal Posting Revisi"] = pd.to_datetime(
        df["Tanggal Posting Revisi"],
        errors="coerce"
    ).fillna(pd.Timestamp("2026-12-31"))

    return df


def is_omspan_dipa(df_raw):
    cols = df_raw.astype(str).apply(lambda x: " ".join(x), axis=1).str.upper()
    return (
        cols.str.contains("OMSPAN").any() or
        cols.str.contains("PAGU_RUPIAH").any() or
        cols.str.contains("KODE_SATKER").any()
    )


# ============================================================
# 🔄 ADAPTER DIPA OMSPAN → FORMAT DIPA STANDAR (FINAL)
# ============================================================
def adapt_dipa_omspan(df_raw):
    df = df_raw.copy()

    # 🔹 cari header otomatis
    for i in range(15):
        row = " ".join(df.iloc[i].astype(str).str.upper())
        if "SATKER" in row and "PAGU" in row:
            df.columns = df.iloc[i]
            df = df.iloc[i+1:]
            break

    df = df.dropna(how="all")

    def find(names):
        for c in df.columns:
            cc = str(c).upper().replace(" ", "").replace("_", "")
            for n in names:
                if n in cc:
                    return c
        return None

    out = pd.DataFrame()

    # ===============================
    # 1️⃣ KODE SATKER (WAJIB)
    # ===============================
    satker_col = find(["KODESATKER", "SATKER"])
    if satker_col is None:
        raise ValueError("Kolom Kode Satker tidak ditemukan di file OMSPAN")

    out["Kode Satker"] = (
        df[satker_col]
        .astype(str)
        .str.extract(r"(\d{6})")[0]
    )

    # ===============================
    # 2️⃣ NAMA SATKER
    # (OMSPAN TIDAK PUNYA → KOSONG DULU)
    # ===============================
    out["Satker"] = pd.NA

    # ===============================
    # 3️⃣ TOTAL PAGU
    # ===============================
    pagu_col = find(["PAGU", "JUMLAH"])
    if pagu_col is None:
        raise ValueError("Kolom Pagu tidak ditemukan di file OMSPAN")

    out["Total Pagu"] = (
        df[pagu_col]
        .astype(str)
        .str.replace(r"[^\d]", "", regex=True)
        .replace("", "0")
        .astype(float)
    )

    # ===============================
    # 4️⃣ NO DIPA (JIKA ADA)
    # ===============================
    dipa_col = find(["DIPA"])
    out["No Dipa"] = df[dipa_col].astype(str) if dipa_col else ""

    # ===============================
    # 5️⃣ TANGGAL POSTING REVISI
    # ===============================
    tgl_col = find([
        "TANGGAL POSTING",
        "TGL POSTING",
        "TANGGAL REKAM",
        "TGL REKAM",
        "TANGGAL APPROVAL",
        "TGL APPROVAL"
    ])

    if tgl_col:
        out["Tanggal Posting Revisi"] = pd.to_datetime(
            df[tgl_col],
            errors="coerce"
        )
    else:
        # fallback aman → 31 Desember (akan disesuaikan di proses utama)
        out["Tanggal Posting Revisi"] = pd.NaT

    # ===============================
    # 6️⃣ METADATA REVISI
    # ===============================
    out["Revisi ke-"] = 0
    out["Jenis Revisi"] = "ANGKA DASAR"

    # ===============================
    # 7️⃣ OWNER & DIGITAL STAMP
    # ===============================
    out["Owner"] = "SATKER"
    out["Digital Stamp"] = "OMSPAN (NON-SPAN)"

    return out.dropna(subset=["Kode Satker"])

def standardize_ikpa_format(df):
    """
    Standardisasi DataFrame IKPA dari GitHub (format sudah header=row).
    Mendeteksi kolom Kode KPPN, Kode BA, Kode Satker, Uraian Satker,
    dan semua indikator IKPA secara otomatis berdasarkan nama kolom.
    """
    import re as _re

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # =========================
    # FILTER HANYA BARIS "NILAI" (jika ada kolom Keterangan)
    # =========================
    ket_col = next((c for c in df.columns if "KETERANGAN" in str(c).upper()), None)
    if ket_col:
        df = df[df[ket_col].astype(str).str.upper().str.contains("NILAI", na=False)]

    if df.empty:
        return df

    # =========================
    # DETEKSI KODE SATKER (6 digit)
    # =========================
    kode_satker_col = None
    for c in df.columns:
        ratio = df[c].astype(str).str.match(r"^\d{6}$").mean()
        if ratio > 0.4:
            kode_satker_col = c
            break
    if kode_satker_col is None and len(df.columns) >= 5:
        kode_satker_col = df.columns[4]

    # =========================
    # DETEKSI KODE BA (3 digit) & KODE KPPN (3-6 digit)
    # Cari dari nama kolom dulu, fallback dari posisi
    # =========================
    def find_col(*names):
        for c in df.columns:
            cu = str(c).upper()
            for n in names:
                if n in cu:
                    return c
        return None

    kode_ba_col   = find_col("KODE BA", "BA") 
    kode_kppn_col = find_col("KODE KPPN", "KPPN")
    uraian_col    = find_col("URAIAN SATKER", "NAMA SATKER")

    # =========================
    # NORMALISASI KODE SATKER
    # =========================
    df["Kode Satker"] = (
        df[kode_satker_col]
        .astype(str)
        .str.extract(r"(\d{6})")[0]
        .fillna("")
    )
    
    df["Kode Satker"] = df["Kode Satker"].apply(normalize_kode_satker)

    if kode_ba_col and kode_ba_col != "Kode BA":
        df["Kode BA"] = df[kode_ba_col].astype(str).str.strip()
    elif "Kode BA" not in df.columns:
        df["Kode BA"] = ""

    if kode_kppn_col and kode_kppn_col != "Kode KPPN":
        df["Kode KPPN"] = df[kode_kppn_col].astype(str).str.strip()
    elif "Kode KPPN" not in df.columns:
        df["Kode KPPN"] = ""

    if uraian_col and uraian_col != "Uraian Satker":
        df["Uraian Satker"] = df[uraian_col].astype(str)
    elif "Uraian Satker" not in df.columns:
        df["Uraian Satker"] = ""

    # =========================
    # BUANG BARIS TIDAK VALID
    # =========================
    df = df[df["Kode Satker"].ne("") & df["Kode Satker"].ne("000000")]
    df = df.reset_index(drop=True)

    return df


#Normalisasi kode BA
def normalize_kode_ba(x):
    try:
        return str(int(x)).zfill(3)
    except:
        return None

# ===============================
# LOAD DATA REFERENSI BA (GITHUB)
# ===============================
def load_reference_ba():
    url = (
        "https://raw.githubusercontent.com/ameernoor/kinerjakeuangansatkerbta/main/templates/Template_Data_Referensi.xlsx"
    )
    ref = pd.read_excel(url, sheet_name=0, dtype=str)
    ref['Kode BA'] = ref['Kode BA'].apply(normalize_kode_ba)
    ref['Nama BA'] = ref['K/L'].astype(str).str.strip()
    return ref

# ===============================
# MAP KODE BA → NAMA BA
# ===============================
def get_ba_map(df_ref_ba):
    return dict(zip(df_ref_ba['Kode BA'], df_ref_ba['Nama BA']))

    
def apply_filter_ba(df):
    selected_ba = st.session_state.get("filter_ba_main", ["SEMUA BA"])
    if not selected_ba or "SEMUA BA" in selected_ba:
        return df
    return df[df['Kode BA'].astype(str).isin(selected_ba)].copy()

# Konfigurasi halaman
st.set_page_config(
    page_title="Dashboard IKPA KPPN Baturaja",
    page_icon="📊",
    layout="wide"
)

# ===============================
# CSS GLOBAL (DIGABUNG JADI SATU)
# ===============================
st.markdown("""
<style>

/* SIDEBAR MENU BUTTON (lama) */
section[data-testid="stSidebar"] div.stButton > button{
    width:100%;
    background:linear-gradient(135deg,#ffffff,#f1f7ff);
    border:none;
    border-radius:12px;
    text-align:left;
    padding:10px 14px;
    font-size:15px;
    color:#1e293b;
    height:auto;
    font-weight:600;
    box-shadow:0 6px 18px rgba(0,0,0,0.08);
    transition:all 0.25s ease;
}
section[data-testid="stSidebar"] div.stButton > button:hover{
    background:linear-gradient(135deg,#e0ecff,#c7dbff);
    transform:translateX(4px);
}
section[data-testid="stSidebar"] div.stButton{
    margin-bottom:6px;
}

/* LAYOUT */
.block-container{
    padding-top:0rem !important;
}
header[data-testid="stHeader"]{
    height:0px;
}
section.main > div{
    padding-top:0rem !important;
}
div[data-testid="stVerticalBlock"] > div:first-child{
    margin-top:0rem !important;
}

/* HERO */
.hero{
    position:relative;
    border-radius:22px;
    overflow:hidden;
    padding:240px 60px;
    margin-bottom:30px;
}
.hero::before{
    content:"";
    position:absolute;
    inset:0;
    background-image:url("https://raw.githubusercontent.com/ameernoor/kinerjakeuangansatkerbta/main/kppn_backgound.png");
    background-size:cover;
    background-position:center 20%;
    filter:blur(0.5px);
    opacity:0.9;
    transform:scale(1.0);
}
.hero::after{
    content:"";
    position:absolute;
    inset:0;
    background:rgba(255,255,255,0.15);
}
.hero-content{
    position:relative;
    z-index:2;
}
.hero-title{
    font-size:56px;
    font-weight:800;
    color:#1f2937;
}
.hero-sub{
    font-size:26px;
    font-weight:500;
    color:#374151;
}

/* MENU CARD */
.menu-container{
    margin-top:20px;
}
.menu-container div.stButton > button{
    height:150px;
    border-radius:18px;
    border:1px solid #dbeafe;
    background:linear-gradient(135deg,#f5f9ff,#e0ecff);
    color:#1e293b;
    font-size:24px;
    font-weight:700;
    text-align:center;
    display:flex;
    flex-direction:column;
    align-items:center;
    justify-content:center;
    line-height:1.4;
    box-shadow:0 8px 25px rgba(0,0,0,0.06);
    transition:all 0.3s ease;
    animation:fadeUp 0.5s ease;
}
.menu-container div.stButton > button:hover{
    transform:translateY(-6px);
    background:linear-gradient(135deg,#e0ecff,#c7dbff);
    box-shadow:0 18px 45px rgba(37,99,235,0.25);
}
@keyframes fadeUp{
    from{ opacity:0; transform:translateY(15px); }
    to{ opacity:1; transform:translateY(0); }
}

/* SIDEBAR */
section[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#eaf4ff,#dbeafe);
}
section[data-testid="stSidebar"] h1{
    color:#1e3a8a;
}
section[data-testid="stSidebar"] hr{
    border-color:#c7dbff;
}
section[data-testid="stSidebar"] .stAlert{
    background:#e6f0ff;
    border:none;
    border-radius:14px;
    color:#1e3a8a;
}

</style>
""", unsafe_allow_html=True)

# ====================================================================
def extract_kode_from_satker_field(s, width=6):
    """
    Jika kolom 'Satker' mengandung '001234 – NAMA SATKER', ambil angka di awal.
    Jika hanya angka (sebagai int/str), return padded string.
    """
    if pd.isna(s):
        return ''
    stxt = str(s).strip()
    # cari angka di awal baris (atau angka pertama)
    m = re.match(r'^\s*0*\d+', stxt)
    if m:
        return normalize_kode_satker(m.group(0), width=width)
    # fallback: cari first group of digits anywhere
    m2 = re.search(r'(\d+)', stxt)
    if m2:
        return normalize_kode_satker(m2.group(1), width=width)
    return ''

        
def register_ikpa_satker(df_final, month, year, source="Manual"):
    
    # ===============================
    # 🔥 NORMALISASI
    # ===============================
    month = normalize_month(month)
    year = str(year)

    # 🔥 FIX: KEY WAJIB ADA
    key = (month, year)

    df = df_final.copy()

    df["Source"] = source
    df["Period"] = f"{month} {year}"

    MONTH_ORDER = {
        "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
        "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
        "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12
    }

    df["Period_Sort"] = f"{int(year):04d}-{MONTH_ORDER.get(month, 0):02d}"

    # ===============================
    # 🔥 FIX NUMERIC
    # ===============================
    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"

    if nilai_col in df.columns:

        df[nilai_col] = df[nilai_col].apply(clean_numeric)

        df = df.sort_values(nilai_col, ascending=False)

        df["Peringkat"] = (
            df[nilai_col]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

    # ===============================
    # 🔥 SIMPAN
    # ===============================
    st.session_state.data_storage[key] = df
    

def find_header_row_by_keywords(uploaded_file, keywords, max_rows=15):
    """
    Mencari baris header Excel berdasarkan BANYAK keyword kolom
    (contoh: 'Nama KPPN', 'KPPN', 'Nama Kantor', dll)
    """
    import pandas as pd

    uploaded_file.seek(0)
    preview = pd.read_excel(
        uploaded_file,
        header=None,
        nrows=max_rows
    )

    keywords = [k.upper() for k in keywords]

    for i in range(preview.shape[0]):
        row_values = (
            preview.iloc[i]
            .astype(str)
            .str.upper()
            .str.strip()
        )

        if any(
            any(k in cell for cell in row_values)
            for k in keywords
        ):
            return i

    return None


def process_excel_digipay(uploaded_file, upload_year):
    """
    Parser DIGIPAY stabil
    - Tidak pakai break
    - Tidak berhenti di tengah
    - Filter tahun langsung saat parsing
    """

    df_raw = pd.read_excel(uploaded_file, header=None)
    processed_rows = []

    for i in range(len(df_raw)):

        row = df_raw.iloc[i]

        # =============================
        # Ambil Tahun
        # =============================
        try:
            tahun_row = int(row[0])
        except:
            continue

        # Skip jika bukan tahun yang dipilih
        if tahun_row != upload_year:
            continue

        # =============================
        # Ambil Kode Satker
        # =============================
        kode_satker = normalize_kode_satker(str(row[3]).strip())

        # Lewati jika kosong / invalid
        if (
            not kode_satker
            or not kode_satker.isdigit()
            or len(kode_satker) != 6
            or kode_satker == "000000"
        ):
            continue

        # =============================
        # Simpan data valid
        # =============================
        processed_rows.append({
            "Tahun": tahun_row,
            "Kode Satker": kode_satker,
            "Nama Satker": str(row[4]).strip(),
            "Nilai Digipay": row[5],
        })

    df_final = pd.DataFrame(processed_rows)

    return df_final


def process_excel_file(uploaded_file, upload_year):
    """
    Parser IKPA Satker universal — mendeteksi otomatis format:
      • Format LAMA (≤2025): 4 baris per satker (NILAI, BOBOT, NILAI AKHIR, NILAI ASPEK)
        - Kode Satker di col 3, Uraian di col 4, indikator mulai col 6
        - Kualitas Aspek dibaca dari baris ke-4 (nilai_aspek)
      • Format BARU (2026+): 3 baris per satker (NILAI, BOBOT, NILAI AKHIR)
        - Kode Satker di col 4, Kode BA di col 3, Uraian di col 5, Kode KPPN di col 2
        - Semua indikator + Nilai Aspek ada di baris NILAI (col 7–20)
    """
    import pandas as pd
    import re

    df_raw = pd.read_excel(uploaded_file, header=None)

    # ===============================
    # DETEKSI BULAN dari baris header
    # ===============================
    month = "UNKNOWN"
    for ri in range(min(5, len(df_raw))):
        for ci in range(min(5, df_raw.shape[1])):
            cell = str(df_raw.iloc[ri, ci]).upper().strip()
            if cell in VALID_MONTHS:
                month = VALID_MONTHS[cell]
                break
        if month != "UNKNOWN":
            break

    if month == "UNKNOWN":
        try:
            month_text = str(df_raw.iloc[1, 0])
            month_raw = month_text.split(":")[-1].strip().upper()
            month = VALID_MONTHS.get(month_raw, "UNKNOWN")
        except:
            pass

    # ===============================
    # DATA MULAI BARIS KE-5 (index 4)
    # ===============================
    df_data = df_raw.iloc[4:].reset_index(drop=True)

    # ===============================
    # DETEKSI FORMAT: 3-baris vs 4-baris
    # Cek apakah ada baris ke-4 dengan label 'NILAI ASPEK' atau 'ASPEK'
    # dalam 12 baris pertama data
    # ===============================
    is_3row_format = True  # default: format 2026 (3 baris)
    check_limit = min(12, len(df_data))
    for ri in range(check_limit):
        cell = str(df_data.iloc[ri, 6]).upper().strip() if df_data.shape[1] > 6 else ""
        if "ASPEK" in cell and "NILAI" in cell:
            is_3row_format = False
            break

    def safe_num(val):
        """Convert Indonesian decimal format (comma) to float."""
        try:
            s = str(val).strip().replace("%", "")
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
            s = re.sub(r"[^\d.\-]", "", s)
            return float(s) if s not in ("", "-") else 0.0
        except:
            return 0.0

    processed_rows = []

    # ===============================
    # FORMAT BARU 2026: 3 baris per satker
    # Header row 2: NO | Periode | Kode KPPN | Kode BA | Kode Satker | Uraian Satker | Keterangan
    #               | RevisiDIPA(7) | Deviasi(8) | NilaiAspekPerencanaan(9)
    #               | Penyerapan(10) | Kontraktual(11) | Tagihan(12) | UP/TUP(13) | NilaiAspekPelaksanaan(14)
    #               | CapaianOutput(15) | NilaiAspekHasil(16)
    #               | NilaiTotal(17) | KonversBobot(18) | Dispensasi(19) | NilaiAkhir(20)
    # ===============================
    if is_3row_format:
        i = 0
        while i < len(df_data):
            nilai = df_data.iloc[i]

            # Pastikan ini baris NILAI (keterangan col 6)
            keterangan = str(nilai[6]).upper().strip() if df_data.shape[1] > 6 else ""
            if keterangan != "NILAI":
                i += 1
                continue

            # Kode Satker di col 4
            kode_satker_raw = str(nilai[4]).strip()
            kode_satker = re.sub(r"\D", "", kode_satker_raw).zfill(6)

            if not kode_satker or kode_satker == "000000":
                i += 3
                continue

            kode_ba   = re.sub(r"\D", "", str(nilai[3]).strip()).zfill(3)
            kode_kppn = re.sub(r"\D", "", str(nilai[2]).strip())
            uraian    = str(nilai[5]).strip()

            # Bulan dari baris BOBOT (col 1)
            if i + 1 < len(df_data):
                bulan_cell = str(df_data.iloc[i + 1, 1]).strip().upper()
                month_detected = VALID_MONTHS.get(bulan_cell, month)
            else:
                month_detected = month

            row = {
                "No": nilai[0],
                "Kode KPPN": kode_kppn,
                "Kode BA":   kode_ba,
                "Kode Satker": kode_satker,
                "Uraian Satker": uraian,

                # Nilai Aspek (sudah ada di baris NILAI untuk format 2026)
                "Kualitas Perencanaan Anggaran":        safe_num(nilai[9]),
                "Kualitas Pelaksanaan Anggaran":         safe_num(nilai[14]),
                "Kualitas Hasil Pelaksanaan Anggaran":   safe_num(nilai[16]),

                # Detail indikator komponen
                "Revisi DIPA":                safe_num(nilai[7]),
                "Deviasi Halaman III DIPA":   safe_num(nilai[8]),
                "Penyerapan Anggaran":         safe_num(nilai[10]),
                "Belanja Kontraktual":         safe_num(nilai[11]),
                "Penyelesaian Tagihan":        safe_num(nilai[12]),
                "Pengelolaan UP dan TUP":      safe_num(nilai[13]),
                "Capaian Output":              safe_num(nilai[15]),

                "Nilai Total":                            safe_num(nilai[17]),
                "Konversi Bobot":                         safe_num(nilai[18]),
                "Dispensasi SPM (Pengurangan)":           safe_num(nilai[19]),
                "Nilai Akhir (Nilai Total/Konversi Bobot)": safe_num(nilai[20]),

                "Bulan": month_detected,
                "Tahun": upload_year,
            }

            processed_rows.append(row)
            i += 3  # lompat ke blok berikutnya

    # ===============================
    # FORMAT LAMA (≤2025): 4 baris per satker
    # col 1=Kode KPPN, 2=Kode BA, 3=Kode Satker, 4=Uraian Satker
    # col 6=RevisiDIPA, 7=Deviasi, 8=Penyerapan, 9=Kontraktual,
    #       10=Tagihan, 11=UP/TUP, 12=CapaianOutput
    # col 13=NilaiTotal, 14=KonversBobot, 15=Dispensasi, 16=NilaiAkhir
    # Baris ke-4 (nilai_aspek): col 6=KualitasPerencanaan, 8=KualitasPelaksanaan, 12=KualitasHasil
    # ===============================
    else:
        i = 0
        while i + 3 < len(df_data):
            nilai       = df_data.iloc[i]
            bobot       = df_data.iloc[i + 1]
            nilai_akhir_row = df_data.iloc[i + 2]
            nilai_aspek = df_data.iloc[i + 3]

            kode_satker_raw = str(nilai[3])
            kode_satker = re.sub(r"\D", "", kode_satker_raw).zfill(6)
            uraian_satker = str(nilai[4]).strip()

            if not kode_satker or kode_satker == "000000":
                i += 4
                continue

            row = {
                "No": nilai[0],
                "Kode KPPN": str(nilai[1]).strip("'"),
                "Kode BA":   str(nilai[2]).strip("'"),
                "Kode Satker": kode_satker,
                "Uraian Satker": uraian_satker,

                # Kualitas dari baris nilai_aspek (format lama)
                "Kualitas Perencanaan Anggaran":        safe_num(nilai_aspek[6]),
                "Kualitas Pelaksanaan Anggaran":         safe_num(nilai_aspek[8]),
                "Kualitas Hasil Pelaksanaan Anggaran":   safe_num(nilai_aspek[12]),

                "Revisi DIPA":                safe_num(nilai[6]),
                "Deviasi Halaman III DIPA":   safe_num(nilai[7]),
                "Penyerapan Anggaran":         safe_num(nilai[8]),
                "Belanja Kontraktual":         safe_num(nilai[9]),
                "Penyelesaian Tagihan":        safe_num(nilai[10]),
                "Pengelolaan UP dan TUP":      safe_num(nilai[11]),
                "Capaian Output":              safe_num(nilai[12]),

                "Nilai Total":                            safe_num(nilai[13]),
                "Konversi Bobot":                         safe_num(nilai[14]),
                "Dispensasi SPM (Pengurangan)":           safe_num(nilai[15]),
                "Nilai Akhir (Nilai Total/Konversi Bobot)": safe_num(nilai[16]),

                "Bulan": month,
                "Tahun": upload_year,
            }

            processed_rows.append(row)
            i += 4

    df_final = pd.DataFrame(processed_rows)

    if df_final.empty:
        return None, "UNKNOWN", upload_year

    return df_final, month, upload_year
    

VALID_MONTHS = {
    "JANUARI": "JANUARI",
    "FEBRUARI": "FEBRUARI",
    "PEBRUARI": "FEBRUARI",
    "MARET": "MARET",
    "APRIL": "APRIL",
    "MEI": "MEI",
    "JUNI": "JUNI",
    "JULI": "JULI",
    "JULY": "JULI",
    "AGUSTUS": "AGUSTUS",
    "AGUSTUSS": "AGUSTUS",
    "SEPTEMBER": "SEPTEMBER",
    "SEPT": "SEPTEMBER",
    "OKTOBER": "OKTOBER",
    "NOVEMBER": "NOVEMBER",
    "NOPEMBER": "NOVEMBER",
    "DESEMBER": "DESEMBER",
}


def post_process_ikpa_satker(df, source="Upload"):
    df = df.copy()

    # =========================
    # 🔥 VALIDASI AWAL
    # =========================
    if df is None or df.empty:
        return df

    # =========================
    # 🔥 KOLOM TEKS YANG TIDAK BOLEH DI-NUMERIK-KAN
    # =========================
    NON_NUMERIC_COLS = {
        "Kode Satker", "Uraian Satker", "Uraian Satker-RINGKAS",
        "Bulan", "Tahun", "Kode BA", "Kode KPPN", "Source", "Period",
        "Period_Sort", "Satker", "Jenis Satker"
    }

    # =========================
    # 🔥 FIX TOTAL NUMERIK (JANGAN SENTUH KOLOM TEKS)
    # =========================
    for col in df.columns:
        if col in NON_NUMERIC_COLS:
            continue
        try:
            df[col] = (
                df[col]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace(r"[^\d.\-]", "", regex=True)
            )
            df[col] = pd.to_numeric(df[col], errors="ignore")
        except:
            pass

    # =========================
    # 🔥 0. FIX TAHUN & BULAN
    # =========================
    if "Tahun" in df.columns:
        df["Tahun"] = pd.to_numeric(df["Tahun"], errors="coerce").fillna(0).astype(int)

    if "Bulan" in df.columns:
        df["Bulan"] = df["Bulan"].apply(normalize_month)

    # =========================
    # 🔥 1. NORMALISASI KODE SATKER
    # =========================
    if "Kode Satker" in df.columns:
        df["Kode Satker"] = (
            df["Kode Satker"]
            .astype(str)
            .str.extract(r"(\d+)")[0]
            .fillna("")
            .str.zfill(6)
        )

    # =========================
    # 🔥 FILTER DATA SAMPAH
    # =========================
    if "Kode Satker" in df.columns:
        df = df[
            (df["Kode Satker"] != "") &
            (df["Kode Satker"] != "000000")
        ]

    # 🔥 FIX: Hanya filter baris null jika kolom Uraian Satker ada
    # DAN tidak semua nilainya null (mencegah data hilang total)
    if "Uraian Satker" in df.columns:
        non_null_mask = df["Uraian Satker"].notna()
        # Hanya filter jika ada minimal beberapa baris valid — jangan kosongkan semua data
        if non_null_mask.sum() > 0:
            df = df[non_null_mask]

    df = df.reset_index(drop=True)


    # =========================
    # 🔥 FIX NILAI IKPA
    # =========================
    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"

    if nilai_col in df.columns:
        df[nilai_col] = pd.to_numeric(df[nilai_col], errors="coerce").fillna(0)
    else:
        st.warning("⚠️ Kolom nilai IKPA tidak ditemukan")
        df[nilai_col] = 0

    # =========================
    # 🔥 RANKING (ANTI CRASH)
    # =========================
    try:
        df = df.sort_values(nilai_col, ascending=False)

        df["Peringkat"] = (
            df[nilai_col]
            .rank(method="dense", ascending=False)
            .fillna(0)
            .astype(int)
        )
    except Exception as e:
        st.warning(f"⚠️ Ranking gagal: {e}")
        df["Peringkat"] = 0

    # =========================
    # 🔥 METADATA
    # =========================
    df["Source"] = source

    if "Bulan" in df.columns and "Tahun" in df.columns:
        df["Period"] = df["Bulan"] + " " + df["Tahun"].astype(str)
    else:
        df["Period"] = "UNKNOWN"

    MONTH_ORDER_LOCAL = {
        "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
        "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
        "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12
    }

    if "Bulan" in df.columns and "Tahun" in df.columns:
        df["Period_Sort"] = (
            df["Tahun"].astype(int).astype(str)
            + "-"
            + df["Bulan"].map(MONTH_ORDER_LOCAL).fillna(0).astype(int).astype(str).str.zfill(2)
        )
    else:
        df["Period_Sort"] = "0000-00"

    # =========================
    # 🔥 MERGE DIPA — hanya jika DATA_DIPA_by_year sudah tersedia
    # =========================
    try:
        dipa_dict = st.session_state.get("DATA_DIPA_by_year", {})
        if dipa_dict:
            df = merge_ikpa_with_dipa(df)
        else:
            df["Total Pagu"] = 0
    except KeyError as e:
        # 'Total Pagu' tidak ditemukan di DIPA → lanjut tanpa merge
        df["Total Pagu"] = 0
    except Exception as e:
        st.warning(f"⚠️ Merge DIPA gagal: {e}")
        df["Total Pagu"] = 0

    # =========================
    # 🔥 KLASIFIKASI
    # =========================
    try:
        df = classify_jenis_satker(df)
    except Exception:
        df["Jenis Satker"] = "SEDANG"

    # =========================
    # 🔥 FINAL KOLOM
    # =========================
    FINAL_COLUMNS = [
        "No", "Kode KPPN", "Kode BA", "Kode Satker", "Uraian Satker",
        "Kualitas Perencanaan Anggaran",
        "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",
        "Revisi DIPA", "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran", "Belanja Kontraktual",
        "Penyelesaian Tagihan", "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total", "Konversi Bobot",
        "Dispensasi SPM (Pengurangan)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
        "Bulan", "Tahun", "Peringkat",
        "Uraian Satker Final", "Satker", "Source",
        "Uraian Satker-RINGKAS",
        "Period", "Period_Sort", "Total Pagu", "Jenis Satker"
    ]

    df = df[[c for c in FINAL_COLUMNS if c in df.columns]]

    # =========================
    # 🔥 FINAL TOUCH
    # =========================
    df = df.fillna({
        "Satker": "TIDAK DIKETAHUI",
        "Total Pagu": 0
    })

    try:
        df = apply_reference_short_names(df)
        df = create_satker_column(df)
    except:
        pass

    return df


# ===============================
# PARSER IKPA KPPN (RINGKAS)
# ===============================
def process_kppn_ringkas(uploaded_file, year, detected_month):
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file)

    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )

    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
    if nilai_col not in df.columns:
        raise ValueError("File IKPA KPPN tidak valid")

    # =========================
    # NORMALISASI DESIMAL
    # =========================
    df = df.applymap(
        lambda x: str(x).replace(",", ".") if isinstance(x, str) else x
    )

    for col in df.columns:
        if col not in ["Nama KPPN", "Bulan", "Tahun", "Source"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # =========================
    # NORMALISASI BULAN
    # =========================
    detected_month = (
        str(detected_month)
        .upper()
    )

    detected_month = VALID_MONTHS.get(detected_month, "JULI")

    df["Bulan"] = detected_month
    df["Tahun"] = year
    df["Source"] = "Upload"

    # =========================
    # DENSE RANKING
    # =========================
    df = df.sort_values(nilai_col, ascending=False)

    df["Peringkat"] = (
        df[nilai_col]
        .rank(method="dense", ascending=False)
        .astype(int)
    )

    return df, detected_month, year


# ===============================
# REPROCESS ALL IKPA SATKER (FINAL)
# ===============================
def reprocess_all_ikpa_satker():
    with st.spinner("🔄 Memproses ulang seluruh IKPA Satker..."):

        # ===============================
        # LOAD DATA DARI GITHUB
        # ===============================
        load_data_from_github()

        # ===============================
        # 🔥 PROSES SEMUA DATA (AUTO MULTI-YEAR)
        # ===============================
        for key in list(st.session_state.data_storage.keys()):
            df = st.session_state.data_storage[key]

            if df is None or df.empty:
                continue

            # ===============================
            # 🔥 VALIDASI TAHUN
            # ===============================
            if "Tahun" not in df.columns:
                continue

            df["Tahun"] = pd.to_numeric(df["Tahun"], errors="coerce").fillna(0).astype(int)

            # ===============================
            # NORMALISASI BULAN & PERIOD
            # ===============================
            if "Bulan" in df.columns:
                df["Bulan"] = df["Bulan"].astype(str).str.upper().str.strip()
                df["Period"] = df["Bulan"] + " " + df["Tahun"].astype(str)

            # ===============================
            # NORMALISASI KODE SATKER
            # ===============================
            if "Kode Satker" in df.columns:
                df["Kode Satker"] = (
                    df["Kode Satker"]
                    .astype(str)
                    .str.extract(r"(\d+)")[0]
                    .str.zfill(6)
                )

            # ===============================
            # 🔥 MERGE SESUAI TAHUN MASING-MASING
            # ===============================
            try:
                df = merge_ikpa_with_dipa(df)
            except Exception as e:
                st.warning(f"⚠️ Merge gagal untuk {key}: {e}")
                df["Total Pagu"] = 0

            # ===============================
            # SIMPAN KEMBALI
            # ===============================
            st.session_state.data_storage[key] = df

        # ===============================
        # FLAG SELESAI
        # ===============================
        st.session_state.ikpa_dipa_merged = True

        st.success("✅ Reprocess IKPA Satker selesai (multi-year siap digunakan)")




import pandas as pd
import re


# ============================================================
# VALID MONTHS MAP
# ============================================================
VALID_MONTHS = {
    "JANUARI": "JANUARI", "JAN": "JANUARI",
    "FEBRUARI": "FEBRUARI", "FEB": "FEBRUARI", "PEBRUARI": "FEBRUARI",
    "MARET": "MARET", "MAR": "MARET",
    "APRIL": "APRIL", "APR": "APRIL",
    "MEI": "MEI",
    "JUNI": "JUNI", "JUN": "JUNI",
    "JULI": "JULI", "JUL": "JULI",
    "AGUSTUS": "AGUSTUS", "AGT": "AGUSTUS", "AGS": "AGUSTUS",
    "SEPTEMBER": "SEPTEMBER", "SEP": "SEPTEMBER", "SEPT": "SEPTEMBER",
    "OKTOBER": "OKTOBER", "OKT": "OKTOBER",
    "NOVEMBER": "NOVEMBER", "NOV": "NOVEMBER", "NOPEMBER": "NOVEMBER",
    "DESEMBER": "DESEMBER", "DES": "DESEMBER",
}

MONTH_NUM = {
    "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
    "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
    "SEPTEMBER": 9, "OKTOBER": 10, "NOVEMBER": 11, "DESEMBER": 12,
}


def _to_float(v):
    """Konversi nilai ke float, mendukung format Indonesia (koma sebagai desimal)."""
    try:
        s = str(v).strip().replace('%', '').strip()
        # Format Indonesia: '99,71' atau '1.234,56'
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        s = re.sub(r'[^\d.\-]', '', s)
        return float(s) if s not in ('', '-') else 0.0
    except Exception:
        return 0.0


def _detect_month_from_raw(df_raw):
    """Cari nama bulan di 6 baris pertama, 5 kolom pertama."""
    for r in range(min(6, df_raw.shape[0])):
        for c in range(min(5, df_raw.shape[1])):
            cell = str(df_raw.iloc[r, c]).upper().strip()
            if cell in VALID_MONTHS:
                return VALID_MONTHS[cell]
            for k, v in VALID_MONTHS.items():
                if len(k) >= 4 and k in cell:
                    return v
    return None


def _detect_month_from_filename(filename):
    """Cari nama bulan di nama file."""
    fname = str(filename).upper()
    for k, v in VALID_MONTHS.items():
        if len(k) >= 4 and k in fname:
            return v
    return None


def _finalize(rows, detected_month, year, source="Upload"):
    """
    Dari list of dict hasil parsing, buat DataFrame akhir yang seragam.
    Lakukan ranking dan tambahkan metadata.
    """
    if not rows:
        return None

    df = pd.DataFrame(rows)

    # Pastikan semua kolom numerik bersih
    NUMERIC_COLS = [
        "Kualitas Perencanaan Anggaran", "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",
        "Revisi DIPA", "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran", "Belanja Kontraktual",
        "Penyelesaian Tagihan", "Pengelolaan UP dan TUP",
        "Capaian Output", "Nilai Total", "Konversi Bobot",
        "Dispensasi SPM (Pengurang)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
    ]
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Metadata
    df["Bulan"] = detected_month or "UNKNOWN"
    df["Tahun"] = int(year)
    df["Source"] = source

    # Ranking
    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
    if nilai_col in df.columns:
        df = df.sort_values(nilai_col, ascending=False).reset_index(drop=True)
        df["Peringkat"] = (
            df[nilai_col]
            .rank(method="dense", ascending=False)
            .astype(int)
        )
    else:
        df["Peringkat"] = 0

    # Nomor urut
    if "No" not in df.columns:
        df.insert(0, "No", range(1, len(df) + 1))
    else:
        df["No"] = range(1, len(df) + 1)

    # Urutan kolom output final
    FINAL_COLS = [
        "No", "Kode KPPN", "Nama KPPN",
        "Kualitas Perencanaan Anggaran",
        "Kualitas Pelaksanaan Anggaran",
        "Kualitas Hasil Pelaksanaan Anggaran",
        "Revisi DIPA", "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran", "Belanja Kontraktual",
        "Penyelesaian Tagihan", "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total", "Konversi Bobot",
        "Dispensasi SPM (Pengurang)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
        "Bulan", "Tahun", "Source", "Peringkat",
    ]
    df = df.reindex(columns=FINAL_COLS, fill_value=0)

    return df


# ============================================================
# MAIN FUNCTION
# ============================================================
def process_excel_file_kppn(uploaded_file, year, detected_month=None):
    """
    Parser IKPA KPPN universal — mendeteksi otomatis Format A, B, atau C.

    Parameters
    ----------
    uploaded_file : file-like object  (Streamlit UploadedFile / BytesIO)
    year          : int   — tahun yang dipilih user
    detected_month: str   — nama bulan yang sudah dideteksi dari luar (opsional)

    Returns
    -------
    (df_out, bulan_str, year_int)  atau  (None, detected_month, year) jika gagal
    """
    try:
        uploaded_file.seek(0)
        df_raw = pd.read_excel(uploaded_file, header=None).fillna("")

        # ─── Deteksi bulan ───────────────────────────────────────────────
        month = detected_month
        if not month:
            month = _detect_month_from_raw(df_raw)
        if not month:
            try:
                month = _detect_month_from_filename(uploaded_file.name)
            except Exception:
                pass
        if not month:
            month = "UNKNOWN"
        month = VALID_MONTHS.get(str(month).upper().strip(), month)

        # ─── Cek apakah file sudah berformat FLAT (Format A) ────────────
        first_row_str = " ".join(df_raw.iloc[0].astype(str).str.upper())
        is_flat = (
            ("KODE KPPN" in first_row_str or "NAMA KPPN" in first_row_str)
            and "KODE SATKER" not in first_row_str
            and "SATKER" not in first_row_str
        )

        if is_flat:
            return _parse_format_a(uploaded_file, year, month)

        # ─── Deteksi Format B vs C berdasarkan jumlah kolom ─────────────
        # Cari baris data pertama (bukan header)
        n_cols = df_raw.shape[1]

        # Format B = 19 kolom, Format C = 15 kolom
        # Gunakan baris header (row 2/3) untuk konfirmasi
        if n_cols >= 18:
            return _parse_format_b(df_raw, year, month)
        else:
            return _parse_format_c(df_raw, year, month)

    except Exception as e:
        import traceback
        traceback.print_exc()
        # st.error(f"❌ ERROR IKPA KPPN: {e}")   # uncomment jika dalam Streamlit
        print(f"ERROR: {e}")
        return None, detected_month, year


# ============================================================
# FORMAT A — FLAT (sudah diproses)
# ============================================================
def _parse_format_a(uploaded_file, year, month):
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.astype(str).str.strip()

    # Normalisasi koma → titik
    df = df.applymap(
        lambda x: str(x).replace(',', '.') if isinstance(x, str) else x
    )

    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
    if nilai_col not in df.columns:
        # coba alias
        for alias in ["Nilai Akhir", "Nilai Akhir (Nilai Total / Konversi Bobot)"]:
            if alias in df.columns:
                df = df.rename(columns={alias: nilai_col})
                break

    # Normalisasi kolom Dispensasi
    for old in ["Dispensasi SPM (Pengurangan)", "Dispensasi SPM Pengurangan"]:
        if old in df.columns:
            df = df.rename(columns={old: "Dispensasi SPM (Pengurang)"})

    # Normalisasi numerik
    skip = {"Kode KPPN", "Nama KPPN", "Bulan", "Tahun", "Source"}
    for col in df.columns:
        if col not in skip:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df["Bulan"] = month
    df["Tahun"] = int(year)
    df["Source"] = "Upload"

    if nilai_col in df.columns:
        df = df.sort_values(nilai_col, ascending=False).reset_index(drop=True)
        df["Peringkat"] = (
            df[nilai_col].rank(method="dense", ascending=False).astype(int)
        )

    return df, month, year


# ============================================================
# FORMAT B — OM-SPAN Modern / MyIntress (≥ 19 kolom, 3 baris/KPPN)
# ============================================================
def _parse_format_b(df_raw, year, month):
    """
    Header ada di baris 2 (row index 2) dan baris 3 (row index 3).
    Data NILAI mulai dari baris 4 (row index 4).
    Struktur per KPPN: NILAI / BOBOT / NILAI AKHIR  (3 baris)

    Pemetaan kolom baris NILAI (index 0-based):
      0  = NO
      1  = Periode (angka bulan)
      2  = Kode KPPN
      3  = Nama KPPN
      4  = Keterangan ('NILAI')
      5  = Revisi DIPA
      6  = Deviasi Halaman III DIPA
      7  = Nilai Aspek Perencanaan
      8  = Penyerapan Anggaran
      9  = Belanja Kontraktual
      10 = Penyelesaian Tagihan
      11 = Pengelolaan UP dan TUP
      12 = Nilai Aspek Pelaksanaan
      13 = Capaian Output
      14 = Nilai Aspek Hasil
      15 = Nilai Total
      16 = Konversi Bobot (misal '100,00%')
      17 = Dispensasi SPM
      18 = Nilai Akhir
    """
    # Cari baris data (pertama yang punya kode numerik di col 2)
    data_start = None
    for i in range(len(df_raw)):
        cell = str(df_raw.iloc[i, 4]).strip().upper()
        if cell in ("NILAI", "NILAI AKHIR"):
            data_start = i
            break

    if data_start is None:
        return None, month, year

    rows_out = []
    i = data_start

    while i < len(df_raw):
        row_nilai = df_raw.iloc[i]
        ket = str(row_nilai.iloc[4]).strip().upper()

        if ket not in ("NILAI", "NILAI AKHIR"):
            i += 1
            continue

        # Ambil nama bulan dari baris BOBOT (row i+1, col 1)
        if i + 1 < len(df_raw):
            bulan_cell = str(df_raw.iloc[i + 1].iloc[1]).strip().upper()
            month_detected = VALID_MONTHS.get(bulan_cell, month)
        else:
            month_detected = month

        kode_kppn = str(row_nilai.iloc[2]).strip()
        nama_kppn = str(row_nilai.iloc[3]).strip()

        rows_out.append({
            "Kode KPPN":                            kode_kppn,
            "Nama KPPN":                            nama_kppn,
            "Revisi DIPA":                          _to_float(row_nilai.iloc[5]),
            "Deviasi Halaman III DIPA":             _to_float(row_nilai.iloc[6]),
            "Kualitas Perencanaan Anggaran":        _to_float(row_nilai.iloc[7]),
            "Penyerapan Anggaran":                  _to_float(row_nilai.iloc[8]),
            "Belanja Kontraktual":                  _to_float(row_nilai.iloc[9]),
            "Penyelesaian Tagihan":                 _to_float(row_nilai.iloc[10]),
            "Pengelolaan UP dan TUP":               _to_float(row_nilai.iloc[11]),
            "Kualitas Pelaksanaan Anggaran":        _to_float(row_nilai.iloc[12]),
            "Capaian Output":                       _to_float(row_nilai.iloc[13]),
            "Kualitas Hasil Pelaksanaan Anggaran":  _to_float(row_nilai.iloc[14]),
            "Nilai Total":                          _to_float(row_nilai.iloc[15]),
            "Konversi Bobot":                       _to_float(row_nilai.iloc[16]),
            "Dispensasi SPM (Pengurang)":           _to_float(row_nilai.iloc[17]),
            "Nilai Akhir (Nilai Total/Konversi Bobot)": _to_float(row_nilai.iloc[18]),
        })

        # Perbarui bulan yang terdeteksi (dari baris terkini)
        month = month_detected

        i += 3  # loncat ke blok berikutnya

    return _finalize(rows_out, month, year), month, year


# ============================================================
# FORMAT C — OM-SPAN Lama (15 kolom, 4 baris/KPPN)
# ============================================================
def _parse_format_c(df_raw, year, month):
    """
    Header ada di baris 2 (row index 2) dan baris 3 (row index 3).
    Data mulai baris 4 (row index 4).
    Struktur per KPPN: NILAI / BOBOT / NILAI AKHIR / NILAI ASPEK  (4 baris)

    Pemetaan kolom baris NILAI (index 0-based):
      0  = No
      1  = Kode KPPN
      2  = Nama KPPN
      3  = Keterangan ('Nilai')
      4  = Revisi DIPA
      5  = Deviasi Halaman III DIPA
      6  = Penyerapan Anggaran
      7  = Belanja Kontraktual
      8  = Penyelesaian Tagihan
      9  = Pengelolaan UP dan TUP
      10 = Capaian Output
      11 = Nilai Total
      12 = Konversi Bobot (misal '100%')
      13 = Dispensasi SPM (Pengurang)
      14 = Nilai Akhir

    Pemetaan kolom baris NILAI ASPEK (index 0-based):
      4  = Kualitas Perencanaan Anggaran
      6  = Kualitas Pelaksanaan Anggaran
      10 = Kualitas Hasil Pelaksanaan Anggaran
    """
    # Cari baris data pertama
    data_start = None
    for i in range(len(df_raw)):
        cell = str(df_raw.iloc[i, 3]).strip().upper()
        if cell in ("NILAI", "NILAI AKHIR"):
            data_start = i
            break

    if data_start is None:
        return None, month, year

    rows_out = []
    i = data_start

    while i + 3 < len(df_raw):
        row_nilai  = df_raw.iloc[i]
        row_aspek  = df_raw.iloc[i + 3]

        ket = str(row_nilai.iloc[3]).strip().upper()
        if ket not in ("NILAI",):
            # Mungkin baris BOBOT/NILAI AKHIR/NILAI ASPEK → skip 1
            i += 1
            continue

        kode_kppn = str(row_nilai.iloc[1]).strip()
        nama_kppn = str(row_nilai.iloc[2]).strip()

        # Skip baris invalid (kode bukan numerik)
        if not re.search(r'\d', kode_kppn):
            i += 4
            continue

        rows_out.append({
            "Kode KPPN":                            kode_kppn,
            "Nama KPPN":                            nama_kppn,
            "Revisi DIPA":                          _to_float(row_nilai.iloc[4]),
            "Deviasi Halaman III DIPA":             _to_float(row_nilai.iloc[5]),
            "Penyerapan Anggaran":                  _to_float(row_nilai.iloc[6]),
            "Belanja Kontraktual":                  _to_float(row_nilai.iloc[7]),
            "Penyelesaian Tagihan":                 _to_float(row_nilai.iloc[8]),
            "Pengelolaan UP dan TUP":               _to_float(row_nilai.iloc[9]),
            "Capaian Output":                       _to_float(row_nilai.iloc[10]),
            "Nilai Total":                          _to_float(row_nilai.iloc[11]),
            "Konversi Bobot":                       _to_float(row_nilai.iloc[12]),
            "Dispensasi SPM (Pengurang)":           _to_float(row_nilai.iloc[13]),
            "Nilai Akhir (Nilai Total/Konversi Bobot)": _to_float(row_nilai.iloc[14]),
            # Aspek dari baris NILAI ASPEK (baris ke-4)
            "Kualitas Perencanaan Anggaran":        _to_float(row_aspek.iloc[4]),
            "Kualitas Pelaksanaan Anggaran":        _to_float(row_aspek.iloc[6]),
            "Kualitas Hasil Pelaksanaan Anggaran":  _to_float(row_aspek.iloc[10]),
        })

        i += 4  # loncat ke blok berikutnya

    return _finalize(rows_out, month, year), month, year


def process_kppn_flat(df):
    # ===============================
    # 1. AMBIL HEADER BARIS KE-2
    # ===============================
    header2 = df.iloc[0]

    new_cols = []
    for col, sub in zip(df.columns, header2):
        if pd.notna(sub):
            new_cols.append(str(sub).strip())
        else:
            new_cols.append(str(col).strip())

    df.columns = new_cols

    # ===============================
    # 2. BUANG BARIS HEADER
    # ===============================
    df = df.iloc[1:].reset_index(drop=True)

    # ===============================
    # 3. FILTER HANYA "NILAI" (SAFE)
    # ===============================
    if "Keterangan" in df.columns:
        df["Keterangan"] = df["Keterangan"].astype(str)
        df = df[df["Keterangan"].str.upper().str.contains("NILAI", na=False)]
    else:
        # format baru → tidak ada kolom Keterangan
        pass

    # ===============================
    # 4. DROP KOLOM TIDAK PERLU
    # ===============================
    df = df.loc[:, ~df.columns.str.contains("Unnamed", case=False)]

    # ===============================
    # 5. CLEAN KOLOM
    # ===============================
    df.columns = (
        df.columns
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )

    # ===============================
    # 6. HAPUS BARIS KOSONG / ZERO
    # ===============================
    df = df.dropna(how="all")

    # optional: buang baris yang semua angka = 0
    df = df[~(df.fillna(0) == 0).all(axis=1)]

    # ===============================
    # 7. RESET INDEX
    # ===============================
    df = df.reset_index(drop=True)

    return df



# ============================================================
# PARSER DIPA (FINAL CLEAN VERSION)
# ============================================================
def load_DATA_DIPA_from_github():
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.error("❌ GitHub token / repo tidak ditemukan.")
        return False

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
    except Exception as e:
        st.error(f"❌ Gagal koneksi GitHub: {e}")
        return False

    try:
        files = repo.get_contents("DATA_DIPA")
    except Exception as e:
        st.error(f"❌ Folder DATA_DIPA tidak ditemukan: {e}")
        return False

    pattern = re.compile(r"^DIPA[_-]?(\d{4})\.xlsx$", re.IGNORECASE)

    # 🔥 FIX PENTING (WAJIB ADA)
    loaded_years = []

    if "DATA_DIPA_by_year" not in st.session_state:
        st.session_state.DATA_DIPA_by_year = {}

    for f in files:


        match = pattern.match(f.name)
        if not match:
            continue

        tahun = int(match.group(1))

        try:
            # ===============================
            # LOAD FILE (header=None — parse_dipa yang handle header)
            # ===============================
            raw = base64.b64decode(f.content)
            df_raw = pd.read_excel(io.BytesIO(raw), header=None)

            if df_raw is None or df_raw.empty:
                continue

            # ===============================
            # PARSE — pakai parse_dipa (terbukti benar dari versi lama)
            # ===============================
            df_parsed = parse_dipa(df_raw)

            if df_parsed is None or df_parsed.empty:
                continue

            # ===============================
            # NORMALISASI TOTAL PAGU (WAJIB)
            # ===============================
            df_parsed["Total Pagu"] = pd.to_numeric(
                df_parsed["Total Pagu"], errors="coerce"
            ).fillna(0)

            # SET TAHUN
            df_parsed["Tahun"] = tahun

            # NORMALISASI KODE SATKER
            df_parsed["Kode Satker"] = (
                df_parsed["Kode Satker"]
                .astype(str)
                .str.extract(r"(\d+)")[0]
                .fillna("")
                .str.zfill(6)
            )

            # ===============================
            # VALIDASI: jika Total Pagu semua 0 → simpan tapi tidak dihitung loaded
            # ===============================
            if df_parsed["Total Pagu"].sum() == 0:
                st.session_state.DATA_DIPA_by_year[tahun] = df_parsed
                continue

            # SIMPAN
            st.session_state.DATA_DIPA_by_year[tahun] = df_parsed
            loaded_years.append(str(tahun))

        except Exception as e:
            st.error(f"Gagal load file {f.name}: {e}")

    # ===============================
    # FINAL RESULT
    # ===============================
    if loaded_years:
        add_notification("Data DIPA berhasil dimuat: " + ", ".join(loaded_years))

    return True


# ===============================
# HELPER EXPORT EXCEL
# ===============================
def to_excel_bytes(df):
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    df = df.copy()

    if "Kode Satker" in df.columns:
        df["Kode Satker"] = (
            df["Kode Satker"]
            .astype(str)
            .str.zfill(6)
        )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data IKPA")

        worksheet = writer.sheets["Data IKPA"]

        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == "Kode Satker":
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    worksheet[f"{col_letter}{row}"].number_format = "@"

    output.seek(0)
    return output.getvalue()

# ============================================================
# LOAD TEMPLATE REFERENSI (TEMPLATES FOLDER SAJA)
# ============================================================
def load_template_referensi_from_github():
    token = st.secrets["GITHUB_TOKEN"]
    repo_name = st.secrets["GITHUB_REPO"]

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    file_path = "templates/Template_Data_Referensi.xlsx"
    existing_file = repo.get_contents(file_path)

    file_content = base64.b64decode(existing_file.content)
    df = pd.read_excel(io.BytesIO(file_content), dtype=str)

    return df, repo, existing_file


# ============================================================
# UPDATE TEMPLATE REFERENSI (REPLACE FILE YANG SAMA)
# ============================================================
def update_template_referensi_github(df_updated, repo, existing_file, message):
    file_path = "templates/Template_Data_Referensi.xlsx"

    excel_bytes = io.BytesIO()
    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
        df_updated.to_excel(writer, index=False)

    excel_bytes.seek(0)

    repo.update_file(
        file_path,
        message,
        excel_bytes.getvalue(),
        existing_file.sha
    )


# Save any file (Excel/template) to your GitHub repo
from github.GithubException import GithubException

def save_file_to_github(content_bytes, filename, folder):
    
    from github import Github, Auth
    from github.GithubException import GithubException
    import streamlit as st

    token = st.secrets["GITHUB_TOKEN"]
    repo_name = st.secrets["GITHUB_REPO"]

    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    path = f"{folder}/{filename}"

    try:
        existing = repo.get_contents(path)

        repo.update_file(
            path,
            f"Update {filename}",
            content_bytes,
            existing.sha
        )

        st.success(f"🔄 UPDATE BERHASIL: {path}")

    except GithubException as e:
        if e.status == 404:
            repo.create_file(
                path,
                f"Create {filename}",
                content_bytes
            )
            st.success(f"✅ CREATE BERHASIL: {path}")
        else:
            st.error(f"❌ GitHub ERROR: {e}")
            raise
        

# ============================
#  LOAD DATA IKPA SATKER DARI GITHUB
# ============================
def load_data_from_github(_cache_buster: int = 0):
    
    import base64
    import io
    import re

    data_storage = {}

    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        st.error("❌ GITHUB_TOKEN / GITHUB_REPO tidak ditemukan")
        return data_storage

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
    except Exception as e:
        st.error(f"❌ Gagal koneksi GitHub: {e}")
        return data_storage

    # ===============================
    # 🔥 AMBIL SEMUA FILE (RECURSIVE)
    # ===============================
    def get_all_files(repo, path="data"):
        files = []
        try:
            contents = repo.get_contents(path)
        except Exception:
            return files

        for item in contents:
            if item.type == "dir":
                files.extend(get_all_files(repo, item.path))
            else:
                files.append(item)

        return files

    contents = get_all_files(repo, "data")


    if not contents:
        st.warning("⚠️ Folder 'data' kosong di GitHub")
        return data_storage

    MONTH_ORDER = {
        "JANUARI": 1, "FEBRUARI": 2, "MARET": 3, "APRIL": 4,
        "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
        "SEPTEMBER": 9, "OKTOBER": 10,
        "NOVEMBER": 11, "NOPEMBER": 11,
        "DESEMBER": 12
    }

    # ===============================
    # LOOP FILE
    # ===============================
    for file in contents:

        if not file.name.endswith(".xlsx"):
            continue

        try:
            decoded = base64.b64decode(file.content)

            df = pd.read_excel(io.BytesIO(decoded))
            df.columns = [str(c).strip() for c in df.columns]

            # 🔥 VALIDASI KOLOM
            if "Kode Satker" not in df.columns:
                st.warning(f"⚠️ Skip file (tidak ada Kode Satker): {file.name}")
                continue

            # ===============================
            # NORMALISASI SATKER
            # ===============================
            df["Kode Satker"] = (
                df["Kode Satker"]
                .astype(str)
                .str.extract(r"(\d+)")[0]
                .fillna("")
                .str.zfill(6)
            )

            df = df[(df["Kode Satker"] != "") & (df["Kode Satker"] != "000000")]

            if df.empty:
                st.warning(f"⚠️ Skip file kosong: {file.name}")
                continue

            # ===============================
            # AMBIL BULAN & TAHUN
            # ===============================
            match = re.search(r"IKPA_(\w+)_(\d{4})", file.name)

            if not match:
                st.warning(f"⚠️ Format nama file salah: {file.name}")
                continue

            month = match.group(1).upper()
            year = match.group(2)

            df["Bulan"] = month
            df["Tahun"] = year

            # ===============================
            # POST PROCESS
            # ===============================
            rows_before = len(df)
            df = post_process_ikpa_satker(df)
            df = normalize_ikpa_columns(df)
            df = ensure_ikpa_columns(df)

            # 🔥 FIX: jika post_process mengosongkan df, coba recover dari data asli
            if df is None or df.empty:
                st.warning(f"⚠️ post_process mengosongkan {file.name} ({rows_before} baris → 0). Coba recover...")
                # baca ulang dan skip post_process yang merusak
                df = pd.read_excel(io.BytesIO(decoded))
                df.columns = [str(c).strip() for c in df.columns]
                if "Kode Satker" in df.columns:
                    df["Kode Satker"] = (
                        df["Kode Satker"].astype(str)
                        .str.extract(r"(\d+)")[0]
                        .fillna("").str.zfill(6)
                    )
                    df = df[(df["Kode Satker"] != "") & (df["Kode Satker"] != "000000")]
                df["Bulan"] = month
                df["Tahun"] = year
                df = normalize_ikpa_columns(df)
                df = ensure_ikpa_columns(df)
                if df.empty:
                    st.error(f"❌ Recovery gagal untuk {file.name}, skip file ini")
                    continue
                st.success(f"✅ Recovery berhasil: {len(df)} baris dari {file.name}")

            # ===============================
            # METADATA
            # ===============================
            month_num = MONTH_ORDER.get(month, 0)

            df["Source"] = "GitHub"
            df["Period"] = f"{month} {year}"
            df["Period_Sort"] = f"{int(year):04d}-{month_num:02d}"

            # ===============================
            # SIMPAN — hanya jika df tidak kosong
            # ===============================
            if df is None or df.empty:
                st.warning(f"⚠️ Skip menyimpan {file.name} karena data kosong")
                continue

            key = (month, year)
            data_storage[key] = df


        except Exception as e:
            st.error(f"❌ Gagal load file {file.name}: {e}")
            

    return data_storage

# load data ikpa kppn
@st.cache_data(show_spinner=False, ttl=300)
@st.cache_data(show_spinner=False, ttl=300)
def load_data_ikpa_kppn_from_github():
    from github import Github, Auth
    import base64, io
    import pandas as pd

    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        return {}

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
    except Exception:
        return {}

    KPPN_PATH = "Data IKPA KPPN"   # ✅ konsisten

    # 🔥 recursive scan semua folder tahun
    def collect_xlsx_files(path):
        all_files = []
        try:
            items = repo.get_contents(path)
        except Exception:
            return all_files

        for item in items:
            if item.type == "dir":
                all_files.extend(collect_xlsx_files(item.path))
            elif item.name.endswith(".xlsx"):
                all_files.append(item)

        return all_files

    xlsx_files = collect_xlsx_files(KPPN_PATH)

    data = {}

    for f in xlsx_files:
        try:
            file_bytes = io.BytesIO(base64.b64decode(f.content))

            header_row = detect_header_row(file_bytes)

            file_bytes.seek(0)
            df = pd.read_excel(file_bytes, header=header_row)
            df = process_kppn_flat(df)

            df.columns = (
                df.columns.astype(str)
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)
            )

            # =========================
            # 🔥 AMBIL TAHUN DARI PATH
            # =========================
            path_parts = f.path.split("/")
            tahun = next((p for p in path_parts if p.isdigit()), str(datetime.now().year))

            # =========================
            # 🔥 AMBIL BULAN DARI NAMA FILE
            # =========================
            name = f.name.upper()
            bulan = "UNKNOWN"

            for m in [
                "JANUARI","FEBRUARI","MARET","APRIL","MEI","JUNI",
                "JULI","AGUSTUS","SEPTEMBER","OKTOBER","NOVEMBER","DESEMBER"
            ]:
                if m in name:
                    bulan = m
                    break

            df["Bulan"] = bulan
            df["Tahun"] = tahun

            key = (bulan, tahun)
            data[key] = df

        except Exception as e:
            pass  # skip file error

    return data


def find_header_row_kkp(uploaded_file, max_rows=10):
    uploaded_file.seek(0)
    preview = pd.read_excel(uploaded_file, header=None, nrows=max_rows)

    for i in range(preview.shape[0]):
        row = preview.iloc[i].astype(str).str.upper()
        if (
            row.str.contains("BA/KL").any()
            and row.str.contains("SATKER").any()
            and row.str.contains("PERIODE").any()
        ):
            return i
    return None

def normalize_kkp_for_dashboard(df):
    df = df.copy()

    # Bulan & Tahun
    df["Bulan"] = df["Periode"].dt.strftime("%Y-%m")
    df["Tahun"] = df["Periode"].dt.year

    # Pastikan string rapi
    for col in ["BA/KL", "Satker", "Nama Pemegang KKP", "Bank Penerbit KKP"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df


# ============================================================
# LOAD DATA KKP FROM GITHUB
# ============================================================
@st.cache_data(show_spinner=False, ttl=300)
def load_kkp_master_from_github():

    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        return pd.DataFrame(), False

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)

        file_path = "data_kkp/KKP_MASTER.xlsx"
        file = repo.get_contents(file_path)

        file_bytes = base64.b64decode(file.content)
        df_master = pd.read_excel(io.BytesIO(file_bytes))

        return df_master, True

    except Exception:
        return pd.DataFrame(), False




# ============================================================
# LOAD DIGIPAY FROM GITHUB
# ============================================================
@st.cache_data(show_spinner=False, ttl=300)
def load_digipay_from_github():
    
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        return pd.DataFrame(), 0

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
        contents = repo.get_contents("data_Digipay")
    except:
        return pd.DataFrame(), 0

    all_df = []
    file_count = 0

    for file in contents:
        if file.name.endswith(".xlsx"):
            try:
                file_content = base64.b64decode(file.content)
                df = pd.read_excel(io.BytesIO(file_content), dtype=str)

                if not df.empty:
                    all_df.append(df)
                    file_count += 1
            except:
                continue

    if all_df:
        return pd.concat(all_df, ignore_index=True), file_count
    return pd.DataFrame(), file_count


# ============================================================
# LOAD CMS FROM GITHUB
# ============================================================
@st.cache_data(show_spinner=False, ttl=300)
def load_cms_from_github():
    
    token = st.secrets.get("GITHUB_TOKEN")
    repo_name = st.secrets.get("GITHUB_REPO")

    if not token or not repo_name:
        return pd.DataFrame(), 0

    try:
        g = Github(auth=Auth.Token(token))
        repo = g.get_repo(repo_name)
        contents = repo.get_contents("data_CMS")
    except Exception:
        return pd.DataFrame(), 0

    all_df = []
    file_count = 0

    for file in contents:
        if file.name.endswith(".xlsx"):
            file_content = base64.b64decode(file.content)
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)

            if not df.empty:
                all_df.append(df)
                file_count += 1

    if all_df:
        return pd.concat(all_df, ignore_index=True), file_count
    return pd.DataFrame(), file_count


# FILE KKP
def process_excel_file_kkp(uploaded_file):
    
    uploaded_file.seek(0)

    # ==========================
    # 1️⃣ BACA TANPA HEADER
    # ==========================
    df_raw = pd.read_excel(uploaded_file, header=None, dtype=str)

    # ==========================
    # 2️⃣ DETEKSI HEADER OTOMATIS
    # ==========================
    header_row = None
    for i in range(min(10, len(df_raw))):
        row_text = " ".join(df_raw.iloc[i].astype(str)).upper()
        if "BA/KL" in row_text and "SATKER" in row_text:
            header_row = i
            break

    if header_row is None:
        return pd.DataFrame()

    # ==========================
    # 3️⃣ SET HEADER MANUAL
    # ==========================
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = df_raw.iloc[header_row]

    df = df.reset_index(drop=True)

    # ==========================
    # 4️⃣ NORMALISASI HEADER
    # ==========================
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.upper()
        .str.replace("\n", " ")
        .str.replace(r"\s+", " ", regex=True)
    )

    # ==========================
    # 5️⃣ EKSTRAK KODE BA
    # ==========================
    if "BA/KL" not in df.columns or "SATKER" not in df.columns:
        return pd.DataFrame()

    df["Kode BA"] = (
        df["BA/KL"]
        .astype(str)
        .str.extract(r"(\d{3})")[0]
        .fillna("")
    )

    df["Kode Satker"] = (
        df["SATKER"]
        .astype(str)
        .str.extract(r"(\d{6})")[0]
        .fillna("")
    )

    # ==========================
    # 6️⃣ NOMOR KARTU (ANTI SCIENTIFIC)
    # ==========================
    if "NOMOR KARTU" in df.columns:
        df["Nomor Kartu"] = (
            df["NOMOR KARTU"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
        )
    else:
        df["Nomor Kartu"] = ""

    # ==========================
    # 7️⃣ BUANG BARIS INVALID
    # ==========================
    df = df[df["Kode Satker"] != ""]
    df = df[df["NO"].notna()]   # 🔥 pastikan No=1 tidak hilang

    if df.empty:
        return pd.DataFrame()

    return df.reset_index(drop=True)

#DATA IKPA KPPN
def get_all_kppn_files(repo, path="Data IKPA KPPN"):
    all_files = []

    try:
        contents = repo.get_contents(path)
    except Exception as e:
        return all_files

    for c in contents:
        if c.type == "dir":
            all_files.extend(get_all_kppn_files(repo, c.path))
        elif c.name.lower().endswith(".xlsx"):  # 🔥 FIX DI SINI
            all_files.append(c.path)

    return all_files

# ============================
#  BACA TEMPLATE FILE
# ============================
def get_template_file():
    try:
        if Path(TEMPLATE_PATH).exists():
            with open(TEMPLATE_PATH, "rb") as f:
                return f.read()
        else:
            if "template_file" in st.session_state:
                return st.session_state.template_file
            return None
    except Exception as e:
        st.error(f"Error membaca template: {e}")
        return None

# Fungsi visualisasi podium/bintang
def create_ranking_chart(df, title, top=True, limit=10):
    """
    Membuat visualisasi ranking dengan bar chart horizontal yang menarik
    (Sekarang menggunakan kolom 'Satker' untuk label agar unik)
    """
    if top:
        df_sorted = df.nlargest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Greens'
        emoji = '🏆'
    else:
        df_sorted = df.nsmallest(limit, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        color_scale = 'Reds'
        emoji = '⚠️'
    
    fig = go.Figure()
    
    colors = px.colors.sequential.Greens if top else px.colors.sequential.Reds
    
    nilai_col = 'Nilai Akhir (Nilai Total/Konversi Bobot)'
    min_val = df_sorted[nilai_col].min()
    max_val = df_sorted[nilai_col].max()

    # use 'Satker' for y labels to keep them unique
    fig.add_trace(go.Bar(
    y=df_sorted['Satker'],
    x=df_sorted[nilai_col],
    orientation='h',
    marker=dict(
        color=df_sorted[nilai_col],
        colorscale='OrRd_r',
        showscale=True,
        cmin=min_val,
        cmax=max_val,
    ),
    text=df_sorted[nilai_col].round(2),
    textposition='outside',
    hovertemplate='<b>%{y}</b><br>Nilai: %{x:.2f}<extra></extra>'
))
    
    fig.update_layout(
        title=f"{emoji} {title}",
        xaxis_title="Nilai Akhir",
        yaxis_title="",
        height=max(400, limit * 40),
        yaxis={'categoryorder': 'total ascending' if not top else 'total descending'},
        showlegend=False
    )
    # ============================
    # Rotated labels 45° di bawah
    # ============================
    annotations = []
    y_positions = list(range(len(df_sorted)))

    for i, satker in enumerate(df_sorted['Satker']):
        annotations.append(dict(
        x=df_sorted[nilai_col].min() - 3,
        y=i,
        text=satker,
        xanchor="right",
        yanchor="middle",
        showarrow=False,
        textangle=45,
        font=dict(size=10),
    ))

    fig.update_layout(annotations=annotations)

    # Sembunyikan label Y-axis
    fig.update_yaxes(showticklabels=False)

    return fig

# ============================================================
# Improved Problem Chart (with sorting, sliders, and filters)
# ============================================================
def get_top_bottom(df, n=10, top=True):
    if df.empty:
        return df
    return (
        df.nlargest(n, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
        if top else
        df.nsmallest(n, 'Nilai Akhir (Nilai Total/Konversi Bobot)')
    )


def make_column_chart(data, title, color_scale, y_min, y_max):
    if data is None or data.empty:
        return None

    nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
    plot_df = data.copy()

    if "Satker" not in plot_df.columns:
        return None

    
    fig = px.bar(
        plot_df.sort_values(nilai_col),
        x=nilai_col,
        y="Satker",
        orientation="h",
        color=nilai_col,
        color_continuous_scale=color_scale,
        title=title
    )

    fig.update_layout(
        xaxis_range=[y_min, y_max],
        xaxis_title="Nilai IKPA",
        yaxis_title="",
        height=450,
        margin=dict(l=10, r=10, t=40, b=20),
        coloraxis_showscale=False,
        showlegend=False
    )

    fig.update_traces(
        texttemplate="%{x:.2f}",
        textposition="outside",
        hovertemplate="<b>%{y}</b><br>Nilai: %{x:.2f}<extra></extra>"
    )

    return fig

# ============================================================
# Problem Chart untuk Dashboard Internal
# ============================================================
def create_problem_chart(df, column, threshold, title, comparison='less', y_min=None, y_max=None, show_yaxis=True):
    
    if comparison == 'less':
        df_filtered = df[df[column] < threshold]
    elif comparison == 'greater':
        df_filtered = df[df[column] > threshold]
    else:
        df_filtered = df.copy()

    # Jika hasil filter kosong → Cegah error
    if df_filtered.empty:
        df_filtered = df.head(1)

    df_filtered = df_filtered.sort_values(by=column, ascending=False)

    # Ambil nilai range untuk colormap
    min_val = df_filtered[column].min()
    max_val = df_filtered[column].max()

    fig = go.Figure()
    fig.add_trace(go.Bar(
    x=df_filtered['Satker'],
    y=df_filtered[column],
    marker=dict(
        color=df_filtered[column],
        colorscale='OrRd_r',
        showscale=True,
        cmin=min_val,
        cmax=max_val,
        colorbar=dict(
            x=1.01,          # ⬅️ DEKATKAN KE CHART
            thickness=12,    # ⬅️ LEBIH RAMPING
            len=0.85         # ⬅️ TIDAK TERLALU TINGGI
        )
        ),
        text=df_filtered[column].round(2),
        textposition='outside',
        textangle=0,
        textfont=dict(family="Arial Black", size=12),
        hovertemplate='<b>%{x}</b><br>Nilai: %{y:.2f}<extra></extra>'
    ))


    # Garis target threshold (tidak berubah)
    fig.add_hline(
        y=threshold,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Target: {threshold}",
        annotation_position="top right"
    )

    # Bold judul dan label axis
    fig.update_layout(
        xaxis=dict(
        tickangle=-45,
        tickmode='linear',
        tickfont=dict(family="Arial Black", size=10),
        automargin=True
    ),
    yaxis=dict(
        tickfont=dict(family="Arial Black", size=11)
        ),
        height=600,
        margin=dict(l=50, r=20, t=80, b=200),
        showlegend=False,
    )

    if not show_yaxis:
        fig.update_yaxes(showticklabels=False)

    return fig


def create_internal_problem_chart_vertical(
    df,
    column,
    threshold,
    title,
    comparison='less',
    show_yaxis=True,
    show_colorbar=True,
    fixed_height=None
):
    import plotly.graph_objects as go
    import pandas as pd

    # ===============================
    # 🔥 VALIDASI AWAL
    # ===============================
    if df is None or column not in df.columns:
        df = pd.DataFrame(columns=["Satker", column])

    df = df.copy()

    # ===============================
    # 🔥 CLEAN DATA
    # ===============================
    df[column] = pd.to_numeric(df[column], errors="coerce")
    df = df.dropna(subset=[column])

    # ===============================
    # 🔥 FILTER < 100
    # ===============================
    if comparison == 'less':
        df_problem = df[df[column] < threshold]
    elif comparison == 'greater':
        df_problem = df[df[column] > threshold]
    else:
        df_problem = df

    # ===============================
    # 🔥 JIKA ADA DATA <100 → tampilkan itu
    # ===============================
    if not df_problem.empty:
        df = df_problem
        note_text = ""
    else:
        # 🔥 fallback: tetap tampilkan 5 terendah biar tidak kosong
        df = df.sort_values(by=column, ascending=True).head(5)
        note_text = " (Semua ≥ 100)"

    df = df.sort_values(by=column, ascending=False)

    jumlah_satker = len(df)

    # ===============================
    # 🎯 HEIGHT
    # ===============================
    BAR_HEIGHT = 38
    BASE_HEIGHT = 260
    MAX_HEIGHT = 1200

    if fixed_height is not None:
        height = fixed_height
    else:
        height = BASE_HEIGHT + (jumlah_satker * BAR_HEIGHT)
        height = min(max(height, 420), MAX_HEIGHT)

    # ===============================
    # 📊 FIGURE
    # ===============================
    fig = go.Figure()

    fig.add_bar(
        x=df["Satker"],
        y=df[column],
        marker=dict(
            color=df[column],
            colorscale="OrRd_r",
            showscale=show_colorbar,
            colorbar=dict(
                thickness=12,
                len=0.85
            ) if show_colorbar else None
        ),
        text=df[column].round(2),
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Nilai: %{y:.2f}<extra></extra>"
    )

    # 🔥 garis target
    fig.add_hline(
        y=threshold,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Target: {threshold}",
        annotation_position="top right"
    )

    # 🔥 kalau semua ≥100 → kasih info di title
    fig.update_layout(
        title=title + note_text,
        height=height,
        margin=dict(l=50, r=20, t=80, b=200),
        xaxis_tickangle=-45,
        showlegend=False
    )

    if not show_yaxis:
        fig.update_yaxes(showticklabels=False)

    return fig

# ===============================================
# Helper to apply reference short names (Simplified)
# ===============================================
def apply_reference_short_names(df):
    df = df.copy()

    # ===============================
    # NORMALISASI KODE SATKER 🔥
    # ===============================
    def clean_kode(x):
        return str(x).strip().replace(".0", "").zfill(6)

    # ===============================
    # VALIDASI KOLOM
    # ===============================
    if 'Kode Satker' not in df.columns:
        df['Kode Satker'] = ''

    df['Kode Satker'] = df['Kode Satker'].apply(clean_kode)

    # ===============================
    # FALLBACK: gunakan Uraian Satker dari file sebagai nama awal
    # ===============================
    uraian_asli = None
    for col in ["Uraian Satker", "Uraian Satker-RINGKAS", "Satker"]:
        if col in df.columns:
            uraian_asli = df[col].astype(str)
            break
    if uraian_asli is None:
        uraian_asli = pd.Series([""] * len(df), index=df.index)

    # ===============================
    # LOAD REFERENCE
    # ===============================
    ref = st.session_state.get("reference_df", pd.DataFrame()).copy()

    # ===============================
    # JIKA TIDAK ADA REFERENCE → FALLBACK ke nama asli
    # ===============================
    if ref.empty or 'Kode Satker' not in ref.columns or 'Uraian Satker-SINGKAT' not in ref.columns:
        df['Uraian Satker-RINGKAS'] = uraian_asli
        df['Satker'] = df['Uraian Satker-RINGKAS'] + " (" + df['Kode Satker'] + ")"
        return df

    ref['Kode Satker'] = ref['Kode Satker'].apply(clean_kode)

    # ===============================
    # HAPUS KOLOM LAMA DULU
    # ===============================
    df = df.drop(columns=['Uraian Satker-RINGKAS'], errors='ignore')

    # ===============================
    # MERGE AMAN
    # ===============================
    df = df.merge(
        ref[['Kode Satker', 'Uraian Satker-SINGKAT']]
        .rename(columns={'Uraian Satker-SINGKAT': 'Uraian Satker-RINGKAS'}),
        on='Kode Satker',
        how='left'
    )

    # ===============================
    # PASTIKAN KOLOM ADA
    # ===============================
    if 'Uraian Satker-RINGKAS' not in df.columns:
        df['Uraian Satker-RINGKAS'] = None

    # ===============================
    # FALLBACK ke nama asli dari file jika referensi tidak punya
    # ===============================
    df['Uraian Satker-RINGKAS'] = df['Uraian Satker-RINGKAS'].fillna(uraian_asli)
    # Pastikan tidak kosong
    df['Uraian Satker-RINGKAS'] = df['Uraian Satker-RINGKAS'].replace("", pd.NA).fillna(uraian_asli)

    # ===============================
    # AUTO RINGKAS TAMBAHAN 🔥
    # ===============================
    df['Uraian Satker-RINGKAS'] = (
        df['Uraian Satker-RINGKAS']
        .astype(str)
        .str.replace("KANTOR KEMENTERIAN AGAMA", "Kemenag", regex=False)
        .str.replace("PENGADILAN AGAMA", "PA", regex=False)
        .str.replace("RUMAH TAHANAN NEGARA", "Rutan", regex=False)
        .str.replace("LEMBAGA PEMASYARAKATAN", "Lapas", regex=False)
        .str.replace("BADAN PUSAT STATISTIK", "BPS", regex=False)
        .str.replace("KANTOR PELAYANAN PERBENDAHARAAN NEGARA", "KPPN", regex=False)
        .str.replace("KANTOR PELAYANAN PAJAK PRATAMA", "KPP Pratama", regex=False)
        .str.replace("KABUPATEN", "Kab.", regex=False)
    )

    # ===============================
    # KOLOM FINAL UNTUK CHART 🔥
    # ===============================
    df['Satker'] = (
        df['Uraian Satker-RINGKAS'] +
        " (" + df['Kode Satker'] + ")"
    )

    return df


# ===============================================
# UPDATED: Helper function to create Satker column consistently
# ===============================================
def create_satker_column(df):
    """
    Creates 'Satker' column consistently across all data sources.
    Should be called after apply_reference_short_names().
    """
    if 'Uraian Satker-RINGKAS' not in df.columns:
        # fallback to older field names
        if 'Uraian Satker Final' in df.columns:
            df['Uraian Satker-RINGKAS'] = df['Uraian Satker Final']
        else:
            df['Uraian Satker-RINGKAS'] = df.get('Uraian Satker', '')

    # Create Satker display using ringkas
    df['Satker'] = (
        df['Uraian Satker-RINGKAS'].astype(str) + 
        ' (' + df['Kode Satker'].astype(str) + ')'
    )
    # Keep backward compatible column
    df['Uraian Satker Final'] = df['Uraian Satker-RINGKAS']
    return df


def merge_ikpa_with_dipa(df):
    df = df.copy()

    # ===============================
    # VALIDASI
    # ===============================
    if df.empty or "Tahun" not in df.columns:
        df["Total Pagu"] = 0
        return df

    # ===============================
    # NORMALISASI TAHUN
    # ===============================
    df["Tahun"] = pd.to_numeric(df["Tahun"], errors="coerce").fillna(0).astype(int)

    tahun_list = df["Tahun"].unique()
    tahun = int(tahun_list[0])

    dipa_dict = st.session_state.get("DATA_DIPA_by_year", {})

    # ===============================
    # AMBIL DIPA
    # ===============================
    df_dipa = dipa_dict.get(tahun)

    if df_dipa is None or df_dipa.empty:
        df["Total Pagu"] = 0
        return df

    df_dipa = df_dipa.copy()
    
    # ===============================
    #  FIX 1: CEK KOLOM PAGU
    # ===============================
    if "Total Pagu" not in df_dipa.columns:
        # Cari kolom alternatif yang mengandung kata PAGU
        pagu_candidates = [c for c in df_dipa.columns if "PAGU" in str(c).upper()]
        if pagu_candidates:
            df_dipa = df_dipa.rename(columns={pagu_candidates[0]: "Total Pagu"})
        else:
            # Tidak ada kolom pagu sama sekali → lanjut tanpa merge
            df["Total Pagu"] = 0
            return df

    # ===============================
    #  FIX 2: SAMAKAN KODE SATKER
    # ===============================
    df["Kode Satker"] = df["Kode Satker"].apply(normalize_kode_satker)
    df_dipa["Kode Satker"] = df_dipa["Kode Satker"].apply(normalize_kode_satker)


    # ===============================
    #  SAFETY KOLOM PAGU
    # ===============================
    if "Total Pagu" not in df_dipa.columns:
        st.error("❌ Kolom 'Total Pagu' tidak ditemukan saat merge")
        st.stop()

    # ===============================
    # 🔥 CLEAN PAGU (FORMAT INDONESIA)
    # ===============================
    df_dipa["Total Pagu"] = (
        df_dipa["Total Pagu"]
        .astype(str)
        .str.replace(".", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    df_dipa["Total Pagu"] = pd.to_numeric(
        df_dipa["Total Pagu"], errors="coerce"
    ).fillna(0)

    # ===============================
    # 🔥 NORMALISASI KODE SATKER (FINAL FIX)
    # ===============================
    def normalize_satker(x):
        if pd.isna(x):
            return ""
        x = ''.join(filter(str.isdigit, str(x)))
        if x == "":
            return ""
        return x[-6:].zfill(6)

    df["Kode Satker"] = df["Kode Satker"].apply(normalize_satker)
    df_dipa["Kode Satker"] = df_dipa["Kode Satker"].apply(normalize_satker)

    # ===============================
    # 🔥 AGREGASI DIPA (ANTI DUPLIKAT)
    # ===============================
    df_dipa = (
        df_dipa
        .groupby("Kode Satker", as_index=False)["Total Pagu"]
        .max()
    )

    # ===============================
    # 🔥 FINAL FORCE MATCH (WAJIB)
    # ===============================
    df["Kode Satker"] = df["Kode Satker"].astype(str).str[-6:]
    df_dipa["Kode Satker"] = df_dipa["Kode Satker"].astype(str).str[-6:]

    # ===============================
    # MERGE
    # ===============================
    df_merge = df.merge(
        df_dipa,
        on="Kode Satker",
        how="left"
    )

    # ===============================
    # HANDLE NILAI KOSONG
    # ===============================
    df_merge["Total Pagu"] = pd.to_numeric(
        df_merge["Total Pagu"], errors="coerce"
    ).fillna(0)

    return df_merge


def classify_jenis_satker(df):
    """
    Menentukan Jenis Satker sebagai IDENTITAS (FINAL)
    """
    df = df.copy()

    df["Total Pagu"] = pd.to_numeric(
        df.get("Total Pagu", 0),
        errors="coerce"
    ).fillna(0)

    # 🚨 PAKSA RESET
    df["Jenis Satker"] = None

    # kalau semua nol → bagikan merata berdasarkan rank/index
    if df["Total Pagu"].sum() == 0:
        n = len(df)
        labels_merata = (
            ["KECIL"] * (n // 3) +
            ["SEDANG"] * (n // 3) +
            ["BESAR"] * (n - 2 * (n // 3))
        )
        df["Jenis Satker"] = labels_merata
        return df

    p40 = df["Total Pagu"].quantile(0.40)
    p70 = df["Total Pagu"].quantile(0.70)

    # Jika p40 == p70 (semua satker pagu sama), pakai rank-based
    if p40 == p70:
        df["_rank"] = df["Total Pagu"].rank(method="first")
        n = len(df)
        cut40 = n * 0.40
        cut70 = n * 0.70
        df["Jenis Satker"] = df["_rank"].apply(
            lambda r: "KECIL" if r <= cut40 else ("SEDANG" if r <= cut70 else "BESAR")
        )
        df = df.drop(columns=["_rank"])
    else:
        df["Jenis Satker"] = pd.cut(
            df["Total Pagu"],
            bins=[-float("inf"), p40, p70, float("inf")],
            labels=["KECIL", "SEDANG", "BESAR"]
        )

    df["Jenis Satker"] = df["Jenis Satker"].astype(str)

    return df



# BAGIAN 4 CHART DASHBOARD UTAMA
def safe_chart(
    df_part,
    jenis,
    top=True,
    color="Greens",
    y_min=None,
    y_max=None,
    thin_bar=False
):
    # ===============================
    # PROTEKSI AWAL
    # ===============================
    if df_part is None or df_part.empty:
        st.info("Tidak ada data.")
        return
    
    df_part = apply_reference_short_names(df_part)


    if "Satker" not in df_part.columns:
        st.warning("Kolom Satker belum siap.")
        return

    kandidat_ikpa = [
        "Nilai Akhir (Nilai Total/Konversi Bobot)",
        "Nilai Total/Konversi Bobot",
        "Nilai Total"
    ]

    nilai_col = next((c for c in kandidat_ikpa if c in df_part.columns), None)
    if nilai_col is None:
        st.warning(f"Kolom IKPA tidak ditemukan untuk {jenis}")
        return

    # ===============================
    # SORT DATA
    # ===============================
    df_sorted = (
        df_part
        .sort_values(nilai_col, ascending=not top)
        .head(10)
        .sort_values(nilai_col, ascending=True)
        .copy()
    )

    # ===============================
    # PROTEKSI PLOTLY
    # ===============================
    df_sorted["Satker"] = df_sorted["Satker"].astype(str).str.strip()
    df_sorted = df_sorted[df_sorted["Satker"] != ""]

    if df_sorted.empty:
        st.info("Tidak ada data valid untuk ditampilkan.")
        return

    # ===============================
    # PLOT
    # ===============================
    fig = px.bar(
        df_sorted,
        x=nilai_col,
        y="Satker",
        orientation="h",
        color=nilai_col,
        color_continuous_scale=color,
        text=nilai_col
    )

    fig.update_traces(
        width=0.65 if thin_bar else 0.8,
        texttemplate="%{text:.2f}",
        textposition="outside",
        cliponaxis=False
    )

    fig.update_layout(
        height=200,
        bargap=0.05,
        margin=dict(l=2, r=2, t=0, b=0),
        xaxis_title=None,
        yaxis_title=None,
        xaxis=dict(range=[y_min, y_max]),
        coloraxis_showscale=False
    )

    st.plotly_chart(fig, use_container_width=True)
    

def get_top_bottom_unique(
    df,
    value_col,
    n=10,
    kode_col="Kode Satker"
):
    """
    Ambil Top N dan Bottom N TANPA satker ganda
    """
    # Top N
    top_df = (
        df.sort_values(value_col, ascending=False)
          .head(n)
          .copy()
    )

    # Kode satker yang sudah tampil di Top
    used_kode = set(top_df[kode_col].astype(str))

    # Bottom N (buang yang sudah muncul di Top)
    bottom_df = (
        df[~df[kode_col].astype(str).isin(used_kode)]
        .sort_values(value_col, ascending=True)
        .head(n)
        .copy()
    )

    return top_df, bottom_df


def dynamic_title(ukuran, kategori, df):
    """
    Judul otomatis sesuai jumlah data
    """
    jumlah = len(df)
    return f"{jumlah} Satker {ukuran} {kategori}"    

# ===============================
# FORMAT TAMPILAN IKPA
# ===============================
MONTH_ABBR = {
    "JANUARI": "Jan",
    "FEBRUARI": "Feb",
    "MARET": "Mar",
    "APRIL": "Apr",
    "MEI": "Mei",
    "JUNI": "Jun",
    "JULI": "Jul",
    "AGUSTUS": "Agu",
    "SEPTEMBER": "Sep",
    "OKTOBER": "Okt",
    "NOVEMBER": "Nov",
    "NOPEMBER": "Nov",
    "DESEMBER": "Des"
}

def format_ikpa_display(x):
    try:
        x = float(x)
        if round(x, 2) == 100:
            return "100"
        return f"{x:.2f}"
    except:
        return x


# -----------------------------------
# AGREGASI DIGIPAY
# -----------------------------------

def clean_nominal(series):
    """
    Membersihkan angka format Indonesia dengan aman
    """

    s = series.astype(str)

    s = s.str.replace(".", "", regex=False)
    s = s.str.replace(",", ".", regex=False)

    return pd.to_numeric(s, errors="coerce").fillna(0)


def generate_digipay_chart(df, periode="Bulanan", tipe="trx", tahun_filter=None):

    df = df.copy()

    df["TAHUN"] = pd.to_numeric(df["TAHUN"], errors="coerce")
    df["BULAN"] = pd.to_numeric(df["BULAN"], errors="coerce")

    df["TRIWULAN"] = ((df["BULAN"] - 1) // 3) + 1

    df["NOMINVOICE"] = clean_nominal(df["NOMINVOICE"])

    if tahun_filter is not None:
        df = df[df["TAHUN"] == int(tahun_filter)]

    if df.empty:
        return pd.DataFrame()

    if periode == "Bulanan":
        group_col = "BULAN"
    elif periode == "Triwulan":
        group_col = "TRIWULAN"
    else:
        group_col = "TAHUN"

    if tipe == "trx":
        grouped = df.groupby(group_col)["NOINVOICE"].nunique()
    else:
        grouped = df.groupby(group_col)["NOMINVOICE"].sum()

    grouped = grouped.sort_index().to_frame("Value")

    grouped["Kumulatif"] = grouped["Value"].cumsum()

    return grouped.reset_index()

def generate_digipay_monthly_from_session(df, tahun_filter=None, tipe="trx"):

    df = df.copy()

    df["TANGGAL"] = pd.to_datetime(df["TANGGAL"], errors="coerce")

    df["Tahun"] = df["TANGGAL"].dt.year
    df["Bulan"] = df["TANGGAL"].dt.month

    df["NOMINVOICE"] = clean_nominal(df["NOMINVOICE"])

    if tahun_filter is not None:
        df = df[df["Tahun"] == tahun_filter]

    if tipe == "trx":

        agg_df = (
            df.groupby(["SATKER","Bulan"])
            .agg(Jumlah_Transaksi=("NOINVOICE","nunique"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["SATKER","Bulan"])
            .agg(Nilai_Transaksi=("NOMINVOICE","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = agg_df.pivot(
        index="SATKER",
        columns="Bulan",
        values=value_col
    ).fillna(0)

    pivot = pivot.reindex(columns=range(1,13), fill_value=0)

    bulan_map = {
        1:"JAN",2:"FEB",3:"MAR",4:"APR",
        5:"MEI",6:"JUN",7:"JUL",8:"AGU",
        9:"SEP",10:"OKT",11:"NOV",12:"DES"
    }

    pivot.columns = [bulan_map[i] for i in pivot.columns]

    return pivot.reset_index()

def generate_digipay_quarterly_from_session(df, tahun_filter=None, tipe="trx"):

    df = df.copy()

    df["TANGGAL"] = pd.to_datetime(df["TANGGAL"], errors="coerce")

    df["Tahun"] = df["TANGGAL"].dt.year
    df["Bulan"] = df["TANGGAL"].dt.month

    df["Triwulan"] = ((df["Bulan"] - 1) // 3) + 1

    df["NOMINVOICE"] = clean_nominal(df["NOMINVOICE"])

    if tahun_filter is not None:
        df = df[df["Tahun"] == tahun_filter]

    if tipe == "trx":

        agg_df = (
            df.groupby(["SATKER","Triwulan"])
            .agg(Jumlah_Transaksi=("NOINVOICE","nunique"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["SATKER","Triwulan"])
            .agg(Nilai_Transaksi=("NOMINVOICE","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = agg_df.pivot(
        index="SATKER",
        columns="Triwulan",
        values=value_col
    ).fillna(0)

    pivot = pivot.reindex(columns=[1,2,3,4], fill_value=0)

    pivot.columns = ["TW1","TW2","TW3","TW4"]

    return pivot.reset_index()

def generate_digipay_yearly_from_session(df, tipe="trx"):

    df = df.copy()

    df["TANGGAL"] = pd.to_datetime(df["TANGGAL"], errors="coerce")

    df["Tahun"] = df["TANGGAL"].dt.year

    df["NOMINVOICE"] = clean_nominal(df["NOMINVOICE"])

    if tipe == "trx":

        agg_df = (
            df.groupby(["SATKER","Tahun"])
            .agg(Jumlah_Transaksi=("NOINVOICE","nunique"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["SATKER","Tahun"])
            .agg(Nilai_Transaksi=("NOMINVOICE","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = (
        agg_df
        .pivot_table(
            index="SATKER",
            columns="Tahun",
            values=value_col,
            fill_value=0
        )
        .sort_index(axis=1)
    )

    pivot.columns = pivot.columns.astype(str)

    return pivot.reset_index()

# -----------------------------------
# AGREGASI KKP
# -----------------------------------
def add_kkp_pagu_column(df_pivot, df_master):
    
    df_master = df_master.copy()

    # normalisasi kode satker
    df_master["Kode Satker"] = df_master["Kode Satker"].astype(str).str.zfill(6)
    df_pivot["Kode Satker"] = df_pivot["Kode Satker"].astype(str).str.zfill(6)

    # normalisasi limit
    df_master["LIMIT KKP"] = (
        df_master["LIMIT KKP"]
        .astype(str)
        .str.replace(r"[^\d]", "", regex=True)
    )

    df_master["LIMIT KKP"] = pd.to_numeric(
        df_master["LIMIT KKP"],
        errors="coerce"
    ).fillna(0)

    # pagu per satker
    pagu_map = (
        df_master
        .groupby("Kode Satker")["LIMIT KKP"]
        .sum()
        .to_dict()
    )

    # mapping pagu
    df_pivot["PAGU KKP"] = df_pivot["Kode Satker"].map(pagu_map).fillna(0)

    # posisi kolom setelah SATKER
    cols = df_pivot.columns.tolist()

    if "PAGU KKP" in cols:
        cols.remove("PAGU KKP")

    insert_pos = 2 if "SATKER" in cols else 1
    cols.insert(insert_pos, "PAGU KKP")

    df_pivot = df_pivot[cols]

    return df_pivot

def generate_kkp_chart(df, periode="Bulanan", tahun_filter=None):
    
    df = df.copy()

    df["PERIODE"] = pd.to_datetime(df["PERIODE"], errors="coerce")

    df["TAHUN"] = df["PERIODE"].dt.year
    df["BULAN"] = df["PERIODE"].dt.month
    df["TRIWULAN"] = df["PERIODE"].dt.quarter

    df["NILAI TRANSAKSI (NILAI SPM)"] = clean_nominal(
        df["NILAI TRANSAKSI (NILAI SPM)"]
    )

    df["LIMIT KKP"] = clean_nominal(df["LIMIT KKP"])

    if tahun_filter is not None:
        df = df[df["TAHUN"] == tahun_filter]

    grouped = (
        df.groupby(["TAHUN","BULAN"])
        .agg(
            NOMINAL=("NILAI TRANSAKSI (NILAI SPM)","sum"),
            PAGU=("LIMIT KKP","sum")
        )
        .reset_index()
    )

    grouped["Kumulatif Nominal"] = grouped["NOMINAL"].cumsum()

    grouped["% Realisasi"] = (
        grouped["Kumulatif Nominal"] /
        grouped["PAGU"].replace(0, pd.NA)
    ) * 100

    grouped["% Realisasi"] = grouped["% Realisasi"].fillna(0)

    return grouped


# -----------------------------------
# KKP BULANAN
# -----------------------------------
def generate_kkp_monthly_from_session(df, tahun_filter=None, tipe="trx"):
    
    df = df.copy()

    df["PERIODE"] = pd.to_datetime(df["PERIODE"], errors="coerce")

    df["Tahun"] = df["PERIODE"].dt.year
    df["Bulan"] = df["PERIODE"].dt.month

    if tahun_filter is not None:
        df = df[df["Tahun"] == tahun_filter]

    satker_map = (
        df[["Kode Satker","SATKER"]]
        .drop_duplicates()
        .set_index("Kode Satker")["SATKER"]
    )

    if tipe == "trx":
    
        agg_df = (
            df.groupby(["Kode Satker","Bulan"])
            .agg(Jumlah_Transaksi=("NOMOR KARTU","count"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["Kode Satker","Bulan"])
            .agg(Nilai_Transaksi=("NILAI TRANSAKSI (NILAI SPM)","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = agg_df.pivot(
        index="Kode Satker",
        columns="Bulan",
        values=value_col
    ).fillna(0)

    pivot = pivot.reindex(columns=range(1,13), fill_value=0)

    bulan_map = {
        1:"JAN",2:"FEB",3:"MAR",4:"APR",
        5:"MEI",6:"JUN",7:"JUL",8:"AGU",
        9:"SEP",10:"OKT",11:"NOV",12:"DES"
    }

    pivot.columns = [bulan_map[i] for i in pivot.columns]

    pivot = pivot.reset_index()

    pivot["SATKER"] = pivot["Kode Satker"].map(satker_map)

    cols = ["Kode Satker","SATKER"] + [c for c in pivot.columns if c not in ["Kode Satker","SATKER"]]

    return pivot[cols]


# -----------------------------------
# KKP TRIWULAN
# -----------------------------------
def generate_kkp_quarterly_from_session(df, tahun_filter=None, tipe="trx"):

    df = df.copy()

    df["PERIODE"] = pd.to_datetime(df["PERIODE"], errors="coerce")

    df["Tahun"] = df["PERIODE"].dt.year
    df["Bulan"] = df["PERIODE"].dt.month

    df["Triwulan"] = ((df["Bulan"] - 1) // 3) + 1

    if tahun_filter is not None:
        df = df[df["Tahun"] == tahun_filter]

    satker_map = (
        df[["Kode Satker","SATKER"]]
        .drop_duplicates()
        .set_index("Kode Satker")["SATKER"]
    )

    if tipe == "trx":
    
        agg_df = (
            df.groupby(["Kode Satker","Triwulan"])
            .agg(Jumlah_Transaksi=("NOMOR KARTU","count"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["Kode Satker","Triwulan"])
            .agg(Nilai_Transaksi=("NILAI TRANSAKSI (NILAI SPM)","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = agg_df.pivot(
        index="Kode Satker",
        columns="Triwulan",
        values=value_col
    ).fillna(0)

    pivot = pivot.reindex(columns=[1,2,3,4], fill_value=0)

    pivot.columns = ["TW1","TW2","TW3","TW4"]

    pivot = pivot.reset_index()

    pivot["SATKER"] = pivot["Kode Satker"].map(satker_map)

    cols = ["Kode Satker","SATKER"] + [c for c in pivot.columns if c not in ["Kode Satker","SATKER"]]

    return pivot[cols]


# -----------------------------------
# KKP TAHUNAN
# -----------------------------------
def generate_kkp_yearly_from_session(df, tipe="trx"):

    df = df.copy()

    df["PERIODE"] = pd.to_datetime(df["PERIODE"], errors="coerce")

    df["Tahun"] = df["PERIODE"].dt.year

    satker_map = (
        df[["Kode Satker","SATKER"]]
        .drop_duplicates()
        .set_index("Kode Satker")["SATKER"]
    )

    if tipe == "trx":
    
        agg_df = (
            df.groupby(["Kode Satker","Tahun"])
            .agg(Jumlah_Transaksi=("NOMOR KARTU","count"))
            .reset_index()
        )

        value_col = "Jumlah_Transaksi"

    else:

        agg_df = (
            df.groupby(["Kode Satker","Tahun"])
            .agg(Nilai_Transaksi=("NILAI TRANSAKSI (NILAI SPM)","sum"))
            .reset_index()
        )

        value_col = "Nilai_Transaksi"

    pivot = (
        agg_df
        .pivot_table(
            index="Kode Satker",
            columns="Tahun",
            values=value_col,
            fill_value=0
        )
        .sort_index(axis=1)
    )

    pivot.columns = pivot.columns.astype(str)

    pivot = pivot.reset_index()

    pivot["SATKER"] = pivot["Kode Satker"].map(satker_map)

    cols = ["Kode Satker","SATKER"] + [c for c in pivot.columns if c not in ["Kode Satker","SATKER"]]

    return pivot[cols]


# -----------------------------------
# WRAPPER KKP
# -----------------------------------
def generate_kkp_from_session(df, periode="Bulanan", tipe="Jumlah Transaksi", tahun_filter=None):

    tipe_val = "trx" if tipe == "Jumlah Transaksi" else "nom"

    if periode == "Bulanan":
        return generate_kkp_monthly_from_session(df, tahun_filter, tipe_val)

    elif periode == "Triwulan":
        return generate_kkp_quarterly_from_session(df, tahun_filter, tipe_val)

    else:
        return generate_kkp_yearly_from_session(df, tipe_val)
    
    
# Persentase Realisasi KKP
def add_kkp_percentage_columns(df_pivot, df_master):
    
    df = df_master.copy()

    df["PERIODE"] = pd.to_datetime(df["PERIODE"], errors="coerce")
    df["LIMIT KKP"] = clean_nominal(df["LIMIT KKP"])

    df["Kode Satker"] = df["Kode Satker"].astype(str).str.zfill(6)
    df_pivot["Kode Satker"] = df_pivot["Kode Satker"].astype(str).str.zfill(6)

    limit_map = (
        df.groupby("Kode Satker")["LIMIT KKP"]
        .sum()
        .to_dict()
    )

    df_pivot = df_pivot.copy()

    value_cols = [
        c for c in df_pivot.columns
        if c not in ["Kode Satker","SATKER","PAGU KKP"]
    ]

    df_pivot[value_cols] = df_pivot[value_cols].astype(float)

    limit_series = df_pivot["Kode Satker"].map(limit_map)

    for col in value_cols:

        persen = (df_pivot[col] / limit_series.replace(0, pd.NA)) * 100
        persen = persen.fillna(0)

        df_pivot.insert(
            df_pivot.columns.get_loc(col) + 1,
            f"{col} % Realisasi KKP",
            persen.round(2)
        )

    return df_pivot
    

# =========================================================================
# PROPORSI CMS
# =========================================================================
def generate_cms_from_session(df_master, periode="Tahunan", tahun_filter=None):
    
    df = df_master.copy()

    # =============================
    # FILTER TAHUN
    # =============================
    if tahun_filter is not None:
        df = df[df["TAHUN"] == tahun_filter]

    # =============================
    # NORMALISASI NUMERIK
    # =============================
    numeric_cols = [
        "JUMLAH TRANSAKSI CMS",
        "NILAI TRANSAKSI CMS",
        "JUMLAH TRANSAKSI KARTU DEBIT",
        "NILAI TRANSAKSI KARTU DEBIT",
        "JUMLAH TRANSAKSI TELLER",
        "NILAI TRANSAKSI TELLER",
    ]

    for col in numeric_cols:

        df[col] = (
            df[col]
            .astype(str)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )

        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)


    # =============================
    # AGREGASI BERDASARKAN KODE SATKER
    # =============================
    df_group = (
        df.groupby("KODE SATKER")
        .agg(
            NAMA_SATKER=("NAMA SATKER", "first"),

            CMS_TRX=("JUMLAH TRANSAKSI CMS", "sum"),
            CMS_NOM=("NILAI TRANSAKSI CMS", "sum"),

            DEBIT_TRX=("JUMLAH TRANSAKSI KARTU DEBIT", "sum"),
            DEBIT_NOM=("NILAI TRANSAKSI KARTU DEBIT", "sum"),

            TELLER_TRX=("JUMLAH TRANSAKSI TELLER", "sum"),
            TELLER_NOM=("NILAI TRANSAKSI TELLER", "sum"),
        )
        .reset_index()
    )

    df_group = df_group.rename(columns={
        "NAMA_SATKER": "NAMA SATKER"
    })
    
    
    # =============================
    # TOTAL
    # =============================
    df_group["TOTAL_TRX"] = (
        df_group["CMS_TRX"]
        + df_group["DEBIT_TRX"]
        + df_group["TELLER_TRX"]
    )

    df_group["TOTAL_NOM"] = (
        df_group["CMS_NOM"]
        + df_group["DEBIT_NOM"]
        + df_group["TELLER_NOM"]
    )

    # =============================
    # PROPORSI CMS
    # =============================
    df_group["Proporsi Transaksi CMS"] = (
        df_group["CMS_TRX"] / df_group["TOTAL_TRX"] * 100
    )

    df_group["Proporsi Nominal CMS"] = (
        df_group["CMS_NOM"] / df_group["TOTAL_NOM"] * 100
    )
    

    # =============================
    # OUTPUT
    # =============================
    result = df_group[
        [
            "KODE SATKER",
            "NAMA SATKER",
            "Proporsi Transaksi CMS",
            "Proporsi Nominal CMS",
        ]
    ]

    return result.fillna(0)

if "cms_master" not in st.session_state:
    st.session_state.cms_master = pd.DataFrame()

# HALAMAN 1: DASHBOARD UTAMA
def page_dashboard():
    
    df_full = pd.DataFrame()
    
    # ===============================
    # LOAD & MAP BA (WAJIB DI SINI)
    # ===============================
    df_ref_ba = load_reference_ba()
    BA_MAP = get_ba_map(df_ref_ba)

    # ===============================
    # SOLUSI 3 — DELAY RENDER SETELAH UPLOAD
    # ===============================
    if st.session_state.get("_just_uploaded"):
        st.session_state["_just_uploaded"] = False
        st.info("🔄 Data baru dimuat, mempersiapkan grafik...")
        st.rerun()

    st.markdown('<div class="hero">', unsafe_allow_html=True)
    st.markdown('<div class="hero-content">', unsafe_allow_html=True)

    st.markdown('<div class="hero-title">Dashboard Utama Kinerja Keuangan</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-sub">Satker Mitra KPPN Baturaja</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    
    if "main_menu" not in st.session_state:
        st.session_state.main_menu = None

    st.markdown('<div class="menu-header">Pilih Menu</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:

        if st.button(
            "📊\n\nIKPA\n\nIndikator Kinerja Pelaksanaan Anggaran",
            use_container_width=True
        ):
            st.session_state.main_menu = "IKPA"
            st.rerun()

    with col2:

        if st.button(
            "💳\n\nDigitalisasi\n\nCMS • DIGIPAY • KKP",
            use_container_width=True
        ):
            st.session_state.main_menu = "Digitalisasi"
            st.rerun()
    
    # ===============================
    # STOP DI SINI JIKA BELUM PILIH
    # ===============================
    if st.session_state.main_menu is None:
        return

    
    # ===============================
    # VALIDASI & PILIH PERIODE (FINAL)
    # ===============================

    data_storage = st.session_state.get("data_storage", {})

    if not isinstance(data_storage, dict) or len(data_storage) == 0:
        st.warning("⚠️ Data IKPA belum tersedia.")
        return
        
    
    # Ambil & urutkan semua periode (bulan, tahun)
    try:
        all_periods = sorted(
            data_storage.keys(),
            key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)),
            reverse=True
        )
    except Exception:
        st.warning("⚠️ Format periode pada data tidak sesuai.")
        return

    if not all_periods:
        st.warning("⚠️ Belum ada data periode IKPA yang valid.")
        return


    # ===============================
    # AMBIL DF AKTIF (GLOBAL)
    # ===============================
    if "selected_period" not in st.session_state:
        st.session_state.selected_period = all_periods[0]
    
    # ===============================
    # NORMALISASI KEY 
    # ===============================
    sel = st.session_state.selected_period
    key = (str(sel[0]).upper().strip(), str(sel[1]))

    df = st.session_state.data_storage.get(key)

    if df is not None:
        df = df.copy()


    st.markdown("""
    <style>

    /* =================================================
    1. KECILKAN HANYA FILTER KODE BA
    ================================================= */
    .filter-ba h3 {
        font-size: 15px !important;
        margin-bottom: 4px !important;
    }

    .filter-ba label {
        font-size: 12px !important;
        margin-bottom: 2px !important;
    }

    .filter-ba div[data-baseweb="select"] {
        font-size: 12px !important;
        min-height: 30px !important;
        margin-bottom: 6px !important;
    }

    .filter-ba div[data-baseweb="tag"] span {
        font-size: 11px !important;
        padding: 2px 6px !important;
    }


    /* =================================================
    2. BESARKAN RADIO PILIH BAGIAN DASHBOARD
    (SETARA ## HEADING)
    ================================================= */
    div[role="radiogroup"] {
        margin-top: 6px !important;
    }

    div[role="radiogroup"] > label {
        font-size: 24px !important;
        font-weight: 700 !important;
        margin-bottom: 6px !important;
    }

    div[role="radiogroup"] label p {
        font-size: 24px !important;
        font-weight: 600 !important;
        margin: 0 16px 0 0 !important;
    }

    /* =================================================
    3. KUNCI SELECTBOX PERIODE (NOVEMBER 2025)
    AGAR TIDAK PERNAH MENGECIL
    ================================================= */
    div[data-testid="stSelectbox"] div[data-baseweb="select"] {
        font-size: 16px !important;
        min-height: 38px !important;
        line-height: 1.4 !important;
    }

    </style>
    """, unsafe_allow_html=True)

    
    # ===============================
    # FILTER KODE BA
    # ===============================
    st.markdown('<div class="filter-ba">', unsafe_allow_html=True)

    st.markdown("🔎 Filter Kode BA")  

    if df is not None and 'Kode BA' in df.columns:

        df['Kode BA'] = df['Kode BA'].apply(normalize_kode_ba)

        ba_codes = sorted(df['Kode BA'].dropna().unique())
        ba_options = ["SEMUA BA"] + ba_codes

        def format_ba(code):
            if code == "SEMUA BA":
                return "SEMUA BA"
            return f"{code} – {BA_MAP.get(code, 'Nama BA tidak ditemukan')}"

        st.multiselect(
            "Pilih Kode BA",
            options=ba_options,
            format_func=format_ba,
            default=st.session_state.get("filter_ba_main", ["SEMUA BA"]),
            key="filter_ba_main"
        )
    else:
        st.warning("Kolom Kode BA tidak tersedia.")

    st.markdown('</div>', unsafe_allow_html=True)  
    
    # ===============================
    # ROUTING MENU UTAMA
    # ===============================

    if st.session_state.main_menu == "IKPA":

        st.markdown("---")

        pass

        # =================================================
        # 🔑 AMANKAN SESSION STATE RADIO (ANTI VALUEERROR)
        # =================================================
        VALID_MAIN_TABS = [
            "🎯 Highlights Satker",
            "🏢 Highlights BA",
            "📋 Data Detail Satker"
        ]

        if "main_tab" in st.session_state:
            if st.session_state.main_tab not in VALID_MAIN_TABS:
                del st.session_state.main_tab

        # ===============================
        # RADIO PILIH BAGIAN DASHBOARD
        # ===============================
        main_tab = st.radio(
            "Pilih Bagian Dashboard",
            VALID_MAIN_TABS,
            key="main_tab",
            horizontal=True
        )

        # -------------------------
        # HIGHLIGHTS
        # -------------------------
        if main_tab == "🎯 Highlights Satker":
            st.markdown("---")
            st.markdown("## 🎯 Highlights Kinerja Satker")

            st.selectbox(
                "Pilih Periode",
                options=all_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}",
                key="selected_period"
            )

            # ===============================
            # NORMALISASI KEY 
            # ===============================
            sel = st.session_state.selected_period
            key = (str(sel[0]).upper().strip(), str(sel[1]))

            df = st.session_state.data_storage.get(key)

            if df is None or df.empty:
                st.warning("Data IKPA belum tersedia.")
                st.stop()

            
            df = apply_reference_short_names(df)
            
            # ===============================
            # NORMALISASI KODE BA (1x SAJA)
            # ===============================
            if 'Kode BA' in df.columns:
                df['Kode BA'] = df['Kode BA'].apply(normalize_kode_ba)
            
            df = apply_filter_ba(df)

            # ===============================
            # GUNAKAN JENIS SATKER DARI LOADER
            # ===============================
            df['Jenis Satker'] = (
                df['Jenis Satker']
                .astype(str)
                .str.upper()
                .str.strip()
            )

            df_kecil  = df[df['Jenis Satker'] == 'KECIL']
            df_sedang = df[df['Jenis Satker'] == 'SEDANG']
            df_besar  = df[df['Jenis Satker'] == 'BESAR']


            # ===============================
            # METRIK UTAMA
            # ===============================
            nilai_col = 'Nilai Akhir (Nilai Total/Konversi Bobot)'

            avg_score = df[nilai_col].mean()
            perfect_df = df[df[nilai_col] == 100]
            below89_df = df[df[nilai_col] < 89]

            # Pastikan kolom Satker tersedia
            def make_satker_col(dd):
                if 'Satker' in dd.columns:
                    return dd
                uraian = dd.get('Uraian Satker-RINGKAS', dd.index.astype(str))
                kode = dd.get('Kode Satker', '')
                dd = dd.copy()
                dd['Satker'] = uraian.astype(str) + " (" + kode.astype(str) + ")"
                return dd

            perfect_df = make_satker_col(perfect_df)
            below89_df = make_satker_col(below89_df)

            jumlah_100 = len(perfect_df)
            jumlah_below = len(below89_df)

            # Tampilan metrik
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📋 Total Satker", len(df))
            with col2:
                st.metric("📈 Rata-rata Nilai", f"{avg_score:.2f}")
            with col3:
                st.metric("⭐ Nilai 100", jumlah_100)
                with st.popover("Lihat daftar satker"):
                    if jumlah_100 == 0:
                        st.write("Tidak ada satker dengan nilai 100.")
                    else:
                        display_df = perfect_df[['Satker']].reset_index(drop=True)
                        display_df.insert(0, 'No', range(1, len(display_df) + 1))
                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(400, len(display_df) * 35 + 38)
                        )
            with col4:
                st.metric("⚠️ Nilai < 89 (Predikat Belum Baik)", jumlah_below)
                with st.popover("Lihat daftar satker"):
                    if jumlah_below == 0:
                        st.write("Tidak ada satker dengan nilai < 89.")
                    else:
                        display_df = below89_df[['Satker']].reset_index(drop=True)
                        display_df.insert(0, 'No', range(1, len(display_df) + 1))
                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(400, len(display_df) * 35 + 38)
                        )

            # ===============================
            # Kontrol Skala Chart
            # ===============================
            st.markdown("---")
            st.markdown("###### Atur Skala Nilai (Sumbu Y)")
            col_min, col_max = st.columns(2)
            with col_min:
                y_min = st.slider("Nilai Minimum (Y-Axis)", 0, 50, 50, 1, key="high_ymin")
            with col_max:
                y_max = st.slider("Nilai Maksimum (Y-Axis)", 51, 110, 110, 1, key="high_ymax")

            # ===============================
            # CHART 6 DALAM 1 TAMPILAN
            # ===============================
            st.markdown("### 📊 Satker Terbaik & Terendah Berdasarkan Nilai IKPA")

            nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"

            # =========================
            # PREPARE DATA (UNIK)
            # =========================
            top_kecil, bottom_kecil = get_top_bottom_unique(df_kecil, nilai_col)
            top_sedang, bottom_sedang = get_top_bottom_unique(df_sedang, nilai_col)
            top_besar, bottom_besar = get_top_bottom_unique(df_besar, nilai_col)

            # =========================
            # BARIS 1 – TERBAIK
            # =========================
            c1, c2, c3 = st.columns(3)

            with c1:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Kecil','Terbaik', top_kecil)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    top_kecil, "KECIL",
                    top=True, color="Greens",
                    y_min=y_min, y_max=y_max
                )

            with c2:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Sedang','Terbaik', top_sedang)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    top_sedang, "SEDANG",
                    top=True, color="Greens",
                    y_min=y_min, y_max=y_max
                )

            with c3:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Besar','Terbaik', top_besar)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    top_besar, "BESAR",
                    top=True, color="Greens",
                    y_min=y_min, y_max=y_max
                )

            # ⬇️ JARAK ANTAR BARIS
            st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

            # =========================
            # BARIS 2 – TERENDAH
            # =========================
            c4, c5, c6 = st.columns(3)

            with c4:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Kecil','Terendah', bottom_kecil)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    bottom_kecil, "KECIL",
                    top=False, color="Reds",
                    y_min=y_min, y_max=y_max
                )

            with c5:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Sedang','Terendah', bottom_sedang)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    bottom_sedang, "SEDANG",
                    top=False, color="Reds",
                    y_min=y_min, y_max=y_max
                )

            with c6:
                st.markdown(
                    f"<div style='margin-top:2px; margin-bottom:6px'><b>"
                    f"{dynamic_title('Besar','Terendah', bottom_besar)}</b></div>",
                    unsafe_allow_html=True
                )
                safe_chart(
                    bottom_besar, "BESAR",
                    top=False, color="Reds",
                    y_min=y_min, y_max=y_max
                )


            # -----------------------------
            # Slider
            # -----------------------------
            # Satker dengan masalah (Deviasi Hal 3 DIPA)
            st.markdown("---")
            st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")
            st.markdown("###### Atur Skala Nilai (Sumbu Y)")

            col_min_dev, col_max_dev = st.columns(2)

            with col_min_dev:
                st.markdown("**Nilai Minimum (Y-Axis)**")
                y_min_dev = st.slider(
                    "",
                    min_value=0,
                    max_value=50,
                    value=40,
                    step=1,
                    key="high_ymin_dev"
                )

            with col_max_dev:
                st.markdown("**Nilai Maksimum (Y-Axis)**")
                y_max_dev = st.slider(
                    "",
                    min_value=51,
                    max_value=110,
                    value=110,
                    step=1,
                    key="high_ymax_dev"
                )

            # -----------------------------
            # ⚠️ JUDUL CHART
            # -----------------------------
            st.markdown("###### ⚠️ Deviasi Hal 3 DIPA Belum Optimal (< 90)")

            # =================================================
            # 🔑 PERBAIKAN UTAMA DIMULAI DI SINI
            # =================================================
            problem_col = "Deviasi Halaman III DIPA"

            # 1️⃣ PAKSA NUMERIK
            df[problem_col] = pd.to_numeric(
                df[problem_col],
                errors="coerce"
            )

            # 2️⃣ FILTER TEGAS (INI KUNCI)
            df_problem = df[
                (df[problem_col].notna()) &
                (df[problem_col] < 90)
            ].copy()

            # 3️⃣ JIKA TIDAK ADA MASALAH → SELESAI
            if df_problem.empty:
                st.success("✅ Semua satker sudah optimal untuk Deviasi Hal 3 DIPA")
            else:
                fig_dev = create_problem_chart(
                    df_problem,              # ⬅️ BUKAN df LAGI
                    column=problem_col,
                    threshold=90,
                    title="",
                    comparison="less",
                    y_min=y_min_dev,
                    y_max=y_max_dev,
                    show_yaxis=True
                )

                st.plotly_chart(fig_dev, use_container_width=True)
                

        # -------------------------
        # HIGHLIGHTS BA
        # -------------------------
        elif main_tab == "🏢 Highlights BA":
            st.markdown("## 🏢 Highlights Kinerja per BA")

            # pilih periode (sama seperti highlights utama)
            st.selectbox(
                "Pilih Periode",
                options=all_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}",
                key="selected_period"
            )

            df = st.session_state.data_storage.get(st.session_state.selected_period)

            if df is None or df.empty:
                st.warning("Data IKPA belum tersedia.")
                st.stop()

            df = df.copy()

            # ===============================
            # NORMALISASI KODE BA
            # ===============================
            if "Kode BA" in df.columns:
                df["Kode BA"] = df["Kode BA"].apply(normalize_kode_ba)
            else:
                st.warning("Kolom Kode BA tidak tersedia.")
                st.stop()

            # ===============================
            # FILTER BA (PAKAI FILTER YANG SAMA)
            # ===============================
            df = apply_filter_ba(df)

            # ===============================
            # HITUNG RATA-RATA IKPA PER BA
            # ===============================
            nilai_col = "Nilai Akhir (Nilai Total/Konversi Bobot)"
            df[nilai_col] = pd.to_numeric(df[nilai_col], errors="coerce")

            df_ba = (
                df.groupby("Kode BA")[nilai_col]
                .mean()
                .reset_index(name="Rata-rata IKPA")
                .dropna()
            )

            # ===============================
            # MAP NAMA BA
            # ===============================
            df_ba["Nama BA"] = df_ba["Kode BA"].map(BA_MAP)
            df_ba["Label BA"] = (
                df_ba["Kode BA"] + " – " +
                df_ba["Nama BA"].fillna("Nama BA tidak ditemukan")
            )

            # ===============================
            # SORT DESC (TINGGI → RENDAH)
            # ===============================
            df_ba = df_ba.sort_values("Rata-rata IKPA", ascending=False)

            if df_ba.empty:
                st.info("Tidak ada data BA yang dapat ditampilkan.")
                st.stop()

            # ===============================
            # CHART VERTIKAL (KHUSUS BA)
            # ===============================
            fig_ba = px.bar(
                df_ba,
                x="Label BA",
                y="Rata-rata IKPA",
                color="Rata-rata IKPA",
                color_continuous_scale="Blues",
                text="Rata-rata IKPA"
            )

            fig_ba.update_traces(
                texttemplate="%{text:.2f}",
                textposition="outside"
            )

            fig_ba.update_layout(
                title="🏢 Rata-rata Nilai IKPA per BA",
                xaxis_title="BA",
                yaxis_title="Rata-rata Nilai IKPA",
                height=600,
                xaxis_tickangle=-45,
                coloraxis_showscale=False,
                margin=dict(l=40, r=20, t=80, b=160)
            )

            st.plotly_chart(fig_ba, use_container_width=True)

            # =================================================
            # 🚨 BA dengan Deviasi Halaman III DIPA Bermasalah
            # (Rata-rata < 90)
            # =================================================
            st.subheader("🚨 BA yang Memerlukan Perhatian Khusus")
            st.markdown("###### Rata-rata Deviasi Halaman III DIPA < 90")

            # -----------------------------
            # Slider Skala Y
            # -----------------------------
            col_min_dev, col_max_dev = st.columns(2)

            with col_min_dev:
                st.markdown("**Nilai Minimum (Y-Axis)**")
                y_min_dev = st.slider(
                    "",
                    min_value=0,
                    max_value=50,
                    value=40,
                    step=1,
                    key="ba_dev_ymin"
                )

            with col_max_dev:
                st.markdown("**Nilai Maksimum (Y-Axis)**")
                y_max_dev = st.slider(
                    "",
                    min_value=51,
                    max_value=110,
                    value=110,
                    step=1,
                    key="ba_dev_ymax"
                )

            # -----------------------------
            # Hitung Rata-rata Deviasi per BA
            # -----------------------------
            problem_col = "Deviasi Halaman III DIPA"

            df[problem_col] = pd.to_numeric(
                df[problem_col],
                errors="coerce"
            )

            df_ba_dev = (
                df.groupby("Kode BA")[problem_col]
                .mean()
                .reset_index(name="Rata-rata Deviasi Halaman III DIPA")
                .dropna()
            )

            # -----------------------------
            # Filter BA Bermasalah (< 90)
            # -----------------------------
            df_ba_problem = df_ba_dev[
                df_ba_dev["Rata-rata Deviasi Halaman III DIPA"] < 90
            ].copy()

            # -----------------------------
            # Map Nama BA & Label
            # -----------------------------
            df_ba_problem["Nama BA"] = df_ba_problem["Kode BA"].map(BA_MAP)
            df_ba_problem["Label BA"] = (
                df_ba_problem["Kode BA"]
                + " – "
                + df_ba_problem["Nama BA"].fillna("Nama BA tidak ditemukan")
            )

            df_ba_problem = df_ba_problem.sort_values(
                "Rata-rata Deviasi Halaman III DIPA",
                ascending=False
            )

            # -----------------------------
            # Render Chart
            # -----------------------------
            if df_ba_problem.empty:
                st.success("✅ Seluruh BA sudah optimal (rata-rata Deviasi ≥ 90).")
            else:
                fig_ba_dev = px.bar(
                    df_ba_problem,
                    x="Label BA",
                    y="Rata-rata Deviasi Halaman III DIPA",
                    color="Rata-rata Deviasi Halaman III DIPA",
                    color_continuous_scale="YlOrRd",
                    text="Rata-rata Deviasi Halaman III DIPA"
                )

                fig_ba_dev.update_traces(
                    texttemplate="%{text:.2f}",
                    textposition="outside"
                )

                fig_ba_dev.update_layout(
                    title="🚨 BA dengan Rata-rata Deviasi Halaman III DIPA < 90",
                    xaxis_title="BA",
                    yaxis_title="Rata-rata Deviasi Halaman III DIPA",
                    height=600,
                    xaxis_tickangle=-45,
                    yaxis=dict(range=[y_min_dev, y_max_dev]),
                    coloraxis_showscale=False,
                    margin=dict(l=40, r=20, t=80, b=160)
                )

                st.plotly_chart(fig_ba_dev, use_container_width=True)


        # -------------------------
        # DATA DETAIL SATKER
        # -------------------------
        elif main_tab == "📋 Data Detail Satker":
            st.markdown("## 📋 Tabel Detail Satker")

            # ===============================
            # 🔎 AMBIL FILTER KODE BA (DARI DASHBOARD UTAMA)
            # ===============================
            selected_ba = st.session_state.get("filter_ba_main", None)

            # persistent sub-tab for Periodik / Detail Satker
            if "active_table_tab" not in st.session_state:
                st.session_state.active_table_tab = "📆 Periodik"

            sub_tab = st.radio(
                "Pilih Mode Tabel",
                ["📆 Periodik", "📋 Detail Satker"],
                key="sub_tab_choice",
                horizontal=True
            )
            st.session_state['active_table_tab'] = sub_tab

            # -------------------------
            # PERIODIK TABLE
            # -------------------------
            if sub_tab == "📆 Periodik":
                st.markdown("#### Periodik — ringkasan per bulan / triwulan / perbandingan")

                # Tentukan tahun yang tersedia
                years = set()
                for k, df_period in st.session_state.data_storage.items():
                    years.update(df_period['Tahun'].astype(str).unique())
                years = sorted([int(y) for y in years if str(y).strip() != ''], reverse=True)

                if not years:
                    st.info("Tidak ada data periodik untuk ditampilkan.")
                    st.stop()

                default_year = years[0]
                selected_year = st.selectbox("Pilih Tahun", options=years, index=0, key='tab_periodik_year_select')

                # session state untuk period_type
                if "period_type" not in st.session_state:
                    st.session_state.period_type = "quarterly"

                period_options = ["quarterly", "monthly", "compare"]
                try:
                    period_index = period_options.index(st.session_state.period_type)
                except ValueError:
                    period_index = 0
                    st.session_state.period_type = "quarterly"

                # Radio button
                period_type = st.radio(
                    "Jenis Periode",
                    options=period_options,
                    format_func=lambda x: {"quarterly": "Triwulan", "monthly": "Bulanan", "compare": "Perbandingan"}.get(x, x),
                    horizontal=True,
                    index=period_index,
                    key="period_type_radio_v2"
                )
                st.session_state.period_type = period_type

                # Pilih indikator (satu untuk semua mode)
                indicator_options = [
                    'Kualitas Perencanaan Anggaran', 'Kualitas Pelaksanaan Anggaran', 'Kualitas Hasil Pelaksanaan Anggaran',
                    'Revisi DIPA', 'Deviasi Halaman III DIPA', 'Penyerapan Anggaran', 'Belanja Kontraktual',
                    'Penyelesaian Tagihan', 'Pengelolaan UP dan TUP', 'Capaian Output', 'Dispensasi SPM (Pengurang)',
                    'Nilai Akhir (Nilai Total/Konversi Bobot)'
                ]
                default_indicator = 'Deviasi Halaman III DIPA'
                selected_indicator = st.selectbox(
                    "Pilih Indikator", 
                    options=indicator_options, 
                    index=indicator_options.index(default_indicator) if default_indicator in indicator_options else 0,
                    key='tab_periodik_indicator_select'
                )
                
                # -------------------------
                # Monthly / Quarterly
                # -------------------------
                if period_type in ['monthly', 'quarterly']:

                    # 1. Gabungkan data per tahun
                    dfs = []
                    for (mon, yr), df_period in st.session_state.data_storage.items():
                        try:
                            if int(yr) == int(selected_year):
                                temp = df_period.copy()

                                # ambil kolom bulan apa pun namanya
                                if 'Bulan' in temp.columns:
                                    temp['Bulan_raw'] = temp['Bulan']
                                elif 'Nama Bulan' in temp.columns:
                                    temp['Bulan_raw'] = temp['Nama Bulan']
                                else:
                                    continue

                                dfs.append(temp)
                        except:
                            continue

                    if not dfs:
                        st.info(f"Tidak ditemukan data untuk tahun {selected_year}.")
                        st.stop()

                    df_year = pd.concat(dfs, ignore_index=True)
                    
                    # ===============================
                    # 2. NORMALISASI KODE BA
                    # ===============================
                    if 'Kode BA' in df_year.columns:
                        df_year['Kode BA'] = df_year['Kode BA'].apply(normalize_kode_ba)

                    # ===============================
                    # 3. APPLY FILTER BA (GLOBAL)
                    # ===============================
                    df_year = apply_filter_ba(df_year)

                    # =========================================================
                    # 2. Normalisasi bulan (SUPER DEFENSIVE)
                    # =========================================================
                    MONTH_FIX = {
                        "JAN": "JANUARI", "JANUARI": "JANUARI",
                        "FEB": "FEBRUARI", "FEBRUARI": "FEBRUARI",
                        "MAR": "MARET", "MRT": "MARET", "MARET": "MARET",
                        "APR": "APRIL", "APRIL": "APRIL",
                        "MEI": "MEI",
                        "JUN": "JUNI", "JUNI": "JUNI",
                        "JUL": "JULI", "JULI": "JULI",
                        "AGT": "AGUSTUS", "AGUSTUS": "AGUSTUS",
                        "SEP": "SEPTEMBER", "SEPTEMBER": "SEPTEMBER",
                        "OKT": "OKTOBER", "OKTOBER": "OKTOBER",
                        "DES": "DESEMBER", "DESEMBER": "DESEMBER"
                    }

                    df_year['Bulan_upper'] = (
                        df_year['Bulan_raw']
                        .astype(str)
                        .str.upper()
                        .str.strip()
                        .map(lambda x: MONTH_FIX.get(x, x))
                    )

                    # =========================================================
                    # 3. Period Column & Order (INI KUNCI)
                    # =========================================================
                    if period_type == 'monthly':
                        df_year['Period_Column'] = df_year['Bulan_upper']
                        df_year['Period_Order'] = df_year['Bulan_upper'].map(MONTH_ORDER)

                    else:  # quarterly
                        def to_quarter(m):
                            return {
                                'MARET': 'Tw I',
                                'JUNI': 'Tw II',
                                'SEPTEMBER': 'Tw III',
                                'DESEMBER': 'Tw IV'
                            }.get(m)

                        quarter_order = {'Tw I':1,'Tw II':2,'Tw III':3,'Tw IV':4}
                        df_year['Period_Column'] = df_year['Bulan_upper'].map(to_quarter)
                        df_year['Period_Order'] = df_year['Period_Column'].map(quarter_order)

                    # =========================================================
                    # 4. PIVOT LANGSUNG (FIXED)
                    # =========================================================

                    # --- pilih nama SATKER TERPENDEK per Kode Satker ---
                    name_map = (
                        df_year
                        .assign(name_len=df_year['Uraian Satker-RINGKAS'].astype(str).str.len())
                        .sort_values('name_len')
                        .groupby('Kode Satker')['Uraian Satker-RINGKAS']
                        .first()
                    )

                    df_pivot = df_year[
                        [
                            'Kode BA',
                            'Kode Satker',
                            'Period_Column',
                            selected_indicator
                        ]
                    ].copy()

                    df_wide = (
                        df_pivot
                        .pivot_table(
                            index=['Kode BA','Kode Satker'],  # ❗ IDENTIFIER ONLY
                            columns='Period_Column',
                            values=selected_indicator,
                            aggfunc='last'
                        )
                        .reset_index()
                    )

                    # --- pasang kembali nama satker ---
                    df_wide['Uraian Satker-RINGKAS'] = df_wide['Kode Satker'].map(name_map)


                    # =========================================================
                    # 5. URUTKAN KOLOM PERIODE 
                    # =========================================================
                    if period_type == 'monthly':
                        ordered_periods = sorted(
                            [c for c in df_wide.columns if c in MONTH_ORDER],
                            key=lambda x: MONTH_ORDER[x]
                        )
                    else:
                        ordered_periods = [
                            c for c in ['Tw I', 'Tw II', 'Tw III', 'Tw IV']
                            if c in df_wide.columns
                        ]


                    # =========================================================
                    # 6. RANKING PERIODIK (BERDASARKAN PERIODE TERAKHIR)
                    # =========================================================
                    if ordered_periods:
                        # pastikan numerik
                        for c in ordered_periods:
                            df_wide[c] = pd.to_numeric(df_wide[c], errors='coerce')

                        # kolom periode TERAKHIR (AMAN: BELUM DI-RENAME)
                        last_col = ordered_periods[-1]

                        # nilai acuan ranking
                        df_wide['Latest_Value'] = df_wide[last_col]

                        # dense ranking (nilai sama = peringkat sama)
                        df_wide['Peringkat'] = (
                            df_wide['Latest_Value']
                            .rank(method='dense', ascending=False)
                            .astype('Int64')
                        )


                    # =========================================================
                    # 7. SUSUN KOLOM DASAR
                    # =========================================================
                    fixed_cols = [
                        "Uraian Satker-RINGKAS",
                        "Peringkat",
                        "Kode BA",
                        "Kode Satker"
                    ]

                    month_cols = [c for c in ordered_periods]
                    df_wide = df_wide[fixed_cols + month_cols]


                    # =========================================================
                    # 8. DISPLAY DATAFRAME
                    # =========================================================
                    df_display = df_wide.copy()

                    # --- format peringkat ---
                    df_display['Peringkat'] = (
                        pd.to_numeric(df_display['Peringkat'], errors='coerce')
                        .astype('Int64')
                        .astype(str)
                    )

                    # --- format kode ---
                    df_display['Kode BA'] = (
                        df_display['Kode BA']
                        .astype(str)
                        .str.replace(r'\.0$', '', regex=True)
                    )

                    df_display['Kode Satker'] = (
                        df_display['Kode Satker']
                        .astype(str)
                        .str.replace(r'\.0$', '', regex=True)
                        .str.zfill(6)
                    )


                    # =========================================================
                    # 9. RENAME BULAN (KHUSUS DISPLAY, AMAN)
                    # =========================================================
                    for c in ordered_periods:
                        if c in df_display.columns:
                            df_display[c] = df_display[c].fillna("–")

                    # --- rename hanya untuk display (SETELAH NaN) ---
                    if period_type == 'monthly':
                        rename_map = {
                            c: MONTH_ABBR.get(c.upper(), c)
                            for c in ordered_periods
                            if c in df_display.columns
                        }
                        df_display.rename(columns=rename_map, inplace=True)



                    # =========================================================
                    # 10. SEARCH
                    # =========================================================
                    search_query = st.text_input(
                        "🔎 Cari (Periodik) – ketik untuk filter di semua kolom",
                        value="",
                        key="tab_periodik_search"
                    )

                    if search_query:
                        q = search_query.strip().lower()
                        mask = df_display.apply(
                            lambda row: row.astype(str).str.lower().str.contains(q, na=False).any(),
                            axis=1
                        )
                        df_display_filtered = df_display[mask].copy()
                    else:
                        df_display_filtered = df_display.copy()


                    # =========================================================
                    # 11. FINAL SORT (WAJIB SEBELUM AGGRID)
                    # =========================================================
                    df_display_filtered["_rank_num"] = pd.to_numeric(
                        df_display_filtered["Peringkat"], errors="coerce"
                    )

                    df_display_filtered = (
                        df_display_filtered
                        .sort_values(
                            by=["_rank_num", "Uraian Satker-RINGKAS"],
                            ascending=[True, True]
                        )
                        .drop(columns="_rank_num")
                        .reset_index(drop=True)
                    )

                    # DATA BULAN JULI DIAMANKAN
                    if "Jul" in df_display_filtered.columns and "Jun" in df_display_filtered.columns:
                        df_display_filtered["Jul"] = (
                            df_display_filtered["Jul"]
                            .replace("–", pd.NA)
                            .fillna(df_display_filtered["Jun"])
                        )
                        
                    # =========================================================
                    # 12. RENDER
                    # =========================================================
                    render_table_pin_satker(df_display_filtered)


                elif period_type == "compare":
                    st.markdown("### 📊 Perbandingan Antara Dua Tahun")

                    # ===============================
                    # 1. GABUNGKAN SELURUH DATA
                    # ===============================
                    all_data = []

                    for (mon, yr), df in st.session_state.data_storage.items():
                        df2 = df.copy()

                        df2["Bulan_upper"] = (
                            df2["Bulan"]
                            .astype(str)
                            .str.upper()
                            .str.strip()
                        )

                        df2["Tahun"] = pd.to_numeric(
                            df2["Tahun"], errors="coerce"
                        ).astype("Int64")

                        if "Kode BA" in df2.columns:
                            df2["Kode BA"] = df2["Kode BA"].apply(normalize_kode_ba)

                        all_data.append(df2)

                    if not all_data:
                        st.warning("Belum ada data IKPA.")
                        st.stop()

                    df_full = pd.concat(all_data, ignore_index=True)

                    # ===============================
                    # 2. FILTER BA VALID (SESUI HIGHLIGHTS)
                    # ===============================
                    latest_period = max(
                        st.session_state.data_storage.keys(),
                        key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0], 0))
                    )

                    df_latest = st.session_state.data_storage[latest_period].copy()
                    df_latest["Kode BA"] = df_latest["Kode BA"].apply(normalize_kode_ba)
                    df_full["Kode BA"] = df_full["Kode BA"].apply(normalize_kode_ba)

                    valid_ba = df_latest["Kode BA"].dropna().unique()
                    df_full = df_full[df_full["Kode BA"].isin(valid_ba)]

                    # ===============================
                    # 3. FILTER BA KHUSUS COMPARE
                    # ===============================
                    if "filter_ba_compare" not in st.session_state:
                        st.session_state["filter_ba_compare"] = ["SEMUA BA"]

                    ba_list = sorted(df_full["Kode BA"].dropna().unique())
                    ba_options = ["SEMUA BA"] + ba_list

                    def format_ba_compare(code):
                        if code == "SEMUA BA":
                            return "SEMUA BA"
                        return f"{code} – {BA_MAP.get(code, 'Nama BA tidak ditemukan')}"

                    selected_ba_compare = st.multiselect(
                        "Pilih Kode BA (Perbandingan)",
                        options=ba_options,
                        format_func=format_ba_compare,
                        key="filter_ba_compare"   # ✅ cukup ini
                    )


                    # ===============================
                    # 4. VALIDASI TAHUN
                    # ===============================
                    available_years = sorted(
                        df_full["Tahun"].dropna().unique().astype(int)
                    )

                    if len(available_years) < 2:
                        st.info("Data tahun tidak cukup untuk perbandingan.")
                        st.stop()

                    colA, colB = st.columns(2)
                    with colA:
                        year_a = st.selectbox(
                            "Tahun A (Awal)",
                            available_years,
                            index=0,
                            key="tahunA_compare"
                        )
                    with colB:
                        year_b = st.selectbox(
                            "Tahun B (Akhir)",
                            available_years,
                            index=1,
                            key="tahunB_compare"
                        )

                    if year_a == year_b:
                        st.info("Pilih dua tahun yang berbeda.")
                        st.stop()

                    # ===============================
                    # 5. FILTER DATA PER TAHUN
                    # ===============================
                    df_a = df_full[df_full["Tahun"] == year_a]
                    df_b = df_full[df_full["Tahun"] == year_b]

                    def extract_tw(df_):
                        return {
                            "Tw I": df_[df_["Bulan_upper"] == "MARET"],
                            "Tw II": df_[df_["Bulan_upper"] == "JUNI"],
                            "Tw III": df_[df_["Bulan_upper"] == "SEPTEMBER"],
                            "Tw IV": df_[df_["Bulan_upper"] == "DESEMBER"],
                        }

                    tw_a = extract_tw(df_a)
                    tw_b = extract_tw(df_b)

                    # ===============================
                    # 6. PILIH SATKER
                    # ===============================
                    satker_list = (
                        df_full[["Kode Satker", "Uraian Satker-RINGKAS"]]
                        .dropna(subset=["Kode Satker"])
                        .copy()
                    )

                    # 🔥 CLEAN DATA
                    satker_list["Kode Satker"] = (
                        satker_list["Kode Satker"]
                        .astype(str)
                        .str.strip()
                    )

                    satker_list["Uraian Satker-RINGKAS"] = (
                        satker_list["Uraian Satker-RINGKAS"]
                        .fillna("TANPA NAMA")
                        .astype(str)
                    )

                    satker_list = satker_list.drop_duplicates().sort_values("Uraian Satker-RINGKAS")

                    # 🔥 AMBIL OPTIONS AMAN
                    satker_options = ["SEMUA SATKER"] + satker_list["Kode Satker"].tolist()

                    # 🔥 BUAT MAPPING BIAR AMAN
                    satker_map = dict(
                        zip(satker_list["Kode Satker"], satker_list["Uraian Satker-RINGKAS"])
                    )

                    selected_satkers = st.multiselect(
                        "Pilih Satker",
                        options=satker_options,
                        default=["SEMUA SATKER"],
                        key="satker_compare",
                        format_func=lambda x: (
                            "SEMUA SATKER"
                            if x == "SEMUA SATKER"
                            else satker_map.get(x, f"SATKER {x}")  # 🔥 anti error
                        )
                    )

                    selected_satkers_final = (
                        satker_list["Kode Satker"].tolist()
                        if "SEMUA SATKER" in selected_satkers
                        else selected_satkers
                    )

                    # ===============================
                    # 7. BANGUN TABEL PERBANDINGAN
                    # ===============================
                    rows = []

                    for _, m in satker_list.iterrows():
                        kode = m["Kode Satker"]
                        if kode not in selected_satkers_final:
                            continue

                        row = {
                            "Kode Satker": kode,
                            "Uraian Satker-RINGKAS": m["Uraian Satker-RINGKAS"]
                        }

                        latest_a, latest_b = None, None
                        has_data = False

                        for tw in ["Tw I", "Tw II", "Tw III", "Tw IV"]:

                            if selected_indicator not in tw_a[tw].columns:
                                continue

                            valA = tw_a[tw].loc[
                                tw_a[tw]["Kode Satker"] == kode,
                                selected_indicator
                            ].values

                            valB = tw_b[tw].loc[
                                tw_b[tw]["Kode Satker"] == kode,
                                selected_indicator
                            ].values

                            valA = valA[0] if len(valA) else None
                            valB = valB[0] if len(valB) else None

                            row[f"{tw} {year_a}"] = valA
                            row[f"{tw} {year_b}"] = valB

                            if valA is not None:
                                latest_a = valA
                                has_data = True
                            if valB is not None:
                                latest_b = valB
                                has_data = True

                        if not has_data:
                            continue

                        row[f"Δ Total ({year_b}-{year_a})"] = (
                            latest_b - latest_a
                            if latest_a is not None and latest_b is not None
                            else None
                        )

                        rows.append(row)

                    if not rows:
                        st.info("Tidak ada data indikator untuk periode yang dipilih.")
                        st.stop()

                    df_compare = pd.DataFrame(rows)

                    # ===============================
                    # 8. TAMPILKAN HASIL (AGGRID)
                    # ===============================
                    st.markdown("### 📋 Hasil Perbandingan")
                    render_table_pin_satker(df_compare)


            # -------------------------
            # DETAIL SATKER 
            # -------------------------
            else:
                st.subheader("📋 Detail Satker")

                # ===============================
                # VALIDASI DATA
                # ===============================
                if not st.session_state.get("data_storage"):
                    st.warning("⚠️ Belum ada data IKPA.")
                    return

                # ===============================
                # PILIH TAHUN & BULAN
                # ===============================
                available_periods = list(st.session_state.data_storage.keys())
                available_years = sorted({int(y) for (_, y) in available_periods}, reverse=True)

                selected_year = st.selectbox(
                    "Pilih Tahun",
                    options=available_years,
                    index=0,
                    key="detail_satker_year"
                )

                months_for_year = [
                    m for (m, y) in available_periods if int(y) == selected_year
                ]

                months_for_year = [
                    VALID_MONTHS.get(m.upper(), m.upper())
                    for (m, y) in available_periods
                    if int(y) == selected_year
                ]

                if not months_for_year:
                    st.info(f"Tidak ada data untuk tahun {selected_year}.")
                    return

                months_for_year = sorted(
                    set(months_for_year),
                    key=lambda m: MONTH_ORDER.get(m, 0)
                )


                selected_month = st.selectbox(
                    "Pilih Bulan",
                    options=months_for_year,
                    format_func=lambda x: x.capitalize(),
                    index=len(months_for_year) - 1,
                    key="detail_satker_month"
                )

                selected_key = (selected_month, str(selected_year))
                df = st.session_state.data_storage.get(selected_key)

                if df is None or df.empty:
                    st.info("Data detail satker tidak tersedia.")
                    return

                df = df.copy()

                # ===============================
                # NORMALISASI & FILTER BA
                # ===============================
                if "Kode BA" in df.columns:
                    df["Kode BA"] = df["Kode BA"].apply(normalize_kode_ba)

                df = apply_filter_ba(df)

                if df.empty:
                    st.info("Tidak ada data sesuai filter BA.")
                    return

                # ===============================
                # MODE TAMPILAN
                # ===============================
                view_mode = st.radio(
                    "Tampilan",
                    options=["aspek", "komponen"],
                    format_func=lambda x: "Berdasarkan Aspek" if x == "aspek" else "Berdasarkan Komponen",
                    horizontal=True,
                    key="detail_view_mode"
                )

                # ===============================
                # KOLOM IDENTITAS (WAJIB)
                # ===============================
                base_cols = [
                    "Uraian Satker-RINGKAS",
                    "Kode Satker",
                    "Peringkat",
                    "Kode BA"
                ]

                if view_mode == "aspek":
                    value_cols = [
                        "Kualitas Perencanaan Anggaran",
                        "Kualitas Pelaksanaan Anggaran",
                        "Kualitas Hasil Pelaksanaan Anggaran",
                        "Nilai Total",
                        "Dispensasi SPM (Pengurang)",
                        "Nilai Akhir (Nilai Total/Konversi Bobot)"
                    ]

                    cols_needed = base_cols + value_cols

                    for col in cols_needed:
                        if col not in df.columns:
                            df[col] = 0  # 🔥 anti KeyError

                    df_display = df[cols_needed].copy()

                else:
                    component_cols = [
                        "Revisi DIPA",
                        "Deviasi Halaman III DIPA",
                        "Penyerapan Anggaran",
                        "Belanja Kontraktual",
                        "Penyelesaian Tagihan",
                        "Pengelolaan UP dan TUP",
                        "Capaian Output"
                    ]

                    value_cols = (
                        component_cols +
                        [
                            "NIlai Total",
                            "Dispensasi SPM (Pengurang)",
                            "Nilai Akhir (Nilai Total/Konversi Bobot)"
                        ]
                    )

                    for c in component_cols:
                        if c not in df.columns:
                            df[c] = 0

                    cols_exist = [c for c in (base_cols + value_cols) if c in df.columns]
                    df_display = df[cols_exist].copy()


                # ===============================
                # SEARCH
                # ===============================
                search_query = st.text_input(
                    "🔎 Cari (ketik untuk filter di semua kolom)",
                    value="",
                    key="detail_satker_search"
                )

                if search_query:
                    q = search_query.lower()
                    mask = df_display.apply(
                        lambda r: r.astype(str).str.lower().str.contains(q, na=False).any(),
                        axis=1
                    )
                    df_display = df_display[mask].copy()

                # ===============================
                # FORMAT KODE
                # ===============================
                df_display["Kode Satker"] = (
                    df_display["Kode Satker"]
                    .astype(str)
                    .str.replace(r"\.0$", "", regex=True)
                    .str.zfill(6)
                )

                df_display["Kode BA"] = (
                    df_display["Kode BA"]
                    .astype(str)
                    .str.replace(r"\.0$", "", regex=True)
                )

                # ===============================
                # TAMPILKAN DENGAN AGGRID
                # ===============================
                render_table_pin_satker(df_display)
            
    elif st.session_state.main_menu == "Digitalisasi":
        st.markdown("## 📊 Digitalisasi")

        menu_digital = st.segmented_control(
            label="",
            options=["📈 Chart Utama", "📋 Tabel Detail"],
            default="📈 Chart Utama"
        )

        st.divider()

        # =====================================================
        # CHART UTAMA
        # =====================================================
        if menu_digital == "📈 Chart Utama":

            st.markdown("## 📊 Chart Utama")

            if "digipay_master" not in st.session_state or "kkp_master" not in st.session_state:
                st.warning("Data Digipay atau KKP belum tersedia")
                st.stop()

            df_digipay = st.session_state.digipay_master.copy()
            df_kkp = st.session_state.kkp_master.copy()
            
            # ===============================
            # FORMAT ANGKA INDONESIA
            # ===============================
            def format_rupiah(x):
                return f"{int(x):,}".replace(",", ".")
            
            # ===============================
            # NORMALISASI KODE SATKER KKP
            # ===============================
            df_kkp["Kode Satker"] = (
                df_kkp["Kode Satker"]
                .astype(str)
                .str.extract(r'(\d+)')[0]   # ambil angka saja
                .str.zfill(6)               # paksa 6 digit
            )

            # ===============================
            # MERGE NAMA SATKER RINGKAS
            # ===============================
            ref = st.session_state.reference_df.copy()

            df_digipay["KDSATKER"] = (
                df_digipay["KDSATKER"]
                .astype(str)
                .str.extract(r'(\d+)')[0]
                .str.zfill(6)
            )

            df_kkp["Kode Satker"] = (
                df_kkp["Kode Satker"]
                .astype(str)
                .str.extract(r'(\d+)')[0]
                .str.zfill(6)
            )

            df_digipay = df_digipay.merge(
                ref[["Kode Satker", "Uraian Satker-SINGKAT"]],
                left_on="KDSATKER",
                right_on="Kode Satker",
                how="left"
            )

            df_kkp = df_kkp.merge(
                ref[["Kode Satker", "Uraian Satker-SINGKAT"]],
                on="Kode Satker",
                how="left"
            )

            # Gunakan nama satker ringkas
            df_digipay["SATKER"] = df_digipay["Uraian Satker-SINGKAT"].fillna(df_digipay["NMSATKER"])
            df_kkp["SATKER"] = df_kkp["Uraian Satker-SINGKAT"].fillna(df_kkp["SATKER"])

            # ===============================
            # BERSIHKAN NAMA SATKER (HAPUS KODE DI DEPAN)
            # ===============================
            df_digipay["SATKER"] = df_digipay["SATKER"].astype(str).str.replace(r'^\d+\s*', '', regex=True)
            df_kkp["SATKER"] = df_kkp["SATKER"].astype(str).str.replace(r'^\d+\s*', '', regex=True)
            
            # ===============================
            # LABEL SATKER UNTUK CHART
            # ===============================
            df_digipay["SATKER"] = df_digipay["KDSATKER"] + " - " + df_digipay["SATKER"]
            df_kkp["SATKER"] = df_kkp["Kode Satker"] + " - " + df_kkp["SATKER"]

            # ===============================
            # NORMALISASI DIGIPAY
            # ===============================
            df_digipay["TAHUN"] = pd.to_numeric(df_digipay["TAHUN"], errors="coerce")
            df_digipay["BULAN"] = pd.to_numeric(df_digipay["BULAN"], errors="coerce")
            df_digipay["TRIWULAN"] = ((df_digipay["BULAN"] - 1) // 3) + 1

            # Bersihkan nominal
            df_digipay["NOMINVOICE"] = (
                df_digipay["NOMINVOICE"]
                .astype(str)
                .str.replace(r"[^\d]", "", regex=True)
            )

            df_digipay["NOMINVOICE"] = pd.to_numeric(
                df_digipay["NOMINVOICE"], errors="coerce"
            ).fillna(0)

            # ===============================
            # FILTER UI
            # ===============================
            col1, col2, col3 = st.columns(3)

            # ===============================
            # PERIODE
            # ===============================
            with col1:
                periode_chart = st.selectbox(
                    "Periode",
                    ["Bulanan", "Triwulan", "Tahunan"]
                )

            # ===============================
            # TAHUN TERBARU
            # ===============================
            tahun_list = sorted(df_digipay["TAHUN"].dropna().astype(int).unique())
            tahun_terbaru = max(tahun_list)

            with col2:
                tahun_chart = st.selectbox(
                    "Tahun",
                    tahun_list,
                    index=tahun_list.index(tahun_terbaru)
                )

            bulan_selected = None
            triwulan_selected = None

            bulan_map = {
                1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
                5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
                9: "September", 10: "Oktober", 11: "November", 12: "Desember"
            }

            # ===============================
            # BULAN TERBARU
            # ===============================
            if periode_chart == "Bulanan":

                bulan_list = sorted(df_digipay["BULAN"].dropna().astype(int).unique())
                bulan_terbaru = max(bulan_list)

                with col3:
                    bulan_selected = st.selectbox(
                        "Bulan",
                        bulan_list,
                        index=bulan_list.index(bulan_terbaru),
                        format_func=lambda x: bulan_map.get(x, x)
                    )

            # ===============================
            # TRIWULAN TERBARU
            # ===============================
            elif periode_chart == "Triwulan":

                tw_list = sorted(df_digipay["TRIWULAN"].dropna().astype(int).unique())
                tw_terbaru = max(tw_list)

                tw_options = ["TW1","TW2","TW3","TW4"]

                with col3:
                    triwulan_selected = st.selectbox(
                        "Triwulan",
                        tw_options,
                        index=tw_terbaru - 1
                    )

            # ===============================
            # TIPE DATA
            # ===============================
            tipe_chart = st.radio(
                "Tipe",
                ["Jumlah Transaksi", "Jumlah Nominal"],
                horizontal=True
            )

            # ===============================
            # FILTER DIGIPAY
            # ===============================
            if periode_chart == "Bulanan":

                df_digipay = df_digipay[
                    (df_digipay["TAHUN"] == tahun_chart) &
                    (df_digipay["BULAN"] <= bulan_selected)
                ]

            elif periode_chart == "Triwulan":

                tw = int(triwulan_selected.replace("TW", ""))

                df_digipay = df_digipay[
                    (df_digipay["TAHUN"] == tahun_chart) &
                    (df_digipay["TRIWULAN"] <= tw)
                ]

            else:

                df_digipay = df_digipay[
                    df_digipay["TAHUN"] <= tahun_chart
                ]

            # ===============================
            # DIGIPAY PER SATKER
            # ===============================
            if tipe_chart == "Jumlah Transaksi":

                digipay_chart = (
                    df_digipay
                    .groupby("SATKER")
                    .agg(Value=("NOINVOICE", "nunique"))
                    .reset_index()
                )

            else:

                digipay_chart = (
                    df_digipay
                    .groupby("SATKER")
                    .agg(Value=("NOMINVOICE", "sum"))
                    .reset_index()
                )

            # ===============================
            # TOP & BOTTOM DIGIPAY
            # ===============================
            digipay_top = digipay_chart.sort_values("Value", ascending=False).head(10)
            digipay_bottom = digipay_chart.sort_values("Value", ascending=True).head(10)

            digipay_top["LABEL"] = digipay_top["Value"].apply(format_rupiah)
            digipay_bottom["LABEL"] = digipay_bottom["Value"].apply(format_rupiah)

            digipay_top = digipay_top.reset_index(drop=True)
            digipay_top["Rank"] = digipay_top.index + 1

            digipay_bottom = digipay_bottom.reset_index(drop=True)
            digipay_bottom["Rank"] = digipay_bottom.index + 1

            col_left, col_right = st.columns(2)

            # ===============================
            # TOP DIGIPAY
            # ===============================
            with col_left:

                fig_digipay_top = px.bar(
                    digipay_top,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#08306B","#6BAED6"],
                    title=f"10 Satker dengan {tipe_chart} Terbesar (Digipay)"
                )

                fig_digipay_top.update_layout(
                    height=550,
                    yaxis={'categoryorder':'total ascending'},
                    coloraxis_showscale=False
                )

                fig_digipay_top.update_traces(textposition="outside")

                st.plotly_chart(fig_digipay_top, use_container_width=True)


            # ===============================
            # BOTTOM DIGIPAY
            # ===============================
            with col_right:

                fig_digipay_bottom = px.bar(
                    digipay_bottom,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#FEE0D2","#DE2D26"],
                    title=f"10 Satker dengan {tipe_chart} Terendah (Digipay)"
                )

                fig_digipay_bottom.update_layout(
                    height=550,
                    yaxis={'categoryorder':'total ascending'},
                    coloraxis_showscale=False
                )

                fig_digipay_bottom.update_traces(textposition="outside")

                st.plotly_chart(fig_digipay_bottom, use_container_width=True)
            
            # ===============================
            # CHART KKP
            # ===============================

            # ===============================
            # AMBIL DATA DARI SESSION
            # ===============================
            df_kkp = st.session_state.kkp_master.copy()

            # ===============================
            # NORMALISASI PERIODE
            # ===============================
            df_kkp["PERIODE"] = pd.to_datetime(df_kkp["PERIODE"], errors="coerce")

            df_kkp["TAHUN"] = df_kkp["PERIODE"].dt.year
            df_kkp["BULAN"] = df_kkp["PERIODE"].dt.month
            df_kkp["TRIWULAN"] = df_kkp["PERIODE"].dt.quarter


            # ===============================
            # NORMALISASI NOMINAL
            # ===============================
            df_kkp["NILAI TRANSAKSI (NILAI SPM)"] = (
                df_kkp["NILAI TRANSAKSI (NILAI SPM)"]
                .astype(str)
                .str.replace(r"[^\d]", "", regex=True)
            )

            df_kkp["NILAI TRANSAKSI (NILAI SPM)"] = pd.to_numeric(
                df_kkp["NILAI TRANSAKSI (NILAI SPM)"],
                errors="coerce"
            ).fillna(0)


            # ===============================
            # NORMALISASI LIMIT KKP
            # ===============================
            df_kkp["LIMIT KKP"] = (
                df_kkp["LIMIT KKP"]
                .astype(str)
                .str.replace(r"[^\d]", "", regex=True)
            )

            df_kkp["LIMIT KKP"] = pd.to_numeric(
                df_kkp["LIMIT KKP"],
                errors="coerce"
            ).fillna(0)


            # ===============================
            # FILTER PERIODE (KUMULATIF)
            # ===============================
            if periode_chart == "Bulanan":

                df_kkp = df_kkp[
                    (df_kkp["TAHUN"] == tahun_chart) &
                    (df_kkp["BULAN"] <= bulan_selected)
                ]

            elif periode_chart == "Triwulan":

                tw = int(triwulan_selected.replace("TW", ""))

                df_kkp = df_kkp[
                    (df_kkp["TAHUN"] == tahun_chart) &
                    (df_kkp["TRIWULAN"] <= tw)
                ]

            else:

                df_kkp = df_kkp[
                    df_kkp["TAHUN"] <= tahun_chart
                ]


            # ===============================
            # MERGE NAMA SATKER RINGKAS (KKP)
            # ===============================

            ref = st.session_state.reference_df.copy()

            # normalisasi kode satker KKP
            df_kkp["Kode Satker"] = (
                df_kkp["Kode Satker"]
                .astype(str)
                .str.extract(r'(\d+)')[0]
                .str.zfill(6)
            )

            # normalisasi kode satker reference
            ref["Kode Satker"] = (
                ref["Kode Satker"]
                .astype(str)
                .str.extract(r'(\d+)')[0]
                .str.zfill(6)
            )

            # merge nama satker ringkas
            df_kkp = df_kkp.merge(
                ref[["Kode Satker", "Uraian Satker-SINGKAT"]],
                on="Kode Satker",
                how="left"
            )

            # gunakan nama satker ringkas
            df_kkp["SATKER"] = (
                df_kkp["Uraian Satker-SINGKAT"]
                .fillna(df_kkp["SATKER"])
            )

            # pastikan tidak ada null
            df_kkp["SATKER"] = df_kkp["SATKER"].fillna("SATKER TIDAK DIKETAHUI")

            # label untuk chart
            df_kkp["SATKER"] = (
                df_kkp["Kode Satker"].astype(str)
                + " "
                + df_kkp["SATKER"].astype(str)
            )
                    

            # ===============================
            # PAGU PER SATKER
            # ===============================
            pagu_satker = (
                df_kkp
                .groupby("SATKER")["LIMIT KKP"]
                .sum()
                .reset_index(name="PAGU")
            )


            # ===============================
            # KKP PER SATKER
            # ===============================
            if tipe_chart == "Jumlah Transaksi":

                kkp_chart = (
                    df_kkp
                    .groupby("SATKER")
                    .size()
                    .reset_index(name="Value")
                )

            else:

                nominal_satker = (
                    df_kkp
                    .groupby("SATKER")
                    .agg(NOMINAL=("NILAI TRANSAKSI (NILAI SPM)", "sum"))
                    .reset_index()
                )

                kkp_chart = nominal_satker.merge(
                    pagu_satker,
                    on="SATKER",
                    how="left"
                )

                kkp_chart["Value"] = (
                    kkp_chart["NOMINAL"] /
                    kkp_chart["PAGU"].replace(0, pd.NA)
                ) * 100

                kkp_chart["Value"] = kkp_chart["Value"].fillna(0)


            # ===============================
            # TOP & BOTTOM
            # ===============================
            kkp_top = kkp_chart.sort_values("Value", ascending=False).head(10)

            if tipe_chart == "Jumlah Transaksi":

                kkp_bottom = (
                    kkp_chart
                    .sort_values("Value", ascending=True)
                    .head(10)
                )

            else:

                kkp_bottom = (
                    kkp_chart[kkp_chart["PAGU"] > 0]
                    .sort_values("Value", ascending=True)
                    .head(10)
                )


            # ===============================
            # LABEL
            # ===============================
            if tipe_chart == "Jumlah Transaksi":

                kkp_top["LABEL"] = kkp_top["Value"].astype(int)
                kkp_bottom["LABEL"] = kkp_bottom["Value"].astype(int)

                title_top = "10 Satker dengan Jumlah Transaksi KKP Terbesar"
                title_bottom = "10 Satker dengan Jumlah Transaksi KKP Terendah"

            else:

                kkp_top["LABEL"] = kkp_top["Value"].apply(lambda x: f"{x:.2f}%")
                kkp_bottom["LABEL"] = kkp_bottom["Value"].apply(lambda x: f"{x:.2f}%")

                title_top = "10 Satker dengan % Realisasi KKP Tertinggi"
                title_bottom = "10 Satker dengan % Realisasi KKP Terendah"


            kkp_top = kkp_top.reset_index(drop=True)
            kkp_top["Rank"] = kkp_top.index + 1

            kkp_bottom = kkp_bottom.reset_index(drop=True)
            kkp_bottom["Rank"] = kkp_bottom.index + 1


            # ===============================
            # CHART
            # ===============================
            col_left, col_right = st.columns(2)


            # ===============================
            # TOP CHART
            # ===============================
            with col_left:

                fig_kkp_top = px.bar(
                    kkp_top,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#00441B","#74C476"],
                    title=title_top
                )

                fig_kkp_top.update_layout(
                    height=550,
                    yaxis={'categoryorder':'total ascending'},
                    coloraxis_showscale=False
                )

                fig_kkp_top.update_traces(textposition="outside")

                st.plotly_chart(fig_kkp_top, use_container_width=True)


            # ===============================
            # BOTTOM CHART
            # ===============================
            with col_right:

                fig_kkp_bottom = px.bar(
                    kkp_bottom,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#FEE0D2","#DE2D26"],
                    title=title_bottom
                )

                fig_kkp_bottom.update_layout(
                    height=550,
                    yaxis={'categoryorder':'total ascending'},
                    coloraxis_showscale=False
                )

                fig_kkp_bottom.update_traces(textposition="outside")

                st.plotly_chart(fig_kkp_bottom, use_container_width=True)

           
            
            # =====================================================
            # CMS CHART
            # =====================================================
            df_cms = st.session_state.cms_master.copy()

            # ===============================
            # NORMALISASI TAHUN
            # ===============================
            df_cms["TAHUN"] = pd.to_numeric(df_cms["TAHUN"], errors="coerce")

            # ===============================
            # NORMALISASI TRIWULAN
            # ===============================
            df_cms["TRIWULAN"] = (
                df_cms["TRIWULAN"]
                .astype(str)
                .str.replace("TW","", regex=False)
            )

            df_cms["TRIWULAN"] = pd.to_numeric(df_cms["TRIWULAN"], errors="coerce")

            # ===============================
            # FILTER CMS
            # ===============================
            if periode_chart == "Bulanan":

                tw = (bulan_selected - 1) // 3 + 1

                df_cms = df_cms[
                    (df_cms["TAHUN"] == tahun_chart) &
                    (df_cms["TRIWULAN"] <= tw)
                ]

            elif periode_chart == "Triwulan":

                tw = int(triwulan_selected.replace("TW",""))

                df_cms = df_cms[
                    (df_cms["TAHUN"] == tahun_chart) &
                    (df_cms["TRIWULAN"] <= tw)
                ]

            else:

                df_cms = df_cms[
                    df_cms["TAHUN"] <= tahun_chart
                ]


            # ===============================
            # DETEKSI KOLOM SATKER
            # ===============================
            if "Kode Satker" in df_cms.columns:
                satker_code_col = "Kode Satker"
            elif "KODE SATKER" in df_cms.columns:
                satker_code_col = "KODE SATKER"
            else:
                st.error("Kolom kode satker CMS tidak ditemukan")
                st.stop()

            if "Nama Satker" in df_cms.columns:
                satker_name_col = "Nama Satker"
            elif "NAMA SATKER" in df_cms.columns:
                satker_name_col = "NAMA SATKER"
            else:
                st.error("Kolom nama satker CMS tidak ditemukan")
                st.stop()


            # ===============================
            # MERGE SATKER RINGKAS
            # ===============================
            ref = st.session_state.reference_df.copy()

            df_cms[satker_code_col] = (
                df_cms[satker_code_col]
                .astype(str)
                .str.extract(r'(\d+)')[0]
                .str.zfill(6)
            )

            df_cms = df_cms.merge(
                ref[["Kode Satker","Uraian Satker-SINGKAT"]],
                left_on=satker_code_col,
                right_on="Kode Satker",
                how="left"
            )

            df_cms["SATKER"] = df_cms["Uraian Satker-SINGKAT"].fillna(df_cms[satker_name_col])
            # ===============================
            # LABEL SATKER CMS
            # ===============================
            df_cms["SATKER"] = df_cms[satker_code_col] + " - " + df_cms["SATKER"]


            # ===============================
            # NORMALISASI NUMERIC
            # ===============================
            cols = [
                "JUMLAH TRANSAKSI CMS",
                "JUMLAH TRANSAKSI KARTU DEBIT",
                "JUMLAH TRANSAKSI TELLER",
                "NILAI TRANSAKSI CMS",
                "NILAI TRANSAKSI KARTU DEBIT",
                "NILAI TRANSAKSI TELLER",
            ]

            for c in cols:
                df_cms[c] = pd.to_numeric(df_cms[c], errors="coerce").fillna(0)


            # ===============================
            # AGREGASI SATKER
            # ===============================
            cms_satker = (
                df_cms
                .groupby("SATKER")
                .agg(
                    CMS_TRX=("JUMLAH TRANSAKSI CMS","sum"),
                    DEBIT_TRX=("JUMLAH TRANSAKSI KARTU DEBIT","sum"),
                    TELLER_TRX=("JUMLAH TRANSAKSI TELLER","sum"),

                    CMS_NOM=("NILAI TRANSAKSI CMS","sum"),
                    DEBIT_NOM=("NILAI TRANSAKSI KARTU DEBIT","sum"),
                    TELLER_NOM=("NILAI TRANSAKSI TELLER","sum"),
                )
                .reset_index()
            )


            # ===============================
            # TOTAL
            # ===============================
            cms_satker["TOTAL_TRX"] = (
                cms_satker["CMS_TRX"]
                + cms_satker["DEBIT_TRX"]
                + cms_satker["TELLER_TRX"]
            )

            cms_satker["TOTAL_NOM"] = (
                cms_satker["CMS_NOM"]
                + cms_satker["DEBIT_NOM"]
                + cms_satker["TELLER_NOM"]
            )


            # ===============================
            # PROPORSI
            # ===============================
            cms_satker["PROPORSI_TRX"] = (
                cms_satker["CMS_TRX"] / cms_satker["TOTAL_TRX"]
            ) * 100

            cms_satker["PROPORSI_NOM"] = (
                cms_satker["CMS_NOM"] / cms_satker["TOTAL_NOM"]
            ) * 100

            cms_satker = cms_satker.fillna(0)


            # ===============================
            # TIPE CHART
            # ===============================
            if tipe_chart == "Jumlah Transaksi":

                cms_chart = cms_satker[["SATKER","PROPORSI_TRX"]].copy()
                cms_chart.rename(columns={"PROPORSI_TRX":"Value"}, inplace=True)

                title_chart = "10 Satker dengan Proporsi Transaksi CMS Tertinggi"

            else:

                cms_chart = cms_satker[["SATKER","PROPORSI_NOM"]].copy()
                cms_chart.rename(columns={"PROPORSI_NOM":"Value"}, inplace=True)

                title_chart = "10 Satker dengan Proporsi Nominal CMS Tertinggi"


            # ===============================
            # BULATKAN NILAI
            # ===============================
            cms_chart["Value"] = cms_chart["Value"].round(2)

            # ===============================
            # FORMAT LABEL PERSEN
            # ===============================
            cms_chart["LABEL"] = cms_chart["Value"].apply(
                lambda x: "100%" if round(x,2) == 100 else f"{x:.2f}%"
            )

            # ===============================
            # TOP & BOTTOM CMS
            # ===============================
            cms_top = cms_chart.sort_values("Value", ascending=False).head(10)
            cms_bottom = cms_chart.sort_values("Value", ascending=True).head(10)

            cms_top = cms_top.reset_index(drop=True)
            cms_top["Rank"] = cms_top.index + 1

            cms_bottom = cms_bottom.reset_index(drop=True)
            cms_bottom["Rank"] = cms_bottom.index + 1

            col_left, col_right = st.columns(2)

            # ===============================
            # TOP CMS
            # ===============================
            with col_left:

                fig_cms_top = px.bar(
                    cms_top,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#3F007D","#BCBDDC"],
                    title="10 Satker dengan Proporsi CMS Tertinggi"
                )

                fig_cms_top.update_layout(
                    height=520,
                    yaxis={'categoryorder':'total ascending'},
                    xaxis=dict(range=[0,105]),
                    margin=dict(r=80),
                    coloraxis_showscale=False
                )

                fig_cms_top.update_traces(
                    textposition="outside",
                    cliponaxis=False,
                    hovertemplate="<b>%{y}</b><br>Proporsi: %{x:.2f}%<extra></extra>"
                )

                st.plotly_chart(fig_cms_top, use_container_width=True)


            # ===============================
            # BOTTOM CMS
            # ===============================
            with col_right:

                fig_cms_bottom = px.bar(
                    cms_bottom,
                    x="Value",
                    y="SATKER",
                    orientation="h",
                    text="LABEL",
                    color="Rank",
                    color_continuous_scale=["#FEE0D2","#DE2D26"],
                    title="10 Satker dengan Proporsi CMS Terendah"
                )

                fig_cms_bottom.update_layout(
                    height=520,
                    yaxis={'categoryorder':'total ascending'},
                    xaxis=dict(range=[0,105]),
                    margin=dict(r=80),
                    coloraxis_showscale=False
                )

                fig_cms_bottom.update_traces(
                    textposition="outside",
                    cliponaxis=False,
                    hovertemplate="<b>%{y}</b><br>Proporsi: %{x:.2f}%<extra></extra>"
                )

                st.plotly_chart(fig_cms_bottom, use_container_width=True)

        # =====================================================
        # TABEL DETAIL
        # =====================================================
        elif menu_digital == "📋 Tabel Detail": 

            st.subheader("📋 Tabel Detail Digitalisasi")

            source_detail = st.radio(
                "",
                ["💰 Digipay", "💳 KKP", "🏦 CMS"],
                horizontal=True
            )

            st.divider()
            

            # =====================================================
            # DIGIPAY
            # =====================================================
            if source_detail == "💰 Digipay":

                if "digipay_master" not in st.session_state:
                    st.warning("Data Digipay belum tersedia")
                    render_table_pin_satker(pd.DataFrame())
                else:

                    df_master = st.session_state.digipay_master.copy()

                    # =====================================================
                    # 1️⃣ DETEKSI KOLOM KODE SATKER
                    # =====================================================
                    possible_kode_cols = [
                        "KODE_SATKER",
                        "KODE SATKER",
                        "Kode Satker",
                        "KDSATKER"
                    ]

                    kode_col = next((c for c in possible_kode_cols if c in df_master.columns), None)

                    if kode_col is None:
                        st.error("Kolom kode satker tidak ditemukan di Digipay")
                        render_table_pin_satker(pd.DataFrame())
                    else:

                        df_master["Kode Satker"] = (
                            df_master[kode_col]
                            .astype(str)
                            .str.extract(r"(\d{6})")[0]
                        )

                        # =====================================================
                        # 2️⃣ MERGE REFERENSI
                        # =====================================================
                        if "reference_df" in st.session_state:

                            ref = st.session_state.reference_df.copy()
                            ref["Kode Satker"] = ref["Kode Satker"].astype(str).str.strip()

                            df_master = df_master.merge(
                                ref[["Kode Satker", "Uraian Satker-SINGKAT"]],
                                on="Kode Satker",
                                how="left"
                            )

                            if "NMSATKER" in df_master.columns:
                                df_master["Nama Satker"] = (
                                    df_master["Uraian Satker-SINGKAT"]
                                    .fillna(df_master["NMSATKER"])
                                )
                            else:
                                df_master["Nama Satker"] = df_master["Uraian Satker-SINGKAT"]

                        else:
                            df_master["Nama Satker"] = df_master.get("NMSATKER", df_master["Kode Satker"])

                        # =====================================================
                        # 3️⃣ NORMALISASI DATA
                        # =====================================================
                        if "TANGGAL" in df_master.columns:

                            df_master["TANGGAL"] = pd.to_datetime(df_master["TANGGAL"], errors="coerce")
                            df_master["TAHUN"] = df_master["TANGGAL"].dt.year
                            df_master["BULAN"] = df_master["TANGGAL"].dt.month

                        else:

                            if "TAHUN" in df_master.columns:
                                df_master["TAHUN"] = pd.to_numeric(df_master["TAHUN"], errors="coerce")

                            if "BULAN" in df_master.columns:
                                df_master["BULAN"] = pd.to_numeric(df_master["BULAN"], errors="coerce")

                        # Bersihkan nominal
                        if "NOMINVOICE" in df_master.columns:

                            df_master["NOMINVOICE"] = (
                                df_master["NOMINVOICE"]
                                .astype(str)
                                .str.replace(r"[^\d]", "", regex=True)
                                .replace("", "0")
                                .astype(float)
                            )

                        # =====================================================
                        # FILTER
                        # =====================================================
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            periode = st.selectbox("Periode", ["Bulanan", "Triwulan", "Tahunan"])

                        with col2:
                            tipe = st.selectbox("Tipe", ["Jumlah Transaksi", "Nilai Transaksi"])

                        if periode != "Tahunan":

                            tahun_list = sorted(df_master["TAHUN"].dropna().unique())
                            latest_year = max(tahun_list) if len(tahun_list) else None

                            with col3:
                                tahun = st.selectbox(
                                    "Tahun",
                                    tahun_list if len(tahun_list) else [None],
                                    index=tahun_list.index(latest_year) if latest_year in tahun_list else 0
                                )

                            if tahun is not None:
                                df_raw = df_master[df_master["TAHUN"] == tahun].copy()
                            else:
                                df_raw = pd.DataFrame()

                        else:
                            df_raw = df_master.copy()

                        # =====================================================
                        # BULANAN
                        # =====================================================
                        if periode == "Bulanan":

                            MONTH_MAP = {
                                1:"JAN",2:"FEB",3:"MAR",4:"APR",
                                5:"MEI",6:"JUN",7:"JUL",8:"AGU",
                                9:"SEP",10:"OKT",11:"NOV",12:"DES"
                            }

                            df_raw["BULAN_NAMA"] = df_raw.get("BULAN", pd.Series()).map(MONTH_MAP)

                            if tipe == "Jumlah Transaksi":

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","BULAN_NAMA"])["NOINVOICE"]
                                    .nunique()
                                    .reset_index(name="Jumlah")
                                )

                                value_col = "Jumlah"

                            else:

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","BULAN_NAMA"])["NOMINVOICE"]
                                    .sum()
                                    .reset_index(name="Nilai")
                                )

                                value_col = "Nilai"

                            df_pivot = df_grouped.pivot_table(
                                index=["Kode Satker","Nama Satker"],
                                columns="BULAN_NAMA",
                                values=value_col,
                                fill_value=0
                            )

                            urutan_bulan = ["JAN","FEB","MAR","APR","MEI","JUN","JUL","AGU","SEP","OKT","NOV","DES"]
                            df_pivot = df_pivot.reindex(columns=urutan_bulan, fill_value=0)

                        # =====================================================
                        # TRIWULAN
                        # =====================================================
                        elif periode == "Triwulan":

                            df_raw["TRIWULAN"] = ((df_raw["BULAN"] - 1) // 3) + 1

                            if tipe == "Jumlah Transaksi":

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","TRIWULAN"])["NOINVOICE"]
                                    .nunique()
                                    .reset_index(name="Jumlah")
                                )

                                value_col = "Jumlah"

                            else:

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","TRIWULAN"])["NOMINVOICE"]
                                    .sum()
                                    .reset_index(name="Nilai")
                                )

                                value_col = "Nilai"

                            df_pivot = df_grouped.pivot_table(
                                index=["Kode Satker","Nama Satker"],
                                columns="TRIWULAN",
                                values=value_col,
                                fill_value=0
                            )

                            df_pivot = df_pivot.reindex(columns=[1,2,3,4], fill_value=0)
                            df_pivot.columns = ["TW1","TW2","TW3","TW4"]

                        # =====================================================
                        # TAHUNAN
                        # =====================================================
                        else:

                            if tipe == "Jumlah Transaksi":

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","TAHUN"])["NOINVOICE"]
                                    .nunique()
                                    .reset_index(name="Jumlah")
                                )

                                value_col = "Jumlah"

                            else:

                                df_grouped = (
                                    df_raw.groupby(["Kode Satker","Nama Satker","TAHUN"])["NOMINVOICE"]
                                    .sum()
                                    .reset_index(name="Nilai")
                                )

                                value_col = "Nilai"

                            df_pivot = (
                                df_grouped
                                .pivot_table(
                                    index=["Kode Satker","Nama Satker"],
                                    columns="TAHUN",
                                    values=value_col,
                                    fill_value=0
                                )
                                .sort_index(axis=1)
                            )

                            df_pivot.columns = df_pivot.columns.astype(str)

                        df_pivot = df_pivot.reset_index()

                        # =====================================================
                        # FORMAT RIBUAN
                        # =====================================================
                        def format_ribuan(x):
                            try:
                                return "{:,.0f}".format(float(x)).replace(",", ".")
                            except:
                                return x

                        for col in df_pivot.columns:
                            if col not in ["Kode Satker","Nama Satker"]:
                                df_pivot[col] = df_pivot[col].apply(format_ribuan)

                        render_table_pin_satker(df_pivot)
                    

            # =====================================================
            # 💳 KKP
            # =====================================================
            elif source_detail == "💳 KKP":

                if "kkp_master" not in st.session_state:
                    st.warning("Data KKP belum tersedia")

                else:

                    df_master = st.session_state.kkp_master.copy()

                    # =============================
                    # NORMALISASI PERIODE
                    # =============================
                    if "PERIODE" in df_master.columns:

                        df_master["PERIODE"] = df_master["PERIODE"].astype(str)

                        df_master["TAHUN"] = df_master["PERIODE"].str[:4].astype(int)
                        df_master["BULAN"] = df_master["PERIODE"].str[5:7].astype(int)

                    else:
                        st.error("Kolom PERIODE tidak ditemukan pada data KKP")
                        st.stop()

                    col1, col2, col3 = st.columns(3)

                    with col1:
                        periode = st.selectbox(
                            "Periode",
                            ["Bulanan", "Triwulan", "Tahunan"],
                            key="kkp_periode"
                        )

                    with col2:
                        tipe = st.selectbox(
                            "Tipe",
                            ["Jumlah Nominal", "Jumlah Transaksi"],
                            key="kkp_tipe"
                        )

                    if periode != "Tahunan":

                        with col3:
                            tahun_list = sorted(df_master["TAHUN"].dropna().unique())
                            latest_year = max(tahun_list)

                            tahun = st.selectbox(
                                "Tahun",
                                tahun_list,
                                index=tahun_list.index(latest_year),
                                key="kkp_tahun"
                            )

                    else:
                        tahun = None

                        
                    df_pivot = generate_kkp_from_session(
                        df_master,
                        periode=periode,
                        tipe=tipe,
                        tahun_filter=tahun
                    )

                    # normalisasi kode satker
                    df_pivot["Kode Satker"] = df_pivot["Kode Satker"].astype(str).str.zfill(6)

                    # tambah pagu
                    df_pivot = add_kkp_pagu_column(df_pivot, df_master)

                    # persen hanya untuk nominal
                    if tipe == "Jumlah Nominal":
                        df_pivot = add_kkp_percentage_columns(df_pivot, df_master)

                    # =============================
                    # FORMAT RIBUAN
                    # =============================
                    def format_ribuan(x):

                        try:
                            return "{:,.0f}".format(float(x)).replace(",", ".")

                        except:
                            return x


                    for col in df_pivot.columns:
    
                        if "% Realisasi KKP" in col:
                            df_pivot[col] = df_pivot[col].apply(lambda x: f"{float(x):.2f}%")

                        elif col not in ["SATKER","Uraian Satker-RINGKAS","Kode Satker"]:
                            df_pivot[col] = df_pivot[col].apply(format_ribuan)

                    # =============================
                    # HILANGKAN KODE SATKER DI DEPAN NAMA
                    # =============================
                    if "SATKER" in df_pivot.columns:
                        df_pivot["SATKER"] = df_pivot["SATKER"].astype(str).apply(
                            lambda x: re.sub(r"^\d{6}\s*", "", x)
                        )
                    
                    render_table_pin_satker(df_pivot)
                                

            # =====================================================
            # 🏦 CMS
            # =====================================================
            elif source_detail == "🏦 CMS":

                if "cms_master" not in st.session_state:
                    st.warning("Data CMS belum tersedia")
                    st.stop()

                df_master = st.session_state.cms_master.copy()
                
                # =============================
                # DETEKSI PERIODE TERBARU CMS
                # =============================
                tw_order = {"TW1":1,"TW2":2,"TW3":3,"TW4":4}

                df_tmp = df_master.copy()
                df_tmp["TW_ORDER"] = df_tmp["TRIWULAN"].map(tw_order)

                latest_row = (
                    df_tmp
                    .sort_values(["TAHUN","TW_ORDER"], ascending=[False,False])
                    .iloc[0]
                )

                latest_year = latest_row["TAHUN"]
                latest_tw = latest_row["TRIWULAN"]

                col1, col2, col3 = st.columns(3)

                with col1:
                    periode = st.selectbox(
                        "Periode",
                        ["Triwulan", "Tahunan"],
                        key="cms_periode"
                    )

                tahun_list = sorted(df_master["TAHUN"].dropna().unique())

                with col2:
                    tahun = st.selectbox(
                        "Tahun",
                        tahun_list,
                        index=tahun_list.index(latest_year),
                        key="cms_tahun"
                    )

                triwulan_selected = None

                if periode == "Triwulan":
                    with col3:
                        tw_list = ["TW1","TW2","TW3","TW4"]
                        triwulan_selected = st.selectbox(
                            "Triwulan",
                            tw_list,
                            index=tw_list.index(latest_tw),
                            key="cms_triwulan"
                        )

                # =============================
                # FILTER TAHUN
                # =============================
                df_master = df_master[df_master["TAHUN"] == tahun]

                # =============================
                # FILTER TRIWULAN
                # =============================
                if triwulan_selected:
                    df_master = df_master[df_master["TRIWULAN"] == triwulan_selected]

                df_pivot = pd.DataFrame()

                try:
                    df_pivot = generate_cms_from_session(
                        df_master,
                        periode=periode,
                        tahun_filter=tahun
                    )
                except Exception as e:
                    st.error(f"Gagal memproses CMS: {e}")
                    st.stop()

                if df_pivot is None or df_pivot.empty:
                    st.warning("Data CMS tidak tersedia untuk periode yang dipilih")
                    st.stop()

                # =============================
                # FORMAT PERSEN
                # =============================
                percent_cols = [
                    "Proporsi Transaksi CMS",
                    "Proporsi Nominal CMS"
                ]

                for col in percent_cols:
                    if col in df_pivot.columns:
                        df_pivot[col] = df_pivot[col].round(2).astype(str) + " %"

                render_table_pin_satker(df_pivot)
                                    
        

# HALAMAN 2: DASHBOARD INTERNAL KPPN (Protected)  
def menu_ews_satker():
    st.subheader("🏛️ Early Warning System Kinerja Keuangan Satker")

    if "data_storage" not in st.session_state or not st.session_state.data_storage:
        st.warning("⚠️ Belum ada data historis yang tersedia.")
        return
    
    # Gabungkan semua data
    all_data = []
    for period, df in st.session_state.data_storage.items():
        df_copy = df.copy()
        # ensure Period & Period_Sort exist
        df_copy['Period'] = f"{period[0]} {period[1]}"
        df_copy['Period_Sort'] = f"{period[1]}-{period[0]}"
        all_data.append(df_copy)
    
    if not all_data:
        st.warning("⚠️ Belum ada data historis yang tersedia.")
        return
    
    df_all = pd.concat(all_data, ignore_index=True)
    
    # Analisis tren dan Early Warning System
    # Gunakan data periode terkini
    latest_period = sorted(st.session_state.data_storage.keys(), key=lambda x: (int(x[1]), MONTH_ORDER.get(x[0].upper(), 0)), reverse=True)[0]
    df_latest = st.session_state.data_storage[latest_period]
    
    # ===============================
    # 🔑 NORMALISASI KODE BA (WAJIB)
    # ===============================
    df_all["Kode BA"] = df_all["Kode BA"].apply(normalize_kode_ba)
    df_latest["Kode BA"] = df_latest["Kode BA"].apply(normalize_kode_ba)
    
    # ===============================
    # LOAD & MAP BA
    # ===============================
    df_ref_ba = load_reference_ba()
    BA_MAP = get_ba_map(df_ref_ba)

    st.markdown('<div class="filter-ba">', unsafe_allow_html=True)
    st.markdown("🔎 Filter Kode BA")
    # 🔒 HANYA BA YANG ADA DI REFERENSI
    ba_codes = sorted(
        [ba for ba in df_all["Kode BA"].dropna().unique() if ba in BA_MAP]
    )
    ba_options = ["SEMUA BA"] + ba_codes

    def format_ba(code):
        if code == "SEMUA BA":
            return "SEMUA BA"
        return f"{code} – {BA_MAP.get(code, 'Nama BA tidak ditemukan')}"

    selected_ba_internal = st.multiselect(
        "Pilih Kode BA",
        options=ba_options,
        default=["SEMUA BA"],
        format_func=format_ba,
        key="filter_ba_internal"
    )

    # ===============================
    # TERAPKAN FILTER BA (GLOBAL INTERNAL)
    # ===============================
    if "SEMUA BA" not in selected_ba_internal:
        df_all = df_all[df_all["Kode BA"].isin(selected_ba_internal)]
        df_latest = df_latest[df_latest["Kode BA"].isin(selected_ba_internal)]

        
    # ===============================
    # 🔑 BUAT LABEL SATKER INTERNAL
    # ===============================
    df_latest = df_latest.copy()

    if "Kode BA" not in df_latest.columns:
        df_latest["Kode BA"] = ""

    df_latest["Kode BA"] = df_latest["Kode BA"].apply(normalize_kode_ba)

    df_latest["Satker_Internal"] = (
        "[" + df_latest["Kode BA"] + "] "
        + df_latest["Uraian Satker-RINGKAS"].astype(str)
        + " (" + df_latest["Kode Satker"].astype(str) + ")"
    )


    st.markdown("---")
    st.subheader("🚨 Satker yang Memerlukan Perhatian Khusus")

    # 🎚️ Pengaturan Sumbu Y
    st.markdown("###### Atur Skala Nilai (Sumbu Y)")
    col_min, col_max = st.columns(2)
    with col_min:
        y_min_int = st.slider(
            "Nilai Minimum (Y-Axis)",
            min_value=0,
            max_value=50,
            value=50,
            step=1,
            key="ymin_internal"
        )
    with col_max:
        y_max_int = st.slider(
            "Nilai Maksimum (Y-Axis)",
            min_value=51,
            max_value=110,
            value=110,
            step=1,
            key="ymax_internal"
        )

    # 📊 Highlights Kinerja Satker yang Perlu Perhatian Khusus
    # ===============================
    # 🔧 ADAPTIVE LAYOUT (INTERNAL)
    # ===============================
    df_tmp = df_latest.copy()
    df_tmp["Satker"] = df_tmp["Satker_Internal"]

    n_up = len(df_tmp[df_tmp['Pengelolaan UP dan TUP'] < 100])

    if n_up >= 15:
        col1, col2 = st.columns([3.5, 1.2])
    elif n_up >= 8:
        col1, col2 = st.columns([3, 1.5])
    else:
        col1, col2 = st.columns([2.5, 1.5])
        
    # ===============================
    # 🔧 SHARED HEIGHT (BIAR TIDAK TINGGI SEBELAH)
    # ===============================
    n_out = len(df_tmp[df_tmp['Capaian Output'] < 100])

    BAR_HEIGHT = 38
    BASE_HEIGHT = 260
    MAX_HEIGHT = 1200

    # ===============================
    # 🔧 SOFT SHARED HEIGHT (NORMAL)
    # ===============================
    MAX_VISIBLE_ITEMS = 10  # ⬅️ kunci utama

    effective_items = min(max(n_up, n_out), MAX_VISIBLE_ITEMS)

    shared_height = BASE_HEIGHT + (effective_items * BAR_HEIGHT)
    shared_height = min(max(shared_height, 380), 700)



    # ======================================================
    # 🔥 FIX NUMERIC WAJIB (ANTI STRING ERROR)
    # ======================================================
    def ensure_numeric(df, cols):
        df = df.copy()
        for col in cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_numeric)
        return df


    # ======================================================
    # 🔥 FILTER AMAN (<100)
    # ======================================================
    def get_problem(df, col):
        if col not in df.columns:
            return pd.DataFrame()

        df = df.copy()
        df[col] = df[col].apply(clean_numeric)

        return df[df[col] < 100]


    # ======================================================
    # 🔥 PREP DATA
    # ======================================================
    df_latest = ensure_numeric(
        df_latest,
        ["Pengelolaan UP dan TUP", "Capaian Output"]
    )

    df_latest["Satker"] = df_latest["Satker_Internal"]


    # ======================================================
    # 🔥 HITUNG JUMLAH DATA
    # ======================================================
    n_up = len(get_problem(df_latest, "Pengelolaan UP dan TUP"))
    n_out = len(get_problem(df_latest, "Capaian Output"))


    # ======================================================
    # 🔥 LAYOUT DINAMIS
    # ======================================================
    if n_up >= 15:
        col1, col2 = st.columns([3.5, 1.2])
    elif n_up >= 8:
        col1, col2 = st.columns([3, 1.5])
    else:
        col1, col2 = st.columns([2.5, 1.5])


    # ======================================================
    # 🔥 SHARED HEIGHT
    # ======================================================
    BAR_HEIGHT = 38
    BASE_HEIGHT = 260
    MAX_VISIBLE_ITEMS = 10

    effective_items = min(max(n_up, n_out), MAX_VISIBLE_ITEMS)

    shared_height = BASE_HEIGHT + (effective_items * BAR_HEIGHT)
    shared_height = min(max(shared_height, 380), 700)

    # 🔥 pastikan kolom numeric bersih
    col_up = "Pengelolaan UP dan TUP"

    if col_up in df.columns:
        df[col_up] = df[col_up].apply(clean_numeric).fillna(0)

    # ======================================================
    # 🔥 KOLOM KIRI — UP TUP
    # ======================================================
    with col1:

        st.markdown("""
        <div style="margin-bottom:6px;">
            <span style="font-size:16px; font-weight:600;">
                ⚠️ Pengelolaan UP dan TUP
            </span><br>
            <span style="font-size:13px; color:#666;">
                Pengelolaan UP dan TUP Belum Optimal (&lt; 100)
            </span>
        </div>
        """, unsafe_allow_html=True)

        df_problem_up = get_problem(df_latest, "Pengelolaan UP dan TUP")

        # 🔥 JIKA SUDAH OPTIMAL → TAMPILKAN NOTIF
        if df_problem_up.empty:

            st.markdown("""
            <div style='padding:16px;
                        background:#ecfdf5;
                        border-radius:12px;
                        color:#065f46;
                        font-weight:600;
                        text-align:center'>
            ✅ Pengelolaan UP/TUP seluruh satker sudah optimal (≥ 100)
            </div>
            """, unsafe_allow_html=True)

        else:

            # 🔥 FIX NUMERIC
            df_problem_up["Pengelolaan UP dan TUP"] = (
                df_problem_up["Pengelolaan UP dan TUP"]
                .apply(clean_numeric)
                .fillna(0)
            )

            y_min = safe_float(y_min_int, 0)
            y_max = safe_float(y_max_int, 110)

            if y_min >= y_max:
                y_min, y_max = 0, 110

            try:
                fig_up = create_internal_problem_chart_vertical(
                    df_problem_up,
                    column='Pengelolaan UP dan TUP',
                    threshold=100,
                    title="Pengelolaan UP dan TUP",
                    comparison='less',
                    show_yaxis=True,
                    show_colorbar=True,
                    fixed_height=shared_height
                )

                fig_up.update_layout(
                    yaxis=dict(range=[y_min, y_max])
                )

                st.plotly_chart(fig_up, use_container_width=True)

            except Exception as e:
                st.error(f"❌ Grafik UP/TUP error: {e}")


    # ======================================================
    # 🔥 KOLOM KANAN — CAPAIAN OUTPUT
    # ======================================================
    with col2:

        st.markdown("""
        <div style="margin-bottom:6px;">
            <span style="font-size:16px; font-weight:600;">
                ⚠️ Capaian Output
            </span><br>
            <span style="font-size:13px; color:#666;">
                Capaian Output Belum Optimal (&lt; 100)
            </span>
        </div>
        """, unsafe_allow_html=True)

        df_problem_out = get_problem(df_latest, "Capaian Output")

        if df_problem_out.empty:

            st.markdown("""
            <div style='padding:16px;
                        background:#ecfdf5;
                        border-radius:12px;
                        color:#065f46;
                        font-weight:600;
                        text-align:center'>
            ✅ Capaian Output seluruh satker sudah optimal (≥ 100)
            </div>
            """, unsafe_allow_html=True)

        else:

            # 🔥 FIX NUMERIC
            df_problem_out["Capaian Output"] = (
                df_problem_out["Capaian Output"]
                .apply(clean_numeric)
                .fillna(0)
            )

            y_min = safe_float(y_min_int, 0)
            y_max = safe_float(y_max_int, 110)

            if y_min >= y_max:
                y_min, y_max = 0, 110

            try:
                fig_output = create_internal_problem_chart_vertical(
                    df_problem_out,
                    column='Capaian Output',
                    threshold=100,
                    title="Capaian Output",
                    comparison='less',
                    show_yaxis=False,
                    show_colorbar=False,
                    fixed_height=shared_height
                )

                fig_output.update_layout(
                    yaxis=dict(range=[y_min, y_max])
                )

                st.plotly_chart(fig_output, use_container_width=True)

            except Exception as e:
                st.error(f"❌ Grafik Output error: {e}")


    warnings = []

    
    # ===============================
    # 📈 ANALISIS TREN
    # ===============================
    st.subheader("📈 Analisis Tren")

    # ======================================================
    # VALIDASI BULAN & TAHUN
    # ======================================================
    df_all["Month_Num"] = (
        df_all["Bulan"]
        .astype(str)
        .str.strip()
        .str.upper()
        .map(MONTH_ORDER)
    )

    if df_all["Month_Num"].isna().any():
        st.error("❌ Ditemukan nama bulan tidak valid.")
        st.stop()

    df_all["Tahun_Int"] = df_all["Tahun"].astype(int)
    df_all["Period_Sort"] = df_all.apply(
        lambda x: f"{x['Tahun_Int']:04d}-{x['Month_Num']:02d}",
        axis=1
    )

    # ======================================================
    # PILIH PERIODE & METRIK
    # ======================================================
    available_periods = sorted(df_all["Period_Sort"].unique())

    col1, col2, col3 = st.columns(3)

    with col1:
        start_period = st.selectbox("Periode Awal", available_periods, index=0)

    with col2:
        end_period = st.selectbox(
            "Periode Akhir",
            available_periods,
            index=len(available_periods) - 1
        )

    with col3:
        metric_options = [
            "Nilai Akhir (Nilai Total/Konversi Bobot)",
            "Kualitas Perencanaan Anggaran",
            "Kualitas Pelaksanaan Anggaran",
            "Kualitas Hasil Pelaksanaan Anggaran",
            "Revisi DIPA",
            "Deviasi Halaman III DIPA",
            "Penyerapan Anggaran",
            "Belanja Kontraktual",
            "Penyelesaian Tagihan",
            "Pengelolaan UP dan TUP",
            "Capaian Output",
        ]
        selected_metric = st.selectbox("Metrik yang Ditampilkan", metric_options)

    if start_period > end_period:
        st.warning("⚠️ Periode awal tidak boleh lebih besar dari periode akhir.")
        st.stop()

    # ======================================================
    # FILTER PERIODE
    # ======================================================
    df_trend = df_all[
        (df_all["Period_Sort"] >= start_period) &
        (df_all["Period_Sort"] <= end_period)
    ].copy()

    if df_trend.empty:
        st.warning("⚠️ Tidak ada data pada periode yang dipilih.")
        st.stop()

    # ======================================================
    # PAKSA SATKER RINGKAS
    # ======================================================
    df_trend = apply_reference_short_names(df_trend)

    df_trend = df_trend[
        df_trend["Uraian Satker-RINGKAS"].notna() &
        (df_trend["Uraian Satker-RINGKAS"].str.strip() != "")
    ].copy()

    df_trend["Nama_Satker_Ringkas"] = (
        df_trend["Uraian Satker-RINGKAS"]
        .astype(str)
        .str.strip()
    )

    # ======================================================
    # LABEL FINAL (UNTUK LEGEND)
    # ======================================================
    df_trend["Satker_Label_Final"] = (
        "[" + df_trend["Kode BA"] + "] "
        + df_trend["Nama_Satker_Ringkas"]
        + " (" + df_trend["Kode Satker"].astype(str) + ")"
    )

    satker_label_map = (
        df_trend[["Kode Satker", "Satker_Label_Final"]]
        .drop_duplicates("Kode Satker")
        .set_index("Kode Satker")["Satker_Label_Final"]
        .to_dict()
    )
    
    all_kode_satker = list(satker_label_map.keys())
    
    # ======================================================
    # 🔑 LABEL SELECTOR DARI DATA REFERENSI (SUMBER ASLI)
    # ======================================================
    ref_df = st.session_state.reference_df.copy()

    # pastikan kode satker string & rapi
    ref_df["Kode Satker"] = ref_df["Kode Satker"].astype(str).str.strip()
    ref_df["Uraian Satker-SINGKAT"] = ref_df["Uraian Satker-SINGKAT"].astype(str).str.strip()

    # hanya satker yang ADA di df_trend
    satker_short_label_map = {
        k: (
            ref_df.loc[
                ref_df["Kode Satker"] == k,
                "Uraian Satker-SINGKAT"
            ].iloc[0]
            + " (" + k + ")"
        )
        for k in satker_label_map.keys()
        if k in ref_df["Kode Satker"].values
    }


    # ======================================================
    # DEFAULT: 5 SATKER TERENDAH (PERIODE TERBARU)
    # ======================================================
    latest_period = df_all["Period_Sort"].max()
    df_latest = df_all[df_all["Period_Sort"] == latest_period].copy()

    bottom_5_kode = (
        df_latest
        .sort_values("Nilai Akhir (Nilai Total/Konversi Bobot)")
        .head(5)["Kode Satker"]
        .astype(str)
        .tolist()
    )

    default_kode = [k for k in bottom_5_kode if k in all_kode_satker]
    if not default_kode:
        default_kode = all_kode_satker[:5]

    # ======================================================
    # 🎯 PILIH SATKER (RINGKAS, SEARCHABLE, TETAP KODE)
    # ======================================================
    options_satker = list(satker_short_label_map.keys())

    safe_default = [k for k in default_kode if k in options_satker]

    # fallback kalau kosong
    if not safe_default:
        safe_default = options_satker[:5]

    selected_kode_satker = st.multiselect(
        label="Pilih Satker (Nama Ringkas)",
        options=options_satker,
        default=safe_default,
        format_func=lambda k: satker_short_label_map[k],
        placeholder="🔍 Ketik nama satker…",
        key="trend_satker_selector",
    )


    # ======================================================
    # 📊 GRAFIK TREN
    # ======================================================
    ordered_periods = (
        df_trend
        .sort_values("Period_Sort")["Period_Sort"]
        .unique()
        .tolist()
    )

    fig = go.Figure()

    for kode in selected_kode_satker:
        d = df_trend[df_trend["Kode Satker"] == kode].sort_values("Period_Sort")
        if d.empty:
            continue

        fig.add_trace(
            go.Scatter(
                x=d["Period_Sort"],
                y=d[selected_metric],
                mode="lines+markers",
                name=satker_short_label_map.get(kode, kode)
            )
        )


    fig.update_layout(
        title=dict(text=f"Tren {selected_metric}", x=0.5),
        xaxis=dict(
            title="Periode",
            categoryorder="array",
            categoryarray=ordered_periods
        ),
        yaxis_title="Nilai",
        height=750,
        hovermode="x unified",
        legend=dict(
            orientation="h",
            x=0.5,
            y=1.03,
            xanchor="center",
            yanchor="bottom",
            font=dict(size=10)
        ),
        margin=dict(l=60, r=40, t=170, b=60)
    )

    st.plotly_chart(fig, use_container_width=True)

    # ======================================================
    # 🚨 EARLY WARNING – TREN MENURUN
    # ======================================================
    warnings = []

    for kode in selected_kode_satker:
        d = (
            df_trend[df_trend["Kode Satker"] == kode]
            .sort_values("Period_Sort")
            .copy()
        )

        if len(d) < 2:
            continue

        prev = pd.to_numeric(d[selected_metric].iloc[-2], errors="coerce")
        last = pd.to_numeric(d[selected_metric].iloc[-1], errors="coerce")

        if pd.isna(prev) or pd.isna(last):
            continue

        if last < prev:
            warnings.append({
                "Satker": satker_short_label_map.get(kode, kode),
                "Sebelum": prev,
                "Terakhir": last,
                "Turun": prev - last
            })


    if warnings:
        st.warning(f"⚠️ Ditemukan {len(warnings)} satker dengan tren menurun!")
        for w in warnings:
            st.markdown(f"""
    **{w['Satker']}**  
    - Nilai sebelumnya: **{w['Sebelum']:.2f}**  
    - Nilai terkini: **{w['Terakhir']:.2f}**  
    - Penurunan: **{w['Turun']:.2f} poin**
    """)
            st.markdown("---")
    else:
        st.success("✅ Tidak ada satker dengan tren menurun.")
        

        
#HIGHLIGHTS
def menu_highlights():
    st.subheader("🎯 Highlights IKPA KPPN")

    # ===============================
    # VALIDASI DATA
    # ===============================
    if "data_storage_kppn" not in st.session_state or not st.session_state.data_storage_kppn:
        st.info("ℹ️ Belum ada data IKPA KPPN yang tersimpan.")
        return

    # ===============================
    # GABUNGKAN DATA IKPA KPPN
    # ===============================
    all_data = []
    for (bulan, tahun), df in st.session_state.data_storage_kppn.items():
        
        # 🔥 WAJIB: COPY DULU
        df_copy = df.copy()

        # 🔥 CLEAN KOLOM (ANTI ERROR)
        df_copy = df_copy.loc[:, ~df_copy.columns.duplicated()]
        df_copy.columns = (
            df_copy.columns.astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
        )

        # metadata
        df_copy["Periode"] = f"{bulan} {tahun}"
        df_copy["Tahun"] = int(tahun)
        df_copy["Bulan"] = bulan

        all_data.append(df_copy)

    df_all = pd.concat(all_data, ignore_index=True)

    # ===============================
    # NORMALISASI KOLOM
    # ===============================
    df_all.columns = (
        df_all.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    # ===============================
    # PERBAIKI UNNAMED (IKPA)
    # ===============================
    rename_map = {
        "Unnamed: 5": "Revisi DIPA",
        "Unnamed: 7": "Deviasi Halaman III DIPA",
        "Unnamed: 8": "Penyerapan Anggaran",
        "Unnamed: 9": "Belanja Kontraktual",
        "Unnamed: 10": "Penyelesaian Tagihan",
        "Unnamed: 11": "Pengelolaan UP dan TUP",
        "Unnamed: 12": "Capaian Output",
    }
    df_all = df_all.rename(columns=rename_map)

    # ===============================
    # 🔑 PASTIKAN PERIOD_SORT ADA
    # ===============================
    if "Period_Sort" not in df_all.columns:
        df_all["Month_Num"] = df_all["Bulan"].str.upper().map(MONTH_ORDER)
        df_all["Period_Sort"] = (
            df_all["Tahun"].astype(str)
            + "-"
            + df_all["Month_Num"].astype(int).astype(str).str.zfill(2)
        )

    # ===============================
    # 📅 FILTER PERIODE (BARU)
    # ===============================
    st.markdown("### 📅 Filter Periode")

    available_periods = sorted(df_all["Period_Sort"].dropna().unique())

    col1, col2 = st.columns(2)

    with col1:
        start_period = st.selectbox(
            "Periode Awal",
            options=available_periods,
            index=0
        )

    with col2:
        end_period = st.selectbox(
            "Periode Akhir",
            options=available_periods,
            index=len(available_periods) - 1
        )

    df_all = df_all[
        (df_all["Period_Sort"] >= start_period) &
        (df_all["Period_Sort"] <= end_period)
    ]

    if df_all.empty:
        st.warning("⚠️ Data kosong pada rentang periode tersebut.")
        return

    add_notification(f"Data IKPA KPPN dimuat ({len(df_all)} baris)")
    

    # ===============================
    # PILIH KPPN
    # ===============================
    kppn_list = sorted(df_all["Nama KPPN"].dropna().unique())
    selected_kppn = st.selectbox("Pilih KPPN", kppn_list)

    df_kppn = df_all[df_all["Nama KPPN"] == selected_kppn].copy()

    # ===============================
    # FILTER BARIS NILAI
    # ===============================
    if "Keterangan" in df_kppn.columns:
        df_kppn = df_kppn[
            df_kppn["Keterangan"].astype(str).str.upper() == "NILAI"
        ]

    # ===============================
    # INDIKATOR
    # ===============================
    indikator_opsi = [
        "Kualitas Perencanaan Anggaran",
        "Revisi DIPA",
        "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran",
        "Belanja Kontraktual",
        "Penyelesaian Tagihan",
        "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total",
        "Nilai Akhir (Nilai Total/Konversi Bobot)"
    ]

    for col in indikator_opsi:
        if col in df_kppn.columns:
            df_kppn[col] = (
                df_kppn[col]
                .astype(str)
                .str.replace("%", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            df_kppn[col] = pd.to_numeric(df_kppn[col], errors="coerce")

    selected_indikator = st.multiselect(
        "Pilih Indikator IKPA KPPN !",
        [c for c in indikator_opsi if c in df_kppn.columns],
        default=["Nilai Akhir (Nilai Total/Konversi Bobot)"]
    )

    if not selected_indikator:
        st.warning("⚠️ Pilih minimal satu indikator.")
        return

    # ===============================
    # URUT PERIODE
    # ===============================
    df_kppn = df_kppn.sort_values("Period_Sort")

    # ===============================
    # LINE CHART
    # ===============================
    fig = go.Figure()

    for indikator in selected_indikator:
        fig.add_trace(
            go.Scatter(
                x=df_kppn["Periode"],
                y=df_kppn[indikator],
                mode="lines+markers",
                name=indikator,
            )
        )

    fig.update_layout(
        title=f"📈 Tren IKPA KPPN – {selected_kppn}",
        xaxis_title="Periode",
        yaxis_title="Nilai",
        height=600,
        hovermode="x unified"
    )

    st.plotly_chart(fig, use_container_width=True)
    
    
    
def menu_tabel_ikpa_kppn():
    
    st.subheader("📊 Tabel IKPA KPPN")

    if "data_storage_kppn" not in st.session_state or not st.session_state.data_storage_kppn:
        st.warning("Data IKPA KPPN belum tersedia")
        return

    # ===============================
    # GABUNGKAN DATA
    # ===============================
    all_data = []

    for (bulan, tahun), df in st.session_state.data_storage_kppn.items():

        temp = df.copy()

        temp = temp.loc[:, ~temp.columns.duplicated()].copy()

        temp.columns = (
            temp.columns.astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
        )

        temp["Bulan"] = bulan
        temp["Tahun"] = int(tahun)

        all_data.append(temp)

    df_all = pd.concat(all_data, ignore_index=True)
    df_all = df_all.loc[:, ~df_all.columns.duplicated()]

    # ===============================
    # NORMALISASI KOLOM
    # ===============================
    rename_map = {
        "Halaman": "Deviasi Halaman III DIPA",
        "Angja": "Penyerapan Anggaran",
        "Kontrak": "Belanja Kontraktual",
        "Tagihan": "Penyelesaian Tagihan",
        "Output": "Capaian Output",
    }
    df_all = df_all.rename(columns=rename_map)

    # ===============================
    # PILIH INDIKATOR
    # ===============================
    indikator_list = [
        "Revisi DIPA",
        "Deviasi Halaman III DIPA",
        "Penyerapan Anggaran",
        "Belanja Kontraktual",
        "Penyelesaian Tagihan",
        "Pengelolaan UP dan TUP",
        "Capaian Output",
        "Nilai Total",
        "Konversi Bobot",
        "Dispensasi SPM (Pengurangan)",
        "Nilai Akhir (Nilai Total/Konversi Bobot)"
    ]

    default_indikator = "Deviasi Halaman III DIPA"

    indikator = st.selectbox(
        "Pilih Indikator IKPA KPPN",
        indikator_list,
        index=indikator_list.index(default_indikator)
    )

    if indikator not in df_all.columns:
        st.error(f"Kolom {indikator} tidak ditemukan")
        return

    # ===============================
    # CLEAN NUMERIC (AMAN)
    # ===============================
    df_all[indikator] = (
        df_all[indikator]
        .astype(str)
        .str.replace("%", "", regex=False)
        .str.replace(",", ".", regex=False)
    )

    df_all[indikator] = pd.to_numeric(df_all[indikator], errors="coerce")

    # ===============================
    # NORMALISASI BULAN
    # ===============================
    df_all["Bulan_upper"] = (
        df_all["Bulan"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    # ===============================
    # PILIH TAHUN
    # ===============================
    years = sorted(df_all["Tahun"].unique(), reverse=True)
    selected_year = st.selectbox("Pilih Tahun", years)

    df_filtered = df_all[df_all["Tahun"] == selected_year]

    # ===============================
    # PILIH MODE
    # ===============================
    mode = st.radio(
        "Jenis Periode",
        ["monthly", "quarterly", "compare"],
        format_func=lambda x: {
            "monthly": "Bulanan",
            "quarterly": "Triwulan",
            "compare": "Perbandingan"
        }[x],
        horizontal=True
    )

    # ===============================
    # BULANAN / TRIWULAN
    # ===============================
    if mode in ["monthly", "quarterly"]:

        df = df_filtered.copy()

        if mode == "monthly":
            df["Period"] = df["Bulan_upper"]
            df["Order"] = df["Bulan_upper"].map(MONTH_ORDER)

        else:
            def to_tw(x):
                return {
                    "MARET": "Tw I",
                    "JUNI": "Tw II",
                    "SEPTEMBER": "Tw III",
                    "DESEMBER": "Tw IV"
                }.get(x)

            df["Period"] = df["Bulan_upper"].map(to_tw)
            df["Order"] = df["Period"].map({
                "Tw I":1,"Tw II":2,"Tw III":3,"Tw IV":4
            })

        # ===============================
        # PIVOT SUDAH ADA KODE KPPN
        # ===============================
        df_pivot = df[
            ["Kode KPPN", "Nama KPPN", "Period", indikator]
        ]

        df_wide = df_pivot.pivot_table(
            index=["Kode KPPN", "Nama KPPN"],
            columns="Period",
            values=indikator,
            aggfunc="last"
        ).reset_index()

        # ===============================
        # URUT KOLOM
        # ===============================
        if mode == "monthly":
            ordered = sorted(
                [c for c in df_wide.columns if c not in ["Kode KPPN", "Nama KPPN"]],
                key=lambda x: MONTH_ORDER.get(x, 99)
            )
        else:
            ordered = [c for c in ["Tw I","Tw II","Tw III","Tw IV"] if c in df_wide.columns]

        df_wide = df_wide[["Kode KPPN", "Nama KPPN"] + ordered]

        # ===============================
        # FORMAT KODE KPPN
        # ===============================
        df_wide["Kode KPPN"] = (
            df_wide["Kode KPPN"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.zfill(3)
        )

        # ===============================
        # RANKING
        # ===============================
        if ordered:
            last_col = ordered[-1]
            df_wide["Peringkat"] = (
                df_wide[last_col]
                .rank(ascending=False, method="dense")
                .astype("Int64")
            )

        # ===============================
        # FORMAT ANGKA (🔥 SESUAI REQUEST)
        # ===============================
        def format_angka(x):
            if pd.isna(x):
                return "–"
            if float(x).is_integer():
                return str(int(x))  # 100 → "100"
            return f"{x:.2f}"      # 99.88 → "99.88"

        for col in df_wide.columns:
            if col not in ["Kode KPPN", "Nama KPPN", "Peringkat"]:
                df_wide[col] = df_wide[col].apply(format_angka)

        df_wide = df_wide.sort_values("Peringkat")

        render_table_pin_satker(df_wide)

    # ===============================
    # COMPARE
    # ===============================
    else:
        st.markdown("### 📊 Perbandingan Antar Tahun (KPPN)")

        # ===============================
        # VALIDASI TAHUN
        # ===============================
        available_years = sorted(df_all["Tahun"].dropna().unique())

        if len(available_years) < 2:
            st.warning("Data minimal harus 2 tahun untuk perbandingan")
            return

        # ===============================
        # PILIH 2 TAHUN (MODEL SATKER)
        # ===============================
        colA, colB = st.columns(2)

        with colA:
            year_a = st.selectbox(
                "Tahun A (Awal)",
                available_years,
                index=0,
                key="kppn_year_a"
            )

        with colB:
            year_b = st.selectbox(
                "Tahun B (Akhir)",
                available_years,
                index=1,
                key="kppn_year_b"
            )

        if year_a == year_b:
            st.warning("Pilih dua tahun yang berbeda")
            return

        # ===============================
        # FILTER DATA
        # ===============================
        df_a = df_all[df_all["Tahun"] == year_a]
        df_b = df_all[df_all["Tahun"] == year_b]

        def extract_tw(df_):
            return {
                "Tw I": df_[df_["Bulan_upper"] == "MARET"],
                "Tw II": df_[df_["Bulan_upper"] == "JUNI"],
                "Tw III": df_[df_["Bulan_upper"] == "SEPTEMBER"],
                "Tw IV": df_[df_["Bulan_upper"] == "DESEMBER"],
            }

        tw_a = extract_tw(df_a)
        tw_b = extract_tw(df_b)

        # ===============================
        # LIST KPPN
        # ===============================
        kppn_list = (
            df_all[["Kode KPPN", "Nama KPPN"]]
            .dropna(subset=["Nama KPPN"])
            .drop_duplicates()
            .sort_values("Nama KPPN")
        )

        rows = []

        for _, m in kppn_list.iterrows():

            kode = m["Kode KPPN"]
            nama = m["Nama KPPN"]

            row = {
                "Kode KPPN": kode,
                "Nama KPPN": nama
            }

            latest_a, latest_b = None, None
            has_data = False

            for tw in ["Tw I", "Tw II", "Tw III", "Tw IV"]:

                valA = tw_a[tw].loc[
                    tw_a[tw]["Nama KPPN"] == nama, indikator
                ].values

                valB = tw_b[tw].loc[
                    tw_b[tw]["Nama KPPN"] == nama, indikator
                ].values

                valA = valA[0] if len(valA) else None
                valB = valB[0] if len(valB) else None

                row[f"{tw} {year_a}"] = valA
                row[f"{tw} {year_b}"] = valB

                if valA is not None:
                    latest_a = valA
                    has_data = True
                if valB is not None:
                    latest_b = valB
                    has_data = True

            if not has_data:
                continue

            row[f"Δ ({year_b}-{year_a})"] = (
                latest_b - latest_a
                if latest_a is not None and latest_b is not None
                else None
            )

            rows.append(row)

        if not rows:
            st.warning("Tidak ada data untuk perbandingan")
            return

        df_compare = pd.DataFrame(rows)

        # ===============================
        # FORMAT ANGKA (SAMA SEPERTI TABEL)
        # ===============================
        def format_angka(x):
            if pd.isna(x):
                return "–"
            if float(x).is_integer():
                return str(int(x))
            return f"{x:.2f}"

        for col in df_compare.columns:
            if col not in ["Kode KPPN", "Nama KPPN"]:
                df_compare[col] = df_compare[col].apply(format_angka)

        # ===============================
        # TAMPILKAN
        # ===============================
        render_table_pin_satker(df_compare)
        

def page_trend():
    st.title("📈 Dashboard Internal KPPN")

    # ===============================
    # AUTHENTICATION
    # ===============================
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi Admin.")
        password = st.text_input("Masukkan Password", type="password")

        if st.button("Login"):
            if password == ADMIN_PASSWORD:
                st.session_state.authenticated = True
                st.success("Login berhasil!")
                st.rerun()
            else:
                st.error("Password salah!")
        return

    # ===============================
    # MENU DASHBOARD INTERNAL
    # ===============================
    menu = st.radio(
        "Pilih Menu",
        [
            "🏛️ Early Warning System Kinerja Keuangan Satker",
            "🎯 IKPA KPPN"
        ],
        horizontal=True
    )

    st.markdown("---")

    # ===============================
    # 🔽 PANGGIL ISI MENU
    # ===============================
    if menu == "🏛️ Early Warning System Kinerja Keuangan Satker":
        menu_ews_satker()

    elif menu == "🎯 IKPA KPPN":
    
        tab1, tab2 = st.tabs([
            "🎯 Highlights IKPA KPPN",
            "📊 Tabel IKPA KPPN"
        ])

        with tab1:
            menu_highlights()

        with tab2:
            menu_tabel_ikpa_kppn()
   
# ============================================================
# HALAMAN 3: ADMIN 
# ============================================================

# ======================================================================================
# DETECT DIPA HEADER (ROBUST VERSION)
# ======================================================================================
def detect_dipa_header(uploaded_file):
    """
    Auto-detect header row dalam file DIPA mentah.
    Returns: DataFrame dengan header yang sudah benar
    """
    try:
        uploaded_file.seek(0)
        
        # Baca 20 baris pertama untuk preview
        preview = pd.read_excel(uploaded_file, header=None, nrows=20, dtype=str)
        
        # Keywords yang PASTI ada di header DIPA
        header_keywords = [
            "satker", "kode", "pagu", "jumlah", "dipa", 
            "tanggal", "revisi", "no", "status"
        ]
        
        header_row = None
        max_matches = 0
        
        # Cari baris dengan keyword terbanyak
        for i in range(len(preview)):
            row_text = " ".join(preview.iloc[i].fillna("").astype(str).str.lower())
            matches = sum(1 for kw in header_keywords if kw in row_text)
            
            if matches > max_matches:
                max_matches = matches
                header_row = i
        
        # Jika tidak ada yang cocok, gunakan baris 0
        if header_row is None or max_matches < 3:
            header_row = 0
        
        # Baca ulang dengan header yang benar
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, header=header_row, dtype=str)
        
        # Bersihkan nama kolom
        df.columns = (
            df.columns.astype(str)
            .str.replace("\n", " ", regex=False)
            .str.replace("\r", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
            .str.upper()  # Normalize to uppercase
        )
        
        # Hapus baris kosong
        df = df.dropna(how='all').reset_index(drop=True)
        
        
        return df
        
    except Exception as e:
        st.error(f"❌ Error deteksi header: {e}")
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file, dtype=str)


# ======================================================================================
# CLEAN DIPA (SUPER ROBUST VERSION)
# ======================================================================================
def clean_dipa(df_raw):
    """
    Membersihkan file DIPA mentah dan mengembalikan format standar.
    """
    
    df = df_raw.copy()
    
    # Hapus kolom Unnamed
    df = df.loc[:, ~df.columns.astype(str).str.contains("Unnamed", case=False)]
    
    # Normalize column names untuk matching yang lebih mudah
    df.columns = df.columns.astype(str).str.upper().str.strip()
    
    st.write("**Kolom yang terdeteksi di file DIPA:**")
    st.write(list(df.columns))
    
    # ====== HELPER: Flexible column finder ======
    def find_col(keywords_list):
        """Find column by multiple possible names"""
        for col in df.columns:
            col_clean = str(col).upper().replace(" ", "").replace("_", "")
            for kw in keywords_list:
                kw_clean = kw.upper().replace(" ", "").replace("_", "")
                if kw_clean in col_clean:
                    return col
        return None
    
    # ====== 1. KODE SATKER & NAMA SATKER ======
    satker_col = find_col([
        "SATKER", "KODESATKER", "KODE SATKER", "KODE SATUAN KERJA",
        "SATUAN KERJA", "SATKER/KPPN"
    ])
    
    if satker_col is None:
        st.error("❌ Kolom Satker tidak ditemukan!")
        st.write("Kolom available:", list(df.columns))
        raise ValueError("Kolom Satker tidak ditemukan")
    
    st.success(f"✅ Kolom Satker ditemukan: {satker_col}")
    
    # Extract kode 6 digit
    df["Kode Satker"] = (
        df[satker_col].astype(str)
        .str.extract(r"(\d{6})", expand=False)
        .fillna("")
    )
    
    # Extract nama satker (remove kode di depan)
    df["Satker"] = (
        df[satker_col].astype(str)
        .str.replace(r"^\d{6}\s*-?\s*", "", regex=True)
        .str.strip()
    )
    
    # Jika nama kosong, gunakan format default
    mask_empty = (df["Satker"] == "") | (df["Satker"].isna())
    df.loc[mask_empty, "Satker"] = df.loc[mask_empty, "Kode Satker"] + " - SATKER"
    
    # ====== 2. TOTAL PAGU ======
    pagu_col = find_col([
        "PAGU", "JUMLAH", "TOTAL PAGU", "TOTALPAGU", "NILAI PAGU",
        "PAGU DIPA", "JUMLAH DIPA"
    ])
    
    if pagu_col:
        st.success(f"✅ Kolom Pagu ditemukan: {pagu_col}")
        df["Total Pagu"] = pd.to_numeric(df[pagu_col], errors="coerce").fillna(0).astype(int)
    else:
        st.warning("⚠️ Kolom Pagu tidak ditemukan, menggunakan 0")
        df["Total Pagu"] = 0
    
    # ====== 3. TANGGAL POSTING REVISI ======
    tgl_col = find_col([
        "TANGGAL POSTING", "TGL POSTING", "TANGGALPOSTING",
        "TANGGAL REVISI", "TGL REVISI", "TANGGAL", "DATE"
    ])
    
    if tgl_col:
        st.success(f"✅ Kolom Tanggal ditemukan: {tgl_col}")
        df["Tanggal Posting Revisi"] = pd.to_datetime(df[tgl_col], errors="coerce")
    else:
        st.warning("⚠️ Kolom Tanggal tidak ditemukan")
        df["Tanggal Posting Revisi"] = pd.NaT
    
    # Extract Tahun
    df["Tahun"] = df["Tanggal Posting Revisi"].dt.year
    df["Tahun"] = df["Tahun"].fillna(datetime.now().year).astype(int)
    
    # ====== 4. KOLOM LAINNYA (OPTIONAL) ======
    
    # NO
    no_col = find_col(["NO", "NOMOR", "NO."])
    df["NO"] = df[no_col].astype(str).str.strip() if no_col else ""
    
    # KEMENTERIAN
    kementerian_col = find_col(["KEMENTERIAN", "KEMENTRIAN", "KL", "K/L", "BA"])
    df["Kementerian"] = df[kementerian_col].astype(str).str.strip() if kementerian_col else ""
    
    # KODE STATUS HISTORY
    status_col = find_col(["STATUS HISTORY", "KODE STATUS", "KODESTATUS", "STATUS"])
    df["Kode Status History"] = df[status_col].astype(str).str.strip() if status_col else ""
    
    # JENIS REVISI
    jenis_col = find_col(["JENIS REVISI", "JENISREVISI", "TIPE REVISI"])
    df["Jenis Revisi"] = df[jenis_col].astype(str).str.strip() if jenis_col else ""
    
    # REVISI KE-
    revisi_col = find_col(["REVISI KE", "REVISIKE", "REVISI"])
    if revisi_col:
        df["Revisi ke-"] = pd.to_numeric(df[revisi_col], errors="coerce").fillna(0).astype(int)
    else:
        df["Revisi ke-"] = 0
    
    # NO DIPA
    nodipa_col = find_col(["NO DIPA", "NODIPA", "NOMOR DIPA", "NO. DIPA"])
    df["No Dipa"] = df[nodipa_col].astype(str).str.strip() if nodipa_col else ""
    
    # TANGGAL DIPA
    tgldipa_col = find_col(["TANGGAL DIPA", "TGL DIPA", "TGLDIPA"])
    if tgldipa_col:
        df["Tanggal Dipa"] = pd.to_datetime(df[tgldipa_col], errors="coerce").dt.strftime("%d-%m-%Y")
    else:
        df["Tanggal Dipa"] = ""
    
    # OWNER
    owner_col = find_col(["OWNER", "PEMILIK"])
    df["Owner"] = df[owner_col].astype(str).str.strip() if owner_col else ""
    
    # DIGITAL STAMP
    stamp_col = find_col(["DIGITAL STAMP", "DIGITALSTAMP", "STAMP", "TTD DIGITAL"])
    df["Digital Stamp"] = df[stamp_col].astype(str).str.strip() if stamp_col else ""
    
    # ====== FINAL: Susun kolom sesuai urutan ======
    final_columns = [
        "Kode Satker", "Satker", "Tahun", "Tanggal Posting Revisi", "Total Pagu",
        "NO", "Kementerian", "Kode Status History", "Jenis Revisi", "Revisi ke-",
        "No Dipa", "Tanggal Dipa", "Digital Stamp"
    ]
    
    df_clean = df[final_columns].copy()
    
    # Ambil revisi terakhir per satker per tahun
    df_clean = df_clean.sort_values(["Kode Satker", "Tahun", "Tanggal Posting Revisi"])
    df_clean = df_clean.groupby(["Kode Satker", "Tahun"], as_index=False).tail(1)
    
    # Filter hanya kode satker yang valid (6 digit)
    df_clean = df_clean[df_clean["Kode Satker"].str.len() == 6]
    
    st.write(f"**Hasil cleaning: {len(df_clean)} baris satker valid**")
    
    return df_clean


# ======================================================================================
# ASSIGN JENIS SATKER
# ======================================================================================
def assign_jenis_satker(df):
    """Klasifikasi satker berdasarkan Total Pagu (persentil)"""

    if df.empty or "Total Pagu" not in df.columns:
        df["Jenis Satker"] = "Satker Kecil"
        return df

    # pastikan numerik
    df["Total Pagu"] = pd.to_numeric(df["Total Pagu"], errors="coerce").fillna(0)

    p40 = df["Total Pagu"].quantile(0.40)
    p70 = df["Total Pagu"].quantile(0.70)

    if df["Total Pagu"].sum() == 0 or p40 == p70:
        # Rank-based jika semua pagu sama / tidak bisa dibagi bins
        df["_rank"] = df["Total Pagu"].rank(method="first")
        n = len(df)
        cut40 = n * 0.40
        cut70 = n * 0.70
        df["Jenis Satker"] = df["_rank"].apply(
            lambda r: "Satker Kecil" if r <= cut40 else ("Satker Sedang" if r <= cut70 else "Satker Besar")
        )
        df = df.drop(columns=["_rank"])
    else:
        df["Jenis Satker"] = pd.cut(
            df["Total Pagu"],
            bins=[-float("inf"), p40, p70, float("inf")],
            labels=["Satker Kecil", "Satker Sedang", "Satker Besar"]
        )

    # rapikan posisi kolom
    cols = list(df.columns)
    if "Jenis Satker" in cols and "Total Pagu" in cols:
        cols.remove("Jenis Satker")
        idx = cols.index("Total Pagu")
        cols.insert(idx + 1, "Jenis Satker")
        df = df[cols]

    return df


# ======================================================================================
# PROCESS UPLOADED DIPA (MAIN FUNCTION)
# ======================================================================================
def process_uploaded_dipa(uploaded_file, save_file_to_github, forced_year=None):
    """Process file DIPA upload user (FIX: tahun dikontrol dari UI)"""

    try:
        st.info("📄 Memulai proses upload DIPA...")

        # ===============================
        # 1️⃣ BACA FILE
        # ===============================
        df_raw = pd.read_excel(uploaded_file, header=None)

        # 🔥 FIX HEADER
        df_raw = fix_dipa_header(df_raw)

        if df_raw is None or df_raw.empty:
            return None, None, "❌ File kosong / header tidak terbaca"

        # ===============================
        # 2️⃣ STANDARDISASI
        # ===============================
        if is_omspan_dipa(df_raw):
            df_adapted = adapt_dipa_omspan(df_raw)

            if df_adapted.empty:
                return None, None, "❌ Data OMSPAN tidak valid"

            df_std = df_adapted.copy()
        else:
            df_std = standardize_dipa(df_raw)

        # ===============================
        # 3️⃣ PAKSA TAHUN DARI UI
        # ===============================
        if forced_year is not None:
            tahun_dipa = int(forced_year)
        else:
            # fallback kalau tidak dikirim dari UI
            if "Tahun" in df_std.columns and not df_std["Tahun"].isna().all():
                tahun_dipa = int(df_std["Tahun"].dropna().mode()[0])
            else:
                tahun_dipa = int(df_std["Tanggal Posting Revisi"].dropna().dt.year.min())

        # 🔒 LOCK TAHUN
        df_std["Tahun"] = tahun_dipa

        # ===============================
        # 4️⃣ FIX TANGGAL
        # ===============================
        df_std["Tanggal Posting Revisi"] = pd.to_datetime(
            df_std["Tanggal Posting Revisi"], errors="coerce"
        )

        mask_na = df_std["Tanggal Posting Revisi"].isna()
        df_std.loc[mask_na, "Tanggal Posting Revisi"] = pd.to_datetime(
            df_std.loc[mask_na, "Tahun"].astype(str) + "-12-31"
        )

        # ===============================
        # 5️⃣ NORMALISASI METADATA
        # ===============================
        df_std["Owner"] = df_std.get("Owner", "SATKER").fillna("SATKER")

        df_std["Digital Stamp"] = (
            df_std.get("Digital Stamp", "OMSPAN (NON-SPAN)")
            .replace("", pd.NA)
            .fillna("OMSPAN (NON-SPAN)")
        )

        # ===============================
        # 6️⃣ VALIDASI (BERSIH, TANPA TAHUN LAMA)
        # ===============================
        st.write(f"**Validasi:** {len(df_std)} baris data")
        st.write(f"**Tahun digunakan:** {tahun_dipa}")

        # ===============================
        # 7️⃣ NORMALISASI KODE SATKER
        # ===============================
        df_std["Kode Satker"] = df_std["Kode Satker"].apply(normalize_kode_satker)

        # ===============================
        # 8️⃣ MERGE REFERENSI
        # ===============================
        if "reference_df" in st.session_state and not st.session_state.reference_df.empty:
            ref = st.session_state.reference_df.copy()
            ref["Kode Satker"] = ref["Kode Satker"].apply(normalize_kode_satker)

            df_std = df_std.merge(
                ref[["Kode BA", "K/L", "Kode Satker"]],
                on="Kode Satker",
                how="left"
            )

            if "Kementerian" in df_std.columns and "K/L" in df_std.columns:
                df_std["Kementerian"] = df_std["Kementerian"].fillna(df_std["K/L"])

        # ===============================
        # 9️⃣ NAMA SATKER FIX
        # ===============================
        if "Satker" not in df_std.columns:
            df_std["Satker"] = pd.NA

        if "Uraian Satker-RINGKAS" in df_std.columns:
            df_std["Satker"] = df_std["Satker"].fillna(df_std["Uraian Satker-RINGKAS"])

        if "Uraian Satker-LENGKAP" in df_std.columns:
            df_std["Satker"] = df_std["Satker"].fillna(df_std["Uraian Satker-LENGKAP"])

        df_std["Satker"] = df_std["Satker"].fillna(
            "SATKER " + df_std["Kode Satker"].astype(str)
        )

        # ===============================
        # 🔟 JENIS SATKER
        # ===============================
        df_std = assign_jenis_satker(df_std)

        # ===============================
        # 1️⃣1️⃣ AMBIL REVISI TERAKHIR
        # ===============================
        df_std = df_std.sort_values(
            ["Kode Satker", "Tanggal Posting Revisi"],
            ascending=[True, False]
        ).drop_duplicates("Kode Satker")

        # ===============================
        # 1️⃣2️⃣ FINAL FORMAT
        # ===============================
        df_std = df_std.reset_index(drop=True)
        df_std["NO"] = df_std.index + 1

        df_std["Tanggal Dipa"] = df_std["Tanggal Posting Revisi"]

        FINAL_COLUMNS = [
            "Kode Satker", "Satker", "Tahun",
            "Tanggal Posting Revisi", "Total Pagu",
            "Jenis Satker", "NO", "Kementerian",
            "Kode Status History", "Jenis Revisi",
            "Revisi ke-", "No Dipa", "Tanggal Dipa",
            "Digital Stamp"
        ]

        df_std = df_std.reindex(columns=FINAL_COLUMNS)

        # ===============================
        # 1️⃣3️⃣ PREVIEW SAJA
        # ===============================
        st.write("**Preview 5 baris pertama:**")
        st.dataframe(df_std.head(5))

        # ===============================
        # VALIDASI TOTAL PAGU SEBELUM RETURN
        # ===============================
        df_std["Total Pagu"] = pd.to_numeric(df_std["Total Pagu"], errors="coerce").fillna(0)
        
        if df_std["Total Pagu"].sum() == 0:
            st.warning("⚠️ Total Pagu semua 0 — cek apakah file DIPA yang diupload benar (harus file mentah SPAN/OMSPAN, bukan file yang sudah diproses)")

        return df_std, tahun_dipa, "✅ Sukses diproses"

    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        return None, None, f"❌ Error: {str(e)}"
    
    
import streamlit as st
import pandas as pd
from github import Github, Auth
import base64

# Fungsi bantu
def get_latest_dipa(dipa_df):
    if 'Tanggal Posting Revisi' in dipa_df.columns:
        dipa_df['Tanggal Posting Revisi'] = pd.to_datetime(dipa_df['Tanggal Posting Revisi'], errors='coerce')
        latest_dipa = dipa_df.sort_values('Tanggal Posting Revisi', ascending=False) \
                              .drop_duplicates(subset='Kode Satker', keep='first')
    elif 'No Revisi Terakhir' in dipa_df.columns:
        latest_dipa = dipa_df.sort_values('No Revisi Terakhir', ascending=False) \
                              .drop_duplicates(subset='Kode Satker', keep='first')
    else:
        latest_dipa = dipa_df.drop_duplicates(subset='Kode Satker', keep='first')
    return latest_dipa


def merge_ikpa_dipa_auto():
    
    if st.session_state.get("ikpa_dipa_merged", False):
        return

    if "data_storage" not in st.session_state:
        return

    if "DATA_DIPA_by_year" not in st.session_state:
        return

    # ===============================
    # 🔥 BUILD DIPA (TANPA FILTER PAGU)
    # ===============================
    valid_dipa_years = {}

    for yr, dipa_df in st.session_state.DATA_DIPA_by_year.items():

        if dipa_df is None or dipa_df.empty:
            continue

        dipa_df = dipa_df.copy()

        # ===============================
        # 🔥 FIX FORMAT PAGU (INDONESIA)
        # ===============================
        if "Total Pagu" not in dipa_df.columns:
            st.error("❌ DIPA tidak memiliki kolom Total Pagu")
            st.stop()

        dipa_df["Total Pagu"] = (
            dipa_df["Total Pagu"]
            .astype(str)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )

        dipa_df["Total Pagu"] = pd.to_numeric(
            dipa_df["Total Pagu"], errors="coerce"
        ).fillna(0)

        # 🔥 LANGSUNG MASUK (TANPA FILTER)
        valid_dipa_years[int(yr)] = dipa_df

    if not valid_dipa_years:
        st.warning("❌ Tidak ada DIPA valid")
        return



    # ===============================
    # 🔥 LOOP IKPA
    # ===============================
    for (bulan, tahun), df_ikpa in st.session_state.data_storage.items():
        
        # ===============================
        # 🔥 VALIDASI IKPA (WAJIB BANGET)
        # ===============================
        if df_ikpa is None or df_ikpa.empty:
            st.warning(f"❌ IKPA kosong untuk {bulan}-{tahun}")
            continue

        if "Kode Satker" not in df_ikpa.columns:
            st.error(f"❌ IKPA tidak punya kolom 'Kode Satker' ({bulan}-{tahun})")
            continue

        # 🔥 NORMALISASI SATKER AWAL
        df_ikpa["Kode Satker"] = (
            df_ikpa["Kode Satker"]
            .astype(str)
            .str.extract(r"(\d+)")[0]
            .fillna("")
            .str.zfill(6)
        )

        jumlah_satker = df_ikpa["Kode Satker"].nunique()


        # 🔥 FILTER DATA RUSAK (threshold diturunkan — awal tahun satker bisa sedikit)
        if jumlah_satker < 5:
            st.error(f"❌ IKPA tidak normal (hanya {jumlah_satker} satker) → SKIP")
            continue

        if df_ikpa is None or df_ikpa.empty:
            continue

        tahun_int = int(tahun)

        dipa = valid_dipa_years.get(tahun_int)

        # ===============================
        # 🔥 FALLBACK TAHUN TERDEKAT (HARUS SEBELUM CEK KOLOM)
        # ===============================
        if dipa is None and valid_dipa_years:
            nearest = min(valid_dipa_years.keys(), key=lambda y: abs(y - tahun_int))
            dipa = valid_dipa_years[nearest]
            st.warning(f"⚠️ DIPA {tahun_int} tidak ada → pakai DIPA {nearest}")

        if dipa is None or dipa.empty:
            st.warning(f"⚠️ Tidak ada DIPA untuk {bulan}-{tahun}, lanjut tanpa merge pagu")
            df_ikpa["Total Pagu"] = 0
            df_ikpa = classify_jenis_satker(df_ikpa)
            df_ikpa = apply_reference_short_names(df_ikpa)
            df_ikpa = create_satker_column(df_ikpa)
            st.session_state.data_storage[(bulan, tahun)] = df_ikpa
            continue

        # ===============================
        # VALIDASI KOLOM PAGU
        # ===============================
        if "Total Pagu" not in dipa.columns:
            st.warning(f"⚠️ Kolom 'Total Pagu' tidak ada di DIPA {tahun_int}, lanjut tanpa merge pagu")
            df_ikpa["Total Pagu"] = 0
            df_ikpa = classify_jenis_satker(df_ikpa)
            df_ikpa = apply_reference_short_names(df_ikpa)
            df_ikpa = create_satker_column(df_ikpa)
            st.session_state.data_storage[(bulan, tahun)] = df_ikpa
            continue

        df_final = df_ikpa.copy()

        # ===============================
        # 🔥 NORMALISASI KODE SATKER
        # ===============================
        df_final["Kode Satker"] = (
            df_final["Kode Satker"]
            .astype(str)
            .str.extract(r"(\d+)")[0]
            .fillna("")
            .str.zfill(6)
        )

        dipa["Kode Satker"] = (
            dipa["Kode Satker"]
            .astype(str)
            .str.extract(r"(\d+)")[0]
            .fillna("")
            .str.zfill(6)
        )

        # ===============================
        # 🔥 AGREGASI DIPA (AMAN)
        # ===============================
        dipa_by_kode = (
            dipa
            .groupby("Kode Satker", as_index=False)["Total Pagu"]
            .max()
        )

        # ===============================
        # 🔥 HAPUS KOLOM LAMA BIAR GA TABRAKAN
        # ===============================
        df_final = df_final.drop(columns=["Total Pagu"], errors="ignore")

        # ===============================
        # 🔥 FINAL NORMALIZATION (WAJIB)
        # ===============================
        df_final["Kode Satker"] = df_final["Kode Satker"].astype(str).str[-6:]
        dipa_by_kode["Kode Satker"] = dipa_by_kode["Kode Satker"].astype(str).str[-6:]
        
        # ===============================
        # 🔥 MERGE
        # ===============================
        df_merged = pd.merge(
            df_final,
            dipa_by_kode,
            on="Kode Satker",
            how="left"
        )

        # ===============================
        # 🔥 FIX KOLOM DUPLIKAT
        # ===============================
        if "Total Pagu_y" in df_merged.columns:
            df_merged["Total Pagu"] = df_merged["Total Pagu_y"]
        elif "Total Pagu_x" in df_merged.columns:
            df_merged["Total Pagu"] = df_merged["Total Pagu_x"]

        df_merged = df_merged.drop(
            columns=[c for c in ["Total Pagu_x", "Total Pagu_y"] if c in df_merged.columns],
            errors="ignore"
        )


        # ===============================
        # 🔥 FINAL CLEAN
        # ===============================
        df_merged["Total Pagu"] = pd.to_numeric(
            df_merged["Total Pagu"], errors="coerce"
        ).fillna(0)

        # ===============================
        # 🔥 POST PROCESS
        # ===============================
        df_merged = classify_jenis_satker(df_merged)
        df_merged = apply_reference_short_names(df_merged)
        df_merged = create_satker_column(df_merged)

        # ===============================
        # SIMPAN
        # ===============================
        st.session_state.data_storage[(bulan, tahun)] = df_merged

    st.session_state.ikpa_dipa_merged = True
    

# ============================================================
# 🔹 Fungsi convert DataFrame ke Excel bytes
# ============================================================
def to_excel_bytes(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

# ============================================================
# 🔹 Fungsi push file ke GitHub
# ============================================================
def push_to_github(file_bytes, repo_path, repo_name, token, commit_message):
    g = Github(auth=Auth.Token(token))
    repo = g.get_repo(repo_name)

    try:
        # Cek apakah file sudah ada
        try:
            contents = repo.get_contents(repo_path)
            repo.update_file(contents.path, commit_message, file_bytes, contents.sha)
            st.success(f"✅ File {repo_path} berhasil diupdate di GitHub")
        except Exception as e_inner:
            # Jika file belum ada atau path salah, buat baru
            repo.create_file(repo_path, commit_message, file_bytes)
            st.success(f"✅ File {repo_path} berhasil dibuat di GitHub")
    except Exception as e:
        st.error(f"❌ Gagal push ke GitHub: {e}")
        
        
        
def detect_header_row(file_or_df, max_scan=15):
    import pandas as pd

    # ===============================
    # HANDLE INPUT
    # ===============================
    if isinstance(file_or_df, pd.DataFrame):
        df_raw = file_or_df.copy()
    else:
        try:
            file_or_df.seek(0)
        except:
            pass

        df_raw = pd.read_excel(file_or_df, header=None, nrows=max_scan)

    # ===============================
    # DETEKSI HEADER (AMAN TANPA .str)
    # ===============================
    keywords = ["KODE", "KPPN", "SATKER", "NILAI"]

    for i in range(min(max_scan, len(df_raw))):

        row = df_raw.iloc[i]

        # ubah semua cell jadi string aman
        row_str = [str(cell).upper() for cell in row]

        score = sum(
            any(k in cell for cell in row_str)
            for k in keywords
        )

        if score >= 2:
            return i

    return 0


def detect_format(df):
    cols = [str(c).upper() for c in df.columns]

    if any("NILAI AKHIR" in c for c in cols):
        return "KPPN"

    if df.shape[0] > 50:
        return "SATKER"

    return "UNKNOWN"

# ============================================================
#  Menu Admin
# ============================================================
def page_admin():
    st.title("🔐 Halaman Administrasi")

    # ===============================
    # 🔑 LOGIN ADMIN
    # ===============================
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if not st.session_state.authenticated:
        st.warning("🔒 Halaman ini memerlukan autentikasi Admin")
        password = st.text_input("Masukkan Password Admin", type="password")

        if st.button("Login"):
            if password == ADMIN_PASSWORD:
                st.session_state.authenticated = True
                st.success("✔ Login berhasil")
                st.rerun()
            else:
                st.error("❌ Password salah")
        return

    st.success("✔ Anda login sebagai Admin")

    # ===============================
    # 🔄 KONTROL DATA (MANUAL OVERRIDE)
    # ===============================
    st.subheader("Manajemen Data")

    # ============================================================
    # JIKA DATA SUDAH SIAP (MERGE BERHASIL)
    # ============================================================
    if st.session_state.get("ikpa_dipa_merged", False):

        add_notification(" Data IKPA & DIPA sudah siap digunakan dan merge berhasil")
        st.caption("Tidak diperlukan proses atau tindakan Admin")

    # ============================================================
    #  JIKA DATA BELUM SIAP (BELUM MERGE / GAGAL)
    # ============================================================
    else:

        st.warning("⚠️ Data belum siap atau perlu diproses")

        # Reset hanya muncul kalau data ada tapi merge gagal
        if st.session_state.get("data_storage") or st.session_state.get("DATA_DIPA_by_year"):
            with st.expander(" Admin Lanjutan (Opsional)"):
                if st.button(" Reset Status Merge"):
                    st.session_state.ikpa_dipa_merged = False
                    st.warning(" Status merge direset. Data akan diproses ulang.")
                    st.rerun()


    # ===============================
    # 📌 TAB MENU
    # ===============================
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📤 Tambah Data",
        "🗑️ Hapus Data",
        "📥 Download Data",
        "📋 Download Template",
        "🕓 Riwayat Aktivitas"
    ])

    # ============================================================
    # TAB 1: UPLOAD DATA (IKPA, DIPA, Referensi)
    # ============================================================
    with tab1:
        # Upload Data IKPA Satker
        st.subheader("📤 Upload Data IKPA Satker")

        upload_year = st.selectbox(
            "Pilih Tahun",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year)
        )

        st.caption(
            "Sistem dapat memproses Data IKPA SATKER yang bersumber dari :\n"
            "1. Aplikasi OM-SPAN, Menu Monev PA → Indikator Pelaksanaan Anggaran → Indikator Pelaksanaan Anggaran SATKER\n"
            "2. Aplikasi MyIntress, Menu Tematik → Indikator Pelaksanaan Anggaran → Indikator Pelaksanaan Anggaran SATKER."
        )
      
        uploaded_files = st.file_uploader(
            "Pilih satu atau beberapa file Excel IKPA Satker",
            type=["xlsx", "xls"],
            accept_multiple_files=True
        )

        if uploaded_files:
    
            st.info("📄 File yang diupload:")
            for f in uploaded_files:
                st.write("•", f.name)

            if st.button("🔄 Proses Semua Data IKPA", type="primary"):

                with st.spinner("Memproses semua file IKPA Satker..."):

                    # ======================
                    # 🔥 INIT STORAGE
                    # ======================
                    if "data_storage" not in st.session_state:
                        st.session_state.data_storage = {}

                    success_count = 0

                    for uploaded_file in uploaded_files:
                        try:
                            # ======================
                            # 🔄 PROSES FILE
                            # ======================
                            uploaded_file.seek(0)
                            result = process_excel_file(uploaded_file, upload_year)

                            # 🔥 HANDLE RESULT AMAN
                            if not result or result[0] is None:
                                st.error(f"❌ Gagal parsing: {uploaded_file.name}")
                                continue

                            df_final, month, year = result

                            # ======================
                            # FIX BULAN (1x SAJA)
                            # ======================
                            if month == "UNKNOWN":
                                import re

                                # 🔥 ambil angka bulan di akhir nama file (paling aman)
                                match = re.search(r"(\d{4})\s*(0[1-9]|1[0-2])", uploaded_file.name)

                                if match:
                                    bulan_num = match.group(2)

                                    bulan_map = {
                                        "01":"JANUARI","02":"FEBRUARI","03":"MARET","04":"APRIL",
                                        "05":"MEI","06":"JUNI","07":"JULI","08":"AGUSTUS",
                                        "09":"SEPTEMBER","10":"OKTOBER","11":"NOVEMBER","12":"DESEMBER"
                                    }

                                    month = bulan_map.get(bulan_num, "UNKNOWN")

                            # 🔥 FINAL FALLBACK (kalau masih gagal)
                            if month == "UNKNOWN":
                                month = "MARET"

                            # ======================
                            # VALIDASI DATA
                            # ======================
                            if df_final.empty:
                                st.error(f"❌ Data kosong: {uploaded_file.name}")
                                continue

                            # ======================
                            # NORMALISASI SATKER
                            # ======================
                            df_final["Kode Satker"] = (
                                df_final["Kode Satker"]
                                .astype(str)
                                .str.extract(r"(\d+)")[0]
                                .fillna("")
                                .str.zfill(6)
                            )

                            # ======================
                            # POST PROCESS
                            # ======================
                            df_final = post_process_ikpa_satker(df_final)
                            df_final = normalize_ikpa_columns(df_final)
                            df_final = ensure_ikpa_columns(df_final)

                            # ======================
                            # 🔥 SIMPAN KE SESSION (FIX UTAMA)
                            # ======================
                            st.session_state.data_storage[(month, str(year))] = df_final.copy()

                            st.success(f"✅ MASUK STORAGE: {month}-{year}")

                            # ======================
                            # 🔥 SIMPAN KE GITHUB
                            # ======================
                            try:
                                excel_bytes = io.BytesIO()
                                with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                                    df_final.to_excel(writer, index=False)

                                excel_bytes.seek(0)

                                save_file_to_github(
                                    excel_bytes.getvalue(),
                                    f"IKPA_{month}_{year}.xlsx",
                                    folder="data"
                                )

                                st.success(f"☁️ Upload GitHub: IKPA_{month}_{year}")

                            except Exception as e:
                                st.error(f"❌ GitHub gagal: {e}")

                            success_count += 1

                        except Exception as e:
                            st.error(f"❌ ERROR FILE {uploaded_file.name}: {e}")


                    if success_count == 0:
                        st.error("❌ Tidak ada file yang berhasil diproses")
                        st.stop()

                    # ======================
                    # 🔥 MERGE IKPA + DIPA
                    # ======================
                    if st.session_state.get("DATA_DIPA_by_year"):
                        merge_ikpa_dipa_auto()

                    st.success("🎉 SEMUA FILE BERHASIL DIPROSES")

                    st.rerun()
            
        
        # Submenu Upload Data IKPA KPPN
        st.markdown("---")
        st.subheader("📝 Upload Data IKPA KPPN")

        # ===============================
        # 📅 PILIH TAHUN
        # ===============================
        upload_year_kppn = st.selectbox(
            "Pilih Tahun",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year),
            key="tahun_kppn"
        )
        
        st.caption(
            "Sistem dapat memproses Data IKPA SATKER yang bersumber dari :\n"
            "1. Aplikasi OM-SPAN, menu Monev PA → Indikator Pelaksanaan Anggaran → Indikator Pelaksanaan Anggaran KPPN\n"
            "2. Aplikasi MyIntress, menu Tematik → Indikator Pelaksanaan Anggaran → Indikator Pelaksanaan Anggaran IKPPN."
        )

        # ===============================
        # 📂 UPLOAD FILE
        # ===============================
        uploaded_file_kppn = st.file_uploader(
            "Pilih file Excel IKPA KPPN",
            type=["xlsx", "xls"],
            key="file_kppn"
        )

        # ===============================
        # 🔐 INISIALISASI SESSION
        # ===============================
        if "data_storage_kppn" not in st.session_state:
            st.session_state.data_storage_kppn = {}

        # ===============================
        # 🚦 VALIDASI FILE
        # ===============================
        if uploaded_file_kppn is not None:
            try:
                # Cari baris header otomatis — coba keyword KPPN-spesifik dulu
                header_row = find_header_row_by_keywords(
                    uploaded_file_kppn,
                    keywords=[
                        "Nama KPPN",
                        "KPPN",
                        "Nama Kantor",
                        "Kantor Pelayanan"
                    ]
                )

                # Fallback: format MyIntress / OM-SPAN modern (ada KODE + NAMA di header)
                if header_row is None:
                    header_row = find_header_row_by_keywords(
                        uploaded_file_kppn,
                        keywords=["Kode", "Nama", "Nilai Akhir", "INDIKATOR"]
                    )

                # Validasi final: cek apakah ini benar file IKPA KPPN dari judulnya
                if header_row is None:
                    uploaded_file_kppn.seek(0)
                    df_peek = pd.read_excel(uploaded_file_kppn, header=None, nrows=5)
                    judul_rows = " ".join(df_peek.iloc[:3].astype(str).values.flatten()).upper()
                    if "KPPN" not in judul_rows and "KANTOR PELAYANAN" not in judul_rows:
                        st.error(
                            "GAGAL UPLOAD!\n\n"
                            "Format file tidak dikenali sebagai **IKPA KPPN**.\n"
                            "Pastikan file bersumber dari OM-SPAN atau MyIntress "
                            "(menu IKPA KPPN)."
                        )
                        st.stop()
                    # Judul ada kata KPPN → parser akan handle formatnya
                    header_row = 0

                # Baca data dengan header yang benar
                uploaded_file_kppn.seek(0)
                df_check = pd.read_excel(
                    uploaded_file_kppn,
                    header=header_row
                )

                # Normalisasi nama kolom
                df_check.columns = (
                    df_check.columns.astype(str)
                    .str.strip()
                    .str.replace(r"\s+", " ", regex=True)
                )

                # SALAH FILE: IKPA SATKER
                if any(col.lower().startswith("nama satker") for col in df_check.columns):
                    st.error(
                        "GAGAL UPLOAD!\n\n"
                        "File yang Anda upload adalah **IKPA SATKER**.\n"
                        "Halaman ini hanya menerima **IKPA KPPN**."
                    )
                    st.stop()

                # ===============================
                # 🔍 DETEKSI BULAN (HEADER + DATA + FILENAME)
                # ===============================
                uploaded_file_kppn.seek(0)
                df_info = pd.read_excel(uploaded_file_kppn, header=None)

                MONTH_MAP = {
                    "JANUARI": "JANUARI", "JAN": "JANUARI",
                    "FEBRUARI": "FEBRUARI", "FEB": "FEBRUARI", "PEBRUARI": "FEBRUARI",
                    "MARET": "MARET", "MAR": "MARET",
                    "APRIL": "APRIL", "APR": "APRIL",
                    "MEI": "MEI",
                    "JUNI": "JUNI", "JUN": "JUNI",
                    "JULI": "JULI", "JUL": "JULI",
                    "AGUSTUS": "AGUSTUS", "AGT": "AGUSTUS", "AGS": "AGUSTUS",
                    "SEPTEMBER": "SEPTEMBER", "SEP": "SEPTEMBER",
                    "OKTOBER": "OKTOBER", "OKT": "OKTOBER",
                    "NOVEMBER": "NOVEMBER", "NOV": "NOVEMBER", "NOPEMBER": "NOVEMBER",
                    "DESEMBER": "DESEMBER", "DES": "DESEMBER"
                }

                month_preview = None

                # 1️⃣ Cari bulan di header (baris & kolom awal) — untuk raw OM-SPAN
                for r in range(min(6, df_info.shape[0])):
                    for c in range(min(5, df_info.shape[1])):
                        cell = str(df_info.iloc[r, c]).upper().strip()
                        # exact match dulu (lebih aman dari substring)
                        if cell in MONTH_MAP:
                            month_preview = MONTH_MAP[cell]
                            break
                        # substring fallback
                        for k, v in MONTH_MAP.items():
                            if len(k) >= 4 and k in cell:
                                month_preview = v
                                break
                        if month_preview:
                            break
                    if month_preview:
                        break

                # 2️⃣ Cari dari kolom "Bulan" jika format FLAT
                # (file sudah diproses sebelumnya, punya kolom Bulan)
                if not month_preview:
                    # Cari indeks kolom "Bulan" di baris header
                    header_row_vals = df_info.iloc[0].astype(str).str.upper().str.strip().tolist()
                    if "BULAN" in header_row_vals:
                        bulan_col_idx = header_row_vals.index("BULAN")
                        # Cari nilai di kolom tersebut di baris data
                        for r in range(1, min(df_info.shape[0], 10)):
                            cell = str(df_info.iloc[r, bulan_col_idx]).upper().strip()
                            if cell in MONTH_MAP:
                                month_preview = MONTH_MAP[cell]
                                break

                # 3️⃣ Fallback: cari di nama file
                if not month_preview:
                    fname = uploaded_file_kppn.name.upper()
                    for k, v in MONTH_MAP.items():
                        if len(k) >= 4 and k in fname:
                            month_preview = v
                            break

                # 4️⃣ Final fallback
                if not month_preview:
                    month_preview = "UNKNOWN"

                period_key_preview = (month_preview, str(upload_year_kppn))

                # ===============================
                # ℹ️ INFO / KONFIRMASI
                # ===============================
                if period_key_preview in st.session_state.data_storage_kppn:
                    st.warning(
                        f"Data IKPA KPPN **{month_preview} {upload_year_kppn}** sudah ada."
                    )
                    confirm_replace = st.checkbox(
                        "Ganti data yang sudah ada",
                        key=f"confirm_replace_kppn_{month_preview}_{upload_year_kppn}"
                    )
                else:
                    confirm_replace = True
                    st.info(
                        f"Akan mengunggah Data IKPA KPPN "
                        f"untuk periode **{month_preview} {upload_year_kppn}**"
                    )

            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
                confirm_replace = False


            # ===============================
            # 🔄 PROSES DATA
            # ===============================
            if st.button(
                " Proses Data IKPA KPPN",
                type="primary",
                disabled=not confirm_replace,
                key="proses_kppn"
            ):

                if uploaded_file_kppn is None:
                    st.error("Silakan upload file terlebih dahulu")
                    st.stop()

                with st.spinner("Memproses data IKPA KPPN..."):

                    df_processed, month, year = process_excel_file_kppn(
                        uploaded_file_kppn,
                        upload_year_kppn,
                        month_preview
                    )


                    if df_processed is None:
                        st.error(" Gagal memproses file IKPA KPPN.")
                        st.stop()

                    period_key = (str(month), str(year))
                    filename = f"IKPA_KPPN_{month}_{year}.xlsx"

                    try:
                        #  Simpan ke session
                        st.session_state.data_storage_kppn[period_key] = df_processed

                        #  Simpan ke GitHub
                        excel_bytes = io.BytesIO()
                        with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                            df_processed.drop(
                                ["Bobot", "Nilai Terbobot"],
                                axis=1,
                                errors="ignore"
                            ).to_excel(
                                writer,
                                index=False,
                                sheet_name="Data IKPA KPPN"
                            )
                        excel_bytes.seek(0)

                        save_file_to_github(
                            excel_bytes.getvalue(),
                            filename,
                            folder=f"Data IKPA KPPN/{year}"
                        )
                        
                        log_activity(
                            menu="Upload Data",
                            action="Upload IKPA KPPN",
                            detail=f"{uploaded_file_kppn.name} | {month} {year}"
                        )

                        st.success(
                            f" Data IKPA KPPN {month} {year} berhasil disimpan."
                        )
                        st.snow()

                    except Exception as e:
                        st.error(f" Gagal menyimpan ke GitHub: {e}")
                        
            
        # ============================================================
        # SUBMENU: UPLOAD DATA DIPA
        # ============================================================
        st.markdown("---")
        st.subheader("📤 Upload Data DIPA")

        # ===============================
        # PILIH TAHUN DIPA
        # ===============================
        selected_year_dipa = st.selectbox(
            "📅 Pilih Tahun DIPA",
            list(range(2022, 2031)),  # 2022 - 2030
            index=3
        )

        st.caption(
            "Sistem dapat memproses Data DIPA yang bersumber dari :\n"
            "1. Aplikasi OM-SPAN → menu Penganggaran → Informasi Revisi DIPA\n"
            "2. Aplikasi MyIntress → menu Anggaran → Download Data Detil"
        )

        uploaded_dipa_file = st.file_uploader(
            "Pilih file Excel DIPA (mentah dari SAS/SMART/Kemenkeu)",
            type=['xlsx', 'xls'],
            key="upload_dipa"
        )

        # Tombol proses DIPA
        if uploaded_dipa_file is not None:
            if st.button("🔄 Proses Data DIPA", type="primary"):
                with st.spinner("Memproses data DIPA..."):

                    # ===============================
                    # INIT STORAGE (JANGAN RESET!)
                    # ===============================
                    if "DATA_DIPA_by_year" not in st.session_state:
                        st.session_state.DATA_DIPA_by_year = {}

                    st.session_state.ikpa_dipa_merged = False
                    st.session_state["_just_uploaded_dipa"] = True

                    try:
                        # ===============================
                        # PROSES FILE
                        # ===============================
                        df_clean, _, status_msg = process_uploaded_dipa(
                            uploaded_dipa_file,
                            save_file_to_github,
                            forced_year=selected_year_dipa
                        )
                        
                        st.write("Preview DIPA:", df_clean.head())
                        st.write("Kolom DIPA:", df_clean.columns.tolist())

                        if df_clean is None:
                            st.error(f"❌ Gagal memproses DIPA: {status_msg}")
                            st.stop()

                        # ===============================
                        # NORMALISASI KODE SATKER
                        # ===============================
                        df_clean["Kode Satker"] = (
                            df_clean["Kode Satker"]
                            .astype(str)
                            .apply(normalize_kode_satker)
                        )

                        # ===============================
                        # FIX UTAMA: PAKSA TAHUN DARI UI
                        # ===============================
                        tahun_dipa = int(selected_year_dipa)

                        # KUNCI TAHUN DI DATA (ANTI BALIK KE FILE)
                        df_clean["Tahun"] = tahun_dipa

                        # ===============================
                        # SIMPAN KE SESSION
                        # ===============================
                        st.session_state.DATA_DIPA_by_year[tahun_dipa] = df_clean.copy()
                        

                        # ===============================
                        # SIMPAN KE GITHUB
                        # ===============================
                        excel_bytes = io.BytesIO()
                        with pd.ExcelWriter(excel_bytes, engine='openpyxl') as writer:
                            df_clean.to_excel(
                                writer,
                                index=False,
                                sheet_name=f"DIPA_{tahun_dipa}"
                            )

                        excel_bytes.seek(0)

                        save_file_to_github(
                            excel_bytes.getvalue(),
                            f"DIPA_{tahun_dipa}.xlsx",
                            folder="DATA_DIPA"
                        )

                        # ===============================
                        # LOG
                        # ===============================
                        log_activity(
                            menu="Upload Data",
                            action="Upload Data DIPA",
                            detail=f"Tahun {tahun_dipa} | {len(df_clean)} satker"
                        )

                        # ===============================
                        # OUTPUT
                        # ===============================
                        st.success(f"✅ Data DIPA tahun {tahun_dipa} berhasil diproses & disimpan.")
                        st.dataframe(df_clean.head(10), use_container_width=True)

                    except Exception as e:
                        st.error(f"❌ Terjadi error saat memproses file DIPA: {e}")
                        
                        
        # ===============================
        # SUBMENU UPLOAD DATA KKP
        # ===============================
        st.markdown("---")
        st.subheader("💳 Upload Data KKP")

        upload_year_kkp = st.selectbox(
            "Pilih Tahun Data KKP",
            list(range(2020, 2031)),
            index=list(range(2020, 2031)).index(datetime.now().year),
            key="tahun_kkp"
        )

        uploaded_file_kkp = st.file_uploader(
            "Pilih file Excel Data KKP",
            type=["xlsx", "xls"],
            key="file_kkp"
        )

        file_valid = False

        # ===============================
        # PREVIEW VALIDASI
        # ===============================
        if uploaded_file_kkp is not None:
            try:
                df_preview = process_excel_file_kkp(uploaded_file_kkp)

                if df_preview.empty:
                    st.warning("⚠️ Struktur file tidak dikenali atau data kosong.")
                else:
                    st.success(f"File terdeteksi ({len(df_preview)} baris). Klik proses untuk menyimpan.")
                    file_valid = True

            except Exception as e:
                st.error(f"Gagal membaca file KKP: {e}")
                file_valid = False


        # ===============================
        # PROSES DATA FINAL
        # ===============================
        if st.button(
            "Proses Data KKP",
            type="primary",
            disabled=not file_valid,
            key="proses_kkp"
        ):

            with st.spinner("Memproses data KKP..."):

                try:
                    df_kkp = process_excel_file_kkp(uploaded_file_kkp)

                    if df_kkp.empty:
                        st.error("Data KKP kosong setelah diproses.")
                        st.stop()

                    UNIQUE_KEY = ["PERIODE", "Kode Satker", "JENIS KKP"]
                    SPM_COL = "NILAI TRANSAKSI (NILAI SPM)"

                    # Validasi kolom wajib
                    for col in UNIQUE_KEY:
                        if col not in df_kkp.columns:
                            st.error(f"Kolom {col} tidak ditemukan.")
                            st.stop()

                    # Pastikan numeric
                    df_kkp[SPM_COL] = pd.to_numeric(
                        df_kkp[SPM_COL],
                        errors="coerce"
                    ).fillna(0)

                    # Ambil master dari session
                    master_df = st.session_state.get("kkp_master", pd.DataFrame())

                    # ===============================
                    # NORMALISASI SUPER AMAN
                    # ===============================
                    for df in [df_kkp, master_df]:
                        if not df.empty:

                            df["PERIODE"] = (
                                df["PERIODE"]
                                .astype(str)
                                .str.replace(".0", "", regex=False)
                                .str.strip()
                                .str.upper()
                            )

                            # ===============================
                            # NORMALISASI KODE SATKER
                            # ===============================
                            df["Kode Satker"] = (
                                pd.to_numeric(df["Kode Satker"], errors="coerce")
                                .fillna(0)
                                .astype(int)
                                .astype(str)
                                .str.zfill(6)
                            )

                            df["JENIS KKP"] = (
                                df["JENIS KKP"]
                                .astype(str)
                                .str.strip()
                                .str.upper()
                            )

                    # =====================================================
                    # UPLOAD PERTAMA (MASTER KOSONG)
                    # =====================================================
                    if master_df.empty:

                        final_df = df_kkp.copy()
                        new_count = len(df_kkp)
                        update_count = 0

                        st.info("Master kosong → langsung jadi database utama")

                    # =====================================================
                    # SUDAH ADA MASTER → MERGE CERDAS
                    # =====================================================
                    else:

                        master_df[SPM_COL] = pd.to_numeric(
                            master_df[SPM_COL],
                            errors="coerce"
                        ).fillna(0)

                        # Gabungkan upload ke master untuk cek
                        merged = df_kkp.merge(
                            master_df,
                            on=UNIQUE_KEY,
                            how="left",
                            indicator=True,
                            suffixes=("", "_old")
                        )

                        # ===============================
                        # DATA BARU
                        # ===============================
                        new_rows = merged[merged["_merge"] == "left_only"]
                        new_count = len(new_rows)

                        # ===============================
                        # DATA SUDAH ADA → CEK SPM
                        # ===============================
                        existing_rows = merged[merged["_merge"] == "both"]

                        updated_rows = []

                        for _, row in existing_rows.iterrows():

                            key_filter = (
                                (master_df["PERIODE"] == row["PERIODE"]) &
                                (master_df["Kode Satker"] == row["Kode Satker"]) &
                                (master_df["JENIS KKP"] == row["JENIS KKP"])
                            )

                            master_row = master_df.loc[key_filter]

                            if not master_row.empty:

                                master_spm = float(master_row.iloc[0][SPM_COL])
                                upload_spm = float(row[SPM_COL])

                                # Ambil yang SPM lebih besar
                                if upload_spm > master_spm:
                                    updated_rows.append(row[df_kkp.columns])

                        update_count = len(updated_rows)

                        # ===============================
                        # HAPUS DATA LAMA YANG DIUPDATE
                        # ===============================
                        if update_count > 0:

                            update_df = pd.DataFrame(updated_rows)

                            master_df = master_df.merge(
                                update_df[UNIQUE_KEY],
                                on=UNIQUE_KEY,
                                how="left",
                                indicator=True
                            )

                            master_df = master_df[master_df["_merge"] == "left_only"]
                            master_df = master_df.drop(columns=["_merge"])

                            master_df = pd.concat([master_df, update_df], ignore_index=True)

                        # ===============================
                        # TAMBAH DATA BARU
                        # ===============================
                        if new_count > 0:
                            master_df = pd.concat(
                                [master_df, new_rows[df_kkp.columns]],
                                ignore_index=True
                            )

                        final_df = master_df.copy()

                    # ===============================
                    # UPDATE SESSION
                    # ===============================
                    st.session_state.kkp_master = final_df.reset_index(drop=True)

  

                    # =====================================================
                    # SIMPAN KE GITHUB (1 FILE MASTER)
                    # =====================================================
                    filename = "KKP_MASTER.xlsx"

                    excel_bytes = io.BytesIO()
                    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                        final_df.to_excel(
                            writer,
                            index=False,
                            sheet_name="Data KKP"
                        )

                    excel_bytes.seek(0)

                    save_file_to_github(
                        excel_bytes.getvalue(),
                        filename,
                        folder="data_kkp"
                    )

                    st.success(
                        f"✅ Proses selesai | "
                        f"{new_count} data baru | "
                        f"{update_count} data diperbarui"
                    )

                    st.snow()

                except Exception as e:
                    st.error(f"Gagal memproses atau menyimpan data KKP: {e}")
            
            
        # ============================================================
        # SUBMENU: Upload Data Referensi
        # ============================================================
        st.markdown("---")
        st.subheader("📚 Upload / Perbarui Data Referensi Satker & K/L")
        st.info("""
        - File referensi ini berisi kolom: **Kode BA, K/L, Kode Satker, Uraian Satker-SINGKAT, Uraian Satker-LENGKAP**  
        - Saat diupload, sistem akan **menggabungkan** dengan data lama:  
        🔹 Jika `Kode Satker` sudah ada → baris lama akan **diganti**  
        🔹 Jika `Kode Satker` belum ada → akan **ditambahkan baru**
        """)

        uploaded_ref = st.file_uploader(
            "📤 Pilih File Data Referensi Satker & K/L",
            type=['xlsx', 'xls'],
            key="ref_upload_file_admin"
        )


        if uploaded_ref is not None:
            try:
                new_ref = pd.read_excel(uploaded_ref)
                new_ref.columns = [c.strip() for c in new_ref.columns]

                required = [
                    'Kode BA', 'K/L', 'Kode Satker',
                    'Uraian Satker-SINGKAT', 'Uraian Satker-LENGKAP'
                ]
                if not all(col in new_ref.columns for col in required):
                    st.error("❌ Kolom wajib tidak lengkap dalam file referensi.")
                    st.stop()

                # Normalisasi Kode Satker
                new_ref['Kode Satker'] = new_ref['Kode Satker'].apply(normalize_kode_satker)

                # ============================================================
                # MERGE DENGAN REFERENSI LAMA
                # ============================================================
                if 'reference_df' in st.session_state and st.session_state.reference_df is not None:
                    old_ref = st.session_state.reference_df.copy()

                    if 'Kode Satker' in old_ref.columns:
                        old_ref['Kode Satker'] = old_ref['Kode Satker'].apply(normalize_kode_satker)

                    merged = pd.concat([old_ref, new_ref], ignore_index=True)
                    merged = merged.drop_duplicates(subset=['Kode Satker'], keep='last')
                    merged['Kode Satker'] = merged['Kode Satker'].astype(str).str.strip()

                    st.session_state.reference_df = merged
                    st.success(f"✅ Data Referensi diperbarui ({len(merged)} total baris).")
                else:
                    st.session_state.reference_df = new_ref
                    add_notification(f"✅ Data Referensi baru dimuat ({len(new_ref)} baris).")

                # ============================================================
                # 🔄 RE-APPLY REFERENSI KE SEMUA DATA IKPA (INI KUNCINYA)
                # ============================================================
                if "data_storage" in st.session_state:
                    new_storage = {}
                    for key, df in st.session_state.data_storage.items():
                        df = apply_reference_short_names(df)
                        df = create_satker_column(df)
                        new_storage[key] = df
                    st.session_state.data_storage = new_storage

                # ============================================================
                # SIMPAN REFERENSI KE GITHUB
                # ============================================================
                try:
                    # ===============================
                    # 1️⃣ LOAD FILE REFERENSI DARI GITHUB
                    # ===============================
                    token = st.secrets["GITHUB_TOKEN"]
                    repo_name = st.secrets["GITHUB_REPO"]

                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)

                    file_path = "templates/Template_Data_Referensi.xlsx"

                    existing_file = repo.get_contents(file_path)
                    file_content = base64.b64decode(existing_file.content)

                    df_existing = pd.read_excel(io.BytesIO(file_content))

                    # ===============================
                    # 2️⃣ TAMBAH ROW BARU
                    # ===============================
                    next_no = len(df_existing) + 1

                    new_row = pd.DataFrame([{
                        "No": next_no,
                        "Kode BA": kode_ba,
                        "K/L": kl,
                        "Kode Satker": kode_satker,
                        "Uraian Satker-SINGKAT": satker_singkat,
                        "Uraian Satker-LENGKAP": satker_lengkap
                    }])

                    df_updated = pd.concat([df_existing, new_row], ignore_index=True)

                    # ===============================
                    # 3️⃣ SIMPAN ULANG (REPLACE FILE YANG SAMA)
                    # ===============================
                    excel_bytes = io.BytesIO()

                    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                        df_updated.to_excel(writer, index=False)

                    excel_bytes.seek(0)

                    repo.update_file(
                        file_path,
                        "Update referensi (manual input)",
                        excel_bytes.getvalue(),
                        existing_file.sha
                    )

                    st.success("✅ Data berhasil ditambahkan dan file referensi diperbarui di GitHub")
                    st.snow()

                except Exception as e:
                    st.error(f"Gagal update referensi: {e}")

                # ============================================================
                # 🔁 CLEAR CACHE & RERUN (WAJIB)
                # ============================================================
                st.cache_data.clear()
                st.rerun()

            except Exception as e:
                st.error(f"❌ Gagal memproses Data Referensi: {e}")
    
    
    
        #UPLOAD MANUAL DATA REFERENSI
        st.markdown("---")
        st.markdown("## Input Data Referensi Manual")

        with st.form("admin_form_referensi_manual", clear_on_submit=True):

            col1, col2 = st.columns(2)

            with col1:
                kode_ba = st.text_input("Kode BA", key="admin_kode_ba")
                kl = st.text_input("K/L", key="admin_kl")
                kode_satker = st.text_input("Kode Satker", key="admin_kode_satker")

            with col2:
                satker_singkat = st.text_input("Uraian Satker-SINGKAT", key="admin_singkat")
                satker_lengkap = st.text_area("Uraian Satker-LENGKAP", key="admin_lengkap")

            submitted = st.form_submit_button("Simpan Data Referensi")

            if submitted:

                if not kode_satker or not satker_lengkap:
                    st.warning("Kode Satker dan Uraian Satker-LENGKAP wajib diisi.")
                    st.stop()

                try:
                    # ===============================
                    # LOAD FILE TEMPLATE DARI GITHUB
                    # ===============================
                    token = st.secrets["GITHUB_TOKEN"]
                    repo_name = st.secrets["GITHUB_REPO"]

                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)

                    file_path = "templates/Template_Data_Referensi.xlsx"
                    existing_file = repo.get_contents(file_path)

                    file_content = base64.b64decode(existing_file.content)
                    df_existing = pd.read_excel(io.BytesIO(file_content), dtype=str)

                    df_existing["Kode Satker"] = df_existing["Kode Satker"].astype(str)

                    # ===============================
                    # CEK DUPLIKASI LANGSUNG DI FILE
                    # ===============================
                    if kode_satker in df_existing["Kode Satker"].values:
                        st.warning("Kode Satker sudah ada dalam template.")
                        st.stop()

                    # ===============================
                    # AUTO NOMOR
                    # ===============================
                    df_existing["No"] = pd.to_numeric(df_existing["No"], errors="coerce")

                    max_no = df_existing["No"].max()

                    if pd.isna(max_no):
                        next_no = 1
                    else:
                        next_no = int(max_no) + 1

                    new_row = pd.DataFrame([{
                        "No": next_no,
                        "Kode BA": kode_ba,
                        "K/L": kl,
                        "Kode Satker": kode_satker,
                        "Uraian Satker-SINGKAT": satker_singkat,
                        "Uraian Satker-LENGKAP": satker_lengkap
                    }])

                    df_updated = pd.concat([df_existing, new_row], ignore_index=True)

                    # ===============================
                    # PASTIKAN KOLOM "No" DI PALING KIRI
                    # ===============================
                    cols = df_updated.columns.tolist()

                    if "No" in cols:
                        cols.remove("No")
                        df_updated = df_updated[["No"] + cols]


                    # ===============================
                    # UPDATE FILE YANG SAMA
                    # ===============================
                    excel_bytes = io.BytesIO()
                    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                        df_updated.to_excel(writer, index=False)

                    excel_bytes.seek(0)

                    repo.update_file(
                        file_path,
                        f"Tambah referensi manual: {kode_satker}",
                        excel_bytes.getvalue(),
                        existing_file.sha
                    )

                    st.success("✅ Data berhasil ditambahkan ke template dan diperbarui di GitHub")
                    st.snow()
                    st.rerun()

                except Exception as e:
                    st.error(f"Gagal update template: {e}")
                    
    
        st.markdown("---")
        st.subheader("Upload Data Digipay")

        # =========================================
        # INIT STORAGE
        # =========================================
        if "digipay_master" not in st.session_state:
            st.session_state.digipay_master = pd.DataFrame()

        uploaded_digipay = st.file_uploader(
            "Upload File Excel Digipay (Multi Sheet)",
            type=["xlsx"],
            key="upload_digipay_admin"
        )

        if uploaded_digipay:

            with st.spinner("Memproses Data Digipay..."):

                xls = pd.ExcelFile(uploaded_digipay)
                all_sheets = []

                for sheet in xls.sheet_names:
                    df_sheet = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df_sheet["SOURCE_SHEET"] = sheet
                    all_sheets.append(df_sheet)

                df_all = pd.concat(all_sheets, ignore_index=True)

                # ====================================
                # NORMALISASI KOLOM
                # ====================================
                df_all.columns = (
                    df_all.columns.astype(str)
                    .str.strip()
                    .str.upper()
                )
                
                # ==========================================
                # 🔥 POTONG KOLOM SEBELUM "TAHUN"
                # ==========================================

                if "TAHUN" in df_all.columns:
                    start_index = df_all.columns.get_loc("TAHUN")
                    df_all = df_all.iloc[:, start_index:]
                    
        
                # ====================================
                #  PERBAIKI LEADING ZERO OTOMATIS
                # ====================================

                # KDKANWIL → 2 digit
                if "KDKANWIL" in df_all.columns:
                    df_all["KDKANWIL"] = (
                        df_all["KDKANWIL"]
                        .astype(str)
                        .str.replace(".0", "", regex=False)
                        .str.strip()
                        .str.zfill(2)
                    )

                # KDKPPN → 3 digit
                if "KDKPPN" in df_all.columns:
                    df_all["KDKPPN"] = (
                        df_all["KDKPPN"]
                        .astype(str)
                        .str.replace(".0", "", regex=False)
                        .str.strip()
                        .str.zfill(3)
                    )

                # KDSATKER → 6 digit
                if "KDSATKER" in df_all.columns:
                    df_all["KDSATKER"] = (
                        df_all["KDSATKER"]
                        .astype(str)
                        .str.replace(".0", "", regex=False)
                        .str.strip()
                        .str.zfill(6)
                    )

                # ==========================================
                # AMBIL HANYA KOLOM RESMI DIGIPAY
                # ==========================================

                valid_columns = [
                    "TAHUN","KDKANWIL","NMKANWIL","KDKPPN","NMKPPN",
                    "KDSATKER","NMSATKER","NOINVOICE","NOMINVOICE",
                    "NMVENDOR","STSBAYAR","TGLBAYAR","BULAN",
                    "TGLINVOICE","KATEGORI","BANK_SATKER",
                    "BANK_VENDOR","SUBKATEGORI","CARA BAYAR"
                ]

                df_all = df_all[[col for col in valid_columns if col in df_all.columns]]
                
                
                # ====================================
                # FILTER OTOMATIS KPPN 109 BATURAJA
                # ====================================
                df_all = df_all[
                    (df_all["KDKPPN"] == "109") &
                    (df_all["NMKPPN"].str.upper() == "BATURAJA")
                ]
                
                # ====================================
                # FILTER HANYA STATUS SUDAH DIBAYAR
                # ====================================
                if "STSBAYAR" in df_all.columns:

                    df_all["STSBAYAR"] = (
                        df_all["STSBAYAR"]
                        .fillna("")
                        .astype(str)
                        .str.upper()
                        .str.strip()
                    )

                    df_all = df_all[
                        df_all["STSBAYAR"].str.contains("SUDAH", na=False)
                    ]

                df_all = df_all[
                    df_all["NOINVOICE"].notna() &
                    (df_all["NOINVOICE"].astype(str).str.strip() != "")
                ]
                
                # ====================================
                # UNIQUE KEY
                # ====================================
                UNIQUE_KEY = [
                    "TAHUN",
                    "KDSATKER",
                    "NOINVOICE",
                    "NOMINVOICE",
                    "TGLINVOICE"
                ]

                df_all = df_all.drop_duplicates(subset=UNIQUE_KEY)

                # ====================================
                # SMART MERGE UPDATE DATABASE
                # ====================================

                # Jika database kosong → langsung simpan
                if st.session_state.digipay_master.empty:
                    st.session_state.digipay_master = df_all.copy()
                    new_count = len(df_all)
                    update_count = 0

                else:
                    master_df = st.session_state.digipay_master.copy()

                    merged = df_all.merge(
                        master_df,
                        on=UNIQUE_KEY,
                        how="left",
                        indicator=True,
                        suffixes=("", "_old")
                    )

                    # Data baru
                    new_rows = merged[merged["_merge"] == "left_only"]
                    new_count = len(new_rows)

                    # Data sudah ada → cek perubahan
                    existing_rows = merged[merged["_merge"] == "both"]
                    updated_rows = []

                    for _, row in existing_rows.iterrows():

                        key_filter = (
                            (master_df["KDSATKER"] == row["KDSATKER"]) &
                            (master_df["NOINVOICE"] == row["NOINVOICE"]) &
                            (master_df["NOMINVOICE"] == row["NOMINVOICE"]) &
                            (master_df["TGLINVOICE"] == row["TGLINVOICE"])
                        )

                        master_row = master_df.loc[key_filter]

                        if not master_row.empty:
                            master_row = master_row.iloc[0]

                            different = False
                            for col in df_all.columns:
                                if col not in UNIQUE_KEY:
                                    if str(master_row[col]) != str(row[col]):
                                        different = True
                                        break

                            if different:
                                updated_rows.append(row[df_all.columns])

                    update_count = len(updated_rows)

                    # Hapus data lama yang akan diupdate
                    if update_count > 0:
                        update_df = pd.DataFrame(updated_rows)

                        master_df = master_df.merge(
                            update_df[UNIQUE_KEY],
                            on=UNIQUE_KEY,
                            how="left",
                            indicator=True
                        )

                        master_df = master_df[master_df["_merge"] == "left_only"]
                        master_df = master_df.drop(columns=["_merge"])

                        master_df = pd.concat([master_df, update_df], ignore_index=True)

                    # Tambahkan data baru
                    if new_count > 0:
                        master_df = pd.concat(
                            [master_df, new_rows[df_all.columns]],
                            ignore_index=True
                        )

                    st.session_state.digipay_master = master_df.copy()
                    
                    # ==========================================
                    # BERSIHKAN DATABASE DIGIPAY FINAL
                    # ==========================================

                    clean_df = st.session_state.digipay_master.copy()

                    # 1️⃣ Potong mulai dari kolom TAHUN
                    if "TAHUN" in clean_df.columns:
                        start_index = clean_df.columns.get_loc("TAHUN")
                        clean_df = clean_df.iloc[:, start_index:]

                    # 2️⃣ Drop kolom UNNAMED kalau masih ada
                    clean_df = clean_df.loc[:, ~clean_df.columns.str.contains("UNNAMED", case=False)]

                    # 3️⃣ Drop kolom kosong total
                    clean_df = clean_df.dropna(axis=1, how="all")

                    # Simpan kembali yang sudah bersih
                    st.session_state.digipay_master = clean_df.copy()

                # ====================================
                # SIMPAN OTOMATIS KE GITHUB
                # ====================================
                excel_bytes = io.BytesIO()

                with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                    st.session_state.digipay_master.to_excel(
                        writer,
                        index=False,
                        sheet_name="DIGIPAY_109_BATURAJA"
                    )

                excel_bytes.seek(0)

                save_file_to_github(
                    excel_bytes.getvalue(),
                    "DIGIPAY_MASTER.xlsx",
                    folder="data_Digipay"
                )

                log_activity(
                    menu="Upload Data",
                    action="Upload Data Digipay",
                    detail=f"{uploaded_digipay.name} | {new_count} baru | {update_count} diperbarui"
                )

                st.success(
                    f"✅ Upload selesai | {new_count} data baru | {update_count} data diperbarui"
                )
                

        # =========================================
        # AUTO LOAD CMS SAAT PAGE DIBUKA
        # =========================================
        if "cms_master" not in st.session_state:
            cms_df, _ = load_cms_from_github()
            st.session_state.cms_master = cms_df

        # ============================================================
        # UPLOAD DATA CMS 
        # ============================================================
        st.markdown("---")
        st.subheader("Upload Data CMS")

        # ============================================================
        # FILTER PERIODE (AUTO DEFAULT TAHUN & TRIWULAN)
        # ============================================================
        now = datetime.now()
        current_year = now.year
        current_month = now.month

        if current_month <= 3:
            default_tw = "TW1"
        elif current_month <= 6:
            default_tw = "TW2"
        elif current_month <= 9:
            default_tw = "TW3"
        else:
            default_tw = "TW4"

        st.markdown("### Filter Periode Data")
        col1, col2 = st.columns(2)

        with col1:
            selected_year = st.selectbox(
                "Pilih Tahun",
                options=list(range(2022, current_year + 1)),
                index=len(list(range(2022, current_year + 1))) - 1
            )

        with col2:
            selected_triwulan = st.selectbox(
                "Pilih Triwulan",
                options=["TW1", "TW2", "TW3", "TW4"],
                index=["TW1", "TW2", "TW3", "TW4"].index(default_tw)
            )

        st.session_state.selected_year = selected_year
        st.session_state.selected_triwulan = selected_triwulan


        uploaded_cms = st.file_uploader(
            "Upload File Excel CMS (Multi Sheet)",
            type=["xlsx"],
            key="upload_cms_admin"
        )

        # ============================================================
        # PROSES FILE
        # ============================================================
        if uploaded_cms:

            with st.spinner("Memproses Data CMS..."):

                xls = pd.ExcelFile(uploaded_cms)
                all_valid_data = []

                for sheet in xls.sheet_names:

                    df_raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=str)

                    header_row = None
                    for i in range(min(20, len(df_raw))):
                        row_text = " ".join(df_raw.iloc[i].astype(str)).upper()
                        if "SATKER" in row_text and "KANWIL" in row_text:
                            header_row = i
                            break

                    if header_row is None:
                        continue

                    df = pd.read_excel(
                        xls,
                        sheet_name=sheet,
                        header=header_row,
                        dtype=str
                    )

                    df.columns = (
                        df.columns.astype(str)
                        .str.replace("\n", " ")
                        .str.replace("\r", " ")
                        .str.strip()
                        .str.upper()
                    )

                    # Cari kolom KPPN
                    col_kppn = None
                    for col in df.columns:
                        test = (
                            df[col]
                            .astype(str)
                            .str.extract(r"(\d+)")[0]
                            .fillna("")
                            .str.zfill(3)
                        )
                        if (test == "109").sum() > 5:
                            col_kppn = col
                            df[col] = test
                            break

                    if not col_kppn:
                        continue

                    # Cari kolom Satker
                    col_satker = None
                    for col in df.columns:
                        test = (
                            df[col]
                            .astype(str)
                            .str.extract(r"(\d+)")[0]
                            .fillna("")
                            .str.zfill(6)
                        )
                        if test.str.len().eq(6).sum() > 5:
                            col_satker = col
                            df[col] = test
                            break

                    if not col_satker:
                        continue

                    df = df[df[col_kppn] == "109"]

                    if df.empty:
                        continue

                    df["TAHUN"] = selected_year
                    df["TRIWULAN"] = selected_triwulan
                    df["SOURCE_SHEET"] = sheet

                    all_valid_data.append(df)

                if not all_valid_data:
                    st.error("❌ Tidak ada data CMS KPPN 109 ditemukan.")
                    st.stop()

                df_final = pd.concat(all_valid_data, ignore_index=True)

                # ============================================================
                # 🔥 SMART OVERWRITE CMS MASTER
                # ============================================================
                UNIQUE_KEY = col_satker
                df_final = df_final.drop_duplicates(subset=[UNIQUE_KEY])

                if st.session_state.cms_master.empty:
                    st.session_state.cms_master = df_final.copy()
                    new_count = len(df_final)
                    overwrite_count = 0
                else:
                    master_df = st.session_state.cms_master.copy()

                    overwrite_mask = master_df[UNIQUE_KEY].isin(df_final[UNIQUE_KEY])
                    overwrite_count = overwrite_mask.sum()

                    master_df = master_df[~overwrite_mask]

                    master_df = pd.concat(
                        [master_df, df_final],
                        ignore_index=True
                    )

                    new_count = len(df_final) - overwrite_count
                    st.session_state.cms_master = master_df.copy()

                # ============================================================
                # SIMPAN KE GITHUB
                # ============================================================
                excel_bytes = io.BytesIO()

                sheet_name = f"CMS_109_{selected_triwulan}_{selected_year}"
                file_name = f"CMS_109_{selected_triwulan}_{selected_year}.xlsx"

                with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                    st.session_state.cms_master.to_excel(
                        writer,
                        index=False,
                        sheet_name=sheet_name
                    )

                excel_bytes.seek(0)

                save_file_to_github(
                    excel_bytes.getvalue(),
                    file_name,
                    folder="data_CMS"
                )

            st.success(
                f"✅ Upload CMS selesai | "
                f"{new_count} data baru | "
                f"{overwrite_count} data diperbarui | "
                f"Disimpan sebagai {file_name}"
            )


    # ============================================================
    # TAB 2: HAPUS DATA
    # ============================================================
    with tab2:
        # Submenu Hapus Data IKPA Satker
        st.subheader("🗑️ Hapus Data IKPA Satker")
        if not st.session_state.data_storage:
            st.info("ℹ️ Belum ada data IKPA tersimpan.")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_delete = st.selectbox(
                "Pilih periode yang akan dihapus",
                options=available_periods,
                format_func=lambda x: f"{x[0].capitalize()} {x[1]}"
            )
            month, year = period_to_delete
            filename = f"data/IKPA_{month}_{year}.xlsx"

            confirm_delete = st.checkbox(
                f"⚠️ Hapus data {month} {year} dari sistem dan GitHub.",
                key=f"confirm_delete_{month}_{year}"
            )

            if st.button("🗑️ Hapus Data IKPA Satker", type="primary") and confirm_delete:
                try:
                    del st.session_state.data_storage[period_to_delete]
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    contents = repo.get_contents(f"data/IKPA_{month}_{year}.xlsx")
                    repo.delete_file(contents.path, f"Delete {filename}", contents.sha)
                    st.success(f"✅ Data {month} {year} dihapus dari sistem & GitHub.")
                    st.snow()
                    st.session_state.activity_log.append({
                        "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Aksi": "Hapus IKPA",
                        "Periode": f"{month} {year}",
                        "Status": "✅ Sukses"
                    })
                except Exception as e:
                    st.error(f"❌ Gagal menghapus data: {e}")
                    
        # Submenu Hapus Data IKPA KPPN
        # ===========================
        # 🗑️ Hapus Data IKPA KPPN
        # ===========================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data IKPA KPPN")

        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]

            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            

            files_kppn = get_all_kppn_files(repo)
            

        except Exception:
            files_kppn = []

        if not files_kppn:
            st.info("ℹ️ Belum ada data IKPA KPPN tersimpan.")

        else:
            # tampilkan nama saja (tanpa path)
            display_files = {f.split("/")[-1]: f for f in files_kppn}

            selected_name = st.selectbox(
                "Pilih data IKPA KPPN yang akan dihapus",
                sorted(display_files.keys(), reverse=True)
            )

            selected_file = display_files[selected_name]

            confirm_delete = st.checkbox(
                f"⚠️ Saya yakin ingin menghapus **{selected_name}**"
            )

            if st.button("🗑️ Hapus Data IKPA KPPN", type="primary") and confirm_delete:
                try:
                    content = repo.get_contents(selected_file)

                    repo.delete_file(
                        content.path,
                        f"Delete {selected_name}",
                        content.sha
                    )

                    st.success(f"✅ {selected_name} berhasil dihapus.")
                    st.snow()
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Gagal menghapus: {e}")


        # Submenu Hapus Data DIPA
        st.markdown("---")
        st.subheader("🗑️ Hapus Data DIPA")
        if not st.session_state.get("DATA_DIPA_by_year"):
            st.info("ℹ️ Belum ada data DIPA tersimpan.")
        else:
            available_years = sorted(st.session_state.DATA_DIPA_by_year.keys(), reverse=True)
            year_to_delete = st.selectbox(
                "Pilih tahun DIPA yang akan dihapus",
                options=available_years,
                format_func=lambda x: f"Tahun {x}",
                key="delete_dipa_year"
            )
            filename_dipa = f"DATA_DIPA/DIPA_{year_to_delete}.xlsx"

            confirm_delete_dipa = st.checkbox(
                f"⚠️ Hapus data DIPA tahun {year_to_delete} dari sistem dan GitHub.",
                key=f"confirm_delete_dipa_{year_to_delete}"
            )

            if st.button("🗑️ Hapus Data DIPA Ini", type="primary", key="btn_delete_dipa") and confirm_delete_dipa:
                try:
                    del st.session_state.DATA_DIPA_by_year[year_to_delete]
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    contents = repo.get_contents(filename_dipa)
                    repo.delete_file(contents.path, f"Delete {filename_dipa}", contents.sha)
                    st.success(f"✅ Data DIPA tahun {year_to_delete} dihapus dari sistem & GitHub.")
                    st.snow()
                    st.session_state.activity_log.append({
                        "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Aksi": "Hapus DIPA",
                        "Periode": f"Tahun {year_to_delete}",
                        "Status": "✅ Sukses"
                    })
                except Exception as e:
                    st.error(f"❌ Gagal menghapus data DIPA: {e}")
        

        # ===============================
        # HAPUS DATA KKP 
        # ===============================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data KKP")

        if "kkp_master" not in st.session_state or st.session_state.kkp_master.empty:

            st.info("Belum ada data KKP yang tersimpan.")

        else:

            confirm_delete = st.checkbox(
                "Saya yakin ingin menghapus seluruh data KKP dari sistem dan GitHub.",
                key="confirm_delete_kkp_master"
            )

            if st.button(
                "🗑️ Hapus Data KKP",
                type="primary",
                disabled=not confirm_delete
            ):

                try:
                    # 🔹 Hapus dari session
                    st.session_state.kkp_master = pd.DataFrame()

                    # 🔹 Hapus dari GitHub
                    token = st.secrets["GITHUB_TOKEN"]
                    repo_name = st.secrets["GITHUB_REPO"]

                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)

                    file_path = "data_kkp/KKP_MASTER.xlsx"

                    try:
                        file = repo.get_contents(file_path)
                        repo.delete_file(
                            file.path,
                            "Delete KKP_MASTER.xlsx",
                            file.sha
                        )
                    except Exception as e:
                        st.warning(f"File tidak ditemukan di GitHub: {e}")

                    st.success("✅ Data KKP berhasil dihapus dari sistem & GitHub.")
                    st.snow()
                    st.rerun()

                except Exception as e:
                    st.error(f"Gagal menghapus data KKP: {e}")


                    
        
        # ============================================================
        # 🗑️ HAPUS DATA DIGIPAY
        # ============================================================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data Digipay")

        if "digipay_master" not in st.session_state or st.session_state.digipay_master.empty:

            st.info("ℹ️ Belum ada Data Digipay tersimpan.")

        else:

            confirm_delete_digipay = st.checkbox(
                "⚠️ Hapus seluruh Data Digipay dari sistem dan GitHub.",
                key="confirm_delete_digipay"
            )

            if st.button("🗑️ Hapus Data Digipay", type="primary") and confirm_delete_digipay:

                try:
                    # ======================================
                    # 1️⃣ Hapus dari session
                    # ======================================
                    st.session_state.digipay_master = pd.DataFrame()

                    # ======================================
                    # 2️⃣ Hapus dari GitHub
                    # ======================================
                    token = st.secrets["GITHUB_TOKEN"]
                    repo_name = st.secrets["GITHUB_REPO"]

                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)

                    file_path = "data_Digipay/DIGIPAY_MASTER.xlsx"

                    try:
                        file = repo.get_contents(file_path)
                        repo.delete_file(
                            file.path,
                            "Delete DIGIPAY_MASTER.xlsx",
                            file.sha
                        )
                    except Exception:
                        pass  # Jika file belum ada, tidak error

                    # ======================================
                    # 3️⃣ Log Aktivitas (optional)
                    # ======================================
                    if "activity_log" not in st.session_state:
                        st.session_state.activity_log = []

                    st.session_state.activity_log.append({
                        "Waktu": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Menu": "Hapus Data",
                        "Aksi": "Hapus Data Digipay",
                        "Status": "✅ Sukses"
                    })

                    st.success("✅ Data Digipay berhasil dihapus dari sistem & GitHub.")
                    st.snow()
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Gagal menghapus Data Digipay: {e}")
                    

        # ============================================================
        # 🗑️ HAPUS DATA CMS
        # ============================================================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data CMS")


        if st.session_state.get("cms_master") is None or st.session_state.cms_master.empty:
            st.info("ℹ️ Belum ada data CMS tersedia.")
        else:

            cms_df = st.session_state.cms_master.copy()

            # NORMALISASI
            cms_df["TAHUN"] = cms_df["TAHUN"].astype(int)
            cms_df["TRIWULAN"] = cms_df["TRIWULAN"].astype(str).str.upper()

            available_years = sorted(cms_df["TAHUN"].unique(), reverse=True)

            col1, col2 = st.columns(2)

            with col1:
                delete_year = st.selectbox(
                    "Pilih Tahun",
                    options=available_years,
                    key="delete_cms_year"
                )


            # ================================
            # DETEKSI TRIWULAN DARI GITHUB
            # ================================
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]

            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)

            contents = repo.get_contents("data_CMS")

            tw_options = []

            for file in contents:
                if file.name.endswith(f"_{delete_year}.xlsx") and "CMS_109_" in file.name:
                    tw = file.name.split("_")[2]
                    tw_options.append(tw)

            tw_options = sorted(list(set(tw_options)))

            with col2:
                delete_tw = st.selectbox(
                    "Pilih Triwulan",
                    options=tw_options,
                    key="delete_cms_tw"
                )

            confirm_delete_cms = st.checkbox(
                f"⚠️ Hapus Data CMS Tahun {delete_year} {delete_tw}",
                key="confirm_delete_cms"
            )

            if st.button("🗑️ Hapus Data CMS", key="btn_delete_cms") and confirm_delete_cms:

                try:
                    before = len(cms_df)

                    cms_df = cms_df[
                        ~(
                            (cms_df["TAHUN"] == delete_year) &
                            (cms_df["TRIWULAN"] == delete_tw)
                        )
                    ]

                    deleted_rows = before - len(cms_df)
                    st.session_state.cms_master = cms_df.copy()

                    # HAPUS FILE GITHUB
                    file_name = f"CMS_109_{delete_tw}_{delete_year}.xlsx"
                    file_path = f"data_CMS/{file_name}"

                    try:
                        file = repo.get_contents(file_path)
                        repo.delete_file(file.path, f"Delete {file_name}", file.sha)
                    except:
                        pass

                    st.success(f"✅ {deleted_rows} data CMS berhasil dihapus")
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ {e}")

        
        # HAPUS DATA REFERENSI
        # =====================================================
        st.markdown("---")
        st.markdown("## Hapus Data Referensi")

        try:
            # ===============================
            # 1️⃣ LOAD DATA DARI GITHUB
            # ===============================
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]

            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)

            file_path = "templates/Template_Data_Referensi.xlsx"
            existing_file = repo.get_contents(file_path)

            file_content = base64.b64decode(existing_file.content)
            df_referensi = pd.read_excel(io.BytesIO(file_content), dtype=str)

            if df_referensi.empty:
                st.info("Data referensi kosong.")
            else:
                # ===============================
                # 2️⃣ BUAT LABEL DROPDOWN
                # ===============================
                df_referensi["Kode Satker"] = df_referensi["Kode Satker"].astype(str)

                df_referensi["Label"] = (
                    df_referensi["Kode Satker"]
                    + " - "
                    + df_referensi["Uraian Satker-SINGKAT"].astype(str)
                )

                selected_label = st.selectbox(
                    "Pilih Satker yang akan dihapus",
                    df_referensi["Label"]
                )

                # ===============================
                # 3️⃣ TOMBOL HAPUS
                # ===============================
                if st.button("Hapus Data Referensi", type="primary"):

                    # Filter data
                    df_updated = df_referensi[
                        df_referensi["Label"] != selected_label
                    ].copy()

                    # Hapus kolom helper
                    df_updated = df_updated.drop(columns=["Label"])

                    # ===============================
                    # 4️⃣ RAPIKAN NOMOR
                    # ===============================
                    df_updated = df_updated.reset_index(drop=True)
                    df_updated["No"] = df_updated.index + 1

                    # ===============================
                    # 5️⃣ UPDATE FILE YANG SAMA
                    # ===============================
                    excel_bytes = io.BytesIO()

                    with pd.ExcelWriter(excel_bytes, engine="openpyxl") as writer:
                        df_updated.to_excel(writer, index=False)

                    excel_bytes.seek(0)

                    repo.update_file(
                        file_path,
                        f"Hapus referensi: {selected_label}",
                        excel_bytes.getvalue(),
                        existing_file.sha
                    )

                    st.success("✅ Data berhasil dihapus dan file template diperbarui")
                    st.rerun()

        except Exception as e:
            st.error(f"Gagal memuat atau menghapus referensi: {e}")
        
        # ==========================================
        # HAPUS DATABASE DIGIPAY
        # ==========================================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data Digipay")

        if "digipay_master" not in st.session_state or st.session_state.digipay_master.empty:

            st.info("ℹ️ Belum ada Data Digipay tersimpan.")

        else:

            confirm_delete = st.checkbox(
                "⚠️ Hapus seluruh Data Digipay dari sistem dan GitHub.",
                key="confirm_delete_digipay_admin"
            )

            if st.button(
                "🗑️ Hapus Data Digipay",
                type="primary",
                key="delete_digipay_year"
            ) and confirm_delete:

                try:
                    # ==========================================
                    # 1️⃣ Hapus dari session
                    # ==========================================
                    st.session_state.digipay_master = pd.DataFrame()

                    # ==========================================
                    # 2️⃣ Hapus dari GitHub
                    # ==========================================
                    token = st.secrets.get("GITHUB_TOKEN")
                    repo_name = st.secrets.get("GITHUB_REPO")

                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)

                    try:
                        contents = repo.get_contents(
                            "data_Digipay/DIGIPAY_MASTER.xlsx"
                        )

                        repo.delete_file(
                            contents.path,
                            "Delete DIGIPAY_MASTER.xlsx",
                            contents.sha
                        )

                    except Exception:
                        pass  # jika file tidak ada, abaikan

                    # ==========================================
                    # 3️⃣ Reset loader flag (agar sinkron)
                    # ==========================================
                    st.session_state.auto_loaded_digipay = False

                    st.success("✅ Data Digipay berhasil dihapus dari sistem & GitHub.")
                    st.snow()

                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Gagal menghapus Data Digipay: {e}")



    # ============================================================
    # TAB 3: DOWNLOAD DATA
    # ============================================================
    with tab3:
        st.subheader("📥 Download IKPA Satker")

        if "data_storage" not in st.session_state or not st.session_state.data_storage:
            st.info("🔹 Data belum tersedia untuk diunduh")
        else:
            available_periods = sorted(st.session_state.data_storage.keys(), reverse=True)
            period_to_download = st.selectbox(
                "Pilih periode untuk download",
                options=available_periods,
                format_func=lambda x: f"{x[0]} {x[1]}"
            )

            df_selected = st.session_state.data_storage.get(period_to_download)
            if df_selected is not None:
                filename = f"IKPA_{period_to_download[0]}_{period_to_download[1]}.xlsx"
                excel_bytes = to_excel_bytes(df_selected)  # pastikan fungsi ini sudah ada
                st.download_button(
                    label=f"Download {filename}",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
         
        # ===========================
        # 📥 Download Data IKPA KPPN (PROCESSED)
        # ===========================
        st.markdown("---")
        st.subheader("📥 Download Data IKPA KPPN")

        # pastikan session ada
        if "data_storage_kppn" not in st.session_state:
            st.session_state.data_storage_kppn = {}

        data_kppn = st.session_state.data_storage_kppn

        # ===============================
        # JIKA BELUM ADA DATA
        # ===============================
        if not data_kppn:
            st.info("ℹ️ Belum ada data IKPA KPPN tersedia untuk diunduh.")

        else:
            # ===============================
            # PILIH DATA
            # ===============================
            display_keys = {
                f"{bulan} {tahun}": (bulan, tahun)
                for (bulan, tahun) in data_kppn.keys()
            }

            selected_label = st.selectbox(
                "Pilih data IKPA KPPN",
                sorted(display_keys.keys(), reverse=True)
            )

            selected_key = display_keys[selected_label]

            df_download = data_kppn[selected_key]

            # ===============================
            # CONVERT KE EXCEL
            # ===============================
            output = io.BytesIO()

            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df_download.to_excel(writer, index=False)

            # ===============================
            # DOWNLOAD BUTTON
            # ===============================
            st.download_button(
                label="📥 Download File IKPA KPPN",
                data=output.getvalue(),
                file_name=f"IKPA_KPPN_{selected_key[0]}_{selected_key[1]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
                
        # ===========================
        # Submenu Download Data DIPA
        # ===========================
        st.markdown("---")
        st.markdown("### 📥 Download Data DIPA")

        if not st.session_state.get("DATA_DIPA_by_year"):
            st.info("ℹ️ Belum ada data DIPA.")
        else:
            available_years = sorted(st.session_state.DATA_DIPA_by_year.keys(), reverse=True)

            year_to_download = st.selectbox(
                "Pilih tahun DIPA",
                options=available_years,
                format_func=lambda x: f"Tahun {x}",
                key="download_dipa_year"
            )

            # Ambil data yang sudah bersih dari load()
            df = st.session_state.DATA_DIPA_by_year[year_to_download].copy()

            # Kolom yang ingin ditampilkan
            desired_columns = [
                "Kode Satker",
                "Satker",
                "Tahun",
                "Tanggal Posting Revisi",
                "Total Pagu",
                "Jenis Satker",
                "NO",
                "Kementerian",
                "Kode Status History",
                "Jenis Revisi",
                "Revisi ke-",
                "No Dipa",
                "Tanggal Dipa",
                "Owner",
                "Digital Stamp"
            ]

            # Filter kolom yang ada
            df = df[[c for c in desired_columns if c in df.columns]]

            # Ambil revisi terbaru
            if "Kode Satker" in df.columns and "Tanggal Posting Revisi" in df.columns:
                df["Tanggal Posting Revisi"] = pd.to_datetime(df["Tanggal Posting Revisi"], errors="coerce")
                df = df.sort_values(
                    by=["Kode Satker", "Tanggal Posting Revisi"],
                    ascending=[True, False]
                ).drop_duplicates(subset="Kode Satker", keep="first")

            # Klasifikasi Satker
            if "Total Pagu" in df.columns:
                df["Total Pagu"] = pd.to_numeric(df["Total Pagu"], errors="coerce").fillna(0)
                p40 = df["Total Pagu"].quantile(0.40)
                p70 = df["Total Pagu"].quantile(0.70)

                if df["Total Pagu"].sum() == 0 or p40 == p70:
                    # Rank-based jika semua pagu sama / tidak bisa dibagi bins
                    df["_rank"] = df["Total Pagu"].rank(method="first")
                    n = len(df)
                    cut40 = n * 0.40
                    cut70 = n * 0.70
                    df["Jenis Satker"] = df["_rank"].apply(
                        lambda r: "Satker Kecil" if r <= cut40 else ("Satker Sedang" if r <= cut70 else "Satker Besar")
                    )
                    df = df.drop(columns=["_rank"])
                else:
                    df["Jenis Satker"] = pd.cut(
                        df["Total Pagu"],
                        bins=[-float("inf"), p40, p70, float("inf")],
                        labels=["Satker Kecil", "Satker Sedang", "Satker Besar"]
                    )


            # Preview
            with st.expander("Preview Data"):
                st.dataframe(df.head(10), use_container_width=True)

            # Export Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=f"DIPA_{year_to_download}")

            output.seek(0)

            st.download_button(
                "📥 Download Excel DIPA",
                data=output,
                file_name=f"DIPA_{year_to_download}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            

        # DOWNLOAD DATA KKP
        st.markdown("---")
        st.subheader("Download Data KKP")

        if "kkp_master" not in st.session_state or st.session_state.kkp_master.empty:
            st.info("Belum ada data KKP yang tersimpan.")
        else:
            buffer = io.BytesIO()

            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                st.session_state.kkp_master.to_excel(
                    writer,
                    index=False,
                    sheet_name="Data KKP"
                )

            buffer.seek(0)

            st.download_button(
                label="⬇️ Download Data KKP",
                data=buffer,
                file_name="KKP_MASTER.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            

        # ==========================================
        #  DOWNLOAD DATABASE DIGIPAY
        # ==========================================
        st.markdown("---")
        st.subheader("📥 Download Data Digipay")

        if "digipay_master" in st.session_state and not st.session_state.digipay_master.empty:

            buffer = io.BytesIO()

            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                st.session_state.digipay_master.to_excel(
                    writer,
                    index=False,
                    sheet_name="DIGIPAY_109_BATURAJA"
                )

            st.download_button(
                label="Download Data Digipay",
                data=buffer.getvalue(),
                file_name="DIGIPAY_KPPN_109_BATURAJA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.info("ℹ️ Database Digipay kosong.")
            
        
        # ============================================================
        # DOWNLOAD DATA CMS 
        # ============================================================
        st.markdown("---")
        st.subheader("📥 Download Data CMS")

        if "cms_master" in st.session_state and not st.session_state.cms_master.empty:

            df_master = st.session_state.cms_master.copy()

            df_master["TAHUN"] = df_master["TAHUN"].astype(str)
            df_master["TRIWULAN"] = df_master["TRIWULAN"].astype(str)

            available_years = sorted(df_master["TAHUN"].unique())
            available_tw = sorted(df_master["TRIWULAN"].unique())

            col1, col2 = st.columns(2)

            with col1:
                selected_year_dl = st.selectbox(
                    "Pilih Tahun",
                    options=available_years,
                    key="download_cms_year"
                )

            with col2:
                selected_tw_dl = st.selectbox(
                    "Pilih Triwulan",
                    options=available_tw,
                    key="download_cms_tw"
                )

            df_download = df_master[
                (df_master["TAHUN"] == selected_year_dl) &
                (df_master["TRIWULAN"] == selected_tw_dl)
            ]

            st.caption(f"Jumlah data: {len(df_download)}")

            if not df_download.empty:

                buffer = io.BytesIO()

                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    df_download.to_excel(writer, index=False)

                buffer.seek(0)

                st.download_button(
                    label="⬇️ Download Data CMS",
                    data=buffer,
                    file_name=f"CMS_{selected_tw_dl}_{selected_year_dl}.xlsx",
                    key="download_cms_btn"
                )

            else:
                st.warning("⚠️ Tidak ada data untuk periode tersebut.")

        else:
            st.info("ℹ️ Belum ada data CMS tersedia.")
    


        # Download Data Satker Tidak Terdaftar
        st.markdown("---")
        st.subheader("📥 Download Data Satker yang Belum Terdaftar di Tabel Referensi")
        
        if st.button("📥 Generate & Download Laporan"):
            st.info("ℹ️ Fitur ini menggunakan data dari session state untuk performa optimal.")

        

    # ============================================================
    # TAB 4: DOWNLOAD TEMPLATE
    # ============================================================
    with tab4:
        st.subheader("📋 Download Template")
        st.markdown("### 📘 Template IKPA")
        try:
            token = st.secrets["GITHUB_TOKEN"]
            repo_name = st.secrets["GITHUB_REPO"]
            g = Github(auth=Auth.Token(token))
            repo = g.get_repo(repo_name)
            file_content = repo.get_contents("templates/Template_IKPA.xlsx")
            template_data = base64.b64decode(file_content.content)
        except Exception:
            template_data = get_template_file()

        if template_data:
            st.download_button(
                label="📥 Download Template IKPA",
                data=template_data,
                file_name="Template_IKPA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
        st.markdown("### 📗 Template Data Referensi Satker & K/L")

        # 🧩 Use latest reference data for template content
        if 'reference_df' in st.session_state and not st.session_state.reference_df.empty:
            template_ref = st.session_state.reference_df.copy()
        else:
            # fallback: try load from GitHub
            try:
                token = st.secrets["GITHUB_TOKEN"]
                repo_name = st.secrets["GITHUB_REPO"]
                g = Github(auth=Auth.Token(token))
                repo = g.get_repo(repo_name)
                ref_content = repo.get_contents("templates/Template_Data_Referensi.xlsx")
                ref_data = base64.b64decode(ref_content.content)
                template_ref = pd.read_excel(io.BytesIO(ref_data))
            except Exception:
                template_ref = pd.DataFrame({
                    'No': [],
                    'Kode BA': [],
                    'K/L': [],
                    'Kode Satker': [],
                    'Uraian Satker-SINGKAT': [],
                    'Uraian Satker-LENGKAP': []
                })

        output_ref = io.BytesIO()
        with pd.ExcelWriter(output_ref, engine='openpyxl') as writer:
            # ✅ PERBAIKAN: Mulai dari A1
            template_ref.to_excel(writer, index=False, sheet_name='Data Referensi',
                                  startrow=0, startcol=0)
            
            # Format header
            workbook = writer.book
            worksheet = writer.sheets['Data Referensi']
            
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        output_ref.seek(0)

        st.download_button(
            label="📥 Download Template Data Referensi",
            data=output_ref,
            file_name="Template_Data_Referensi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        
        st.subheader("📥 Download Data Referensi Terbaru")
        try:
            df_ref, repo, existing_file = load_template_referensi_from_github()

            add_notification("Data referensi berhasil dimuat dari GitHub")

            # ============================
            # CONVERT TO EXCEL (IN-MEMORY)
            # ============================
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_ref.to_excel(writer, index=False)

            output.seek(0)

            # ============================
            # DOWNLOAD BUTTON
            # ============================
            st.download_button(
                label="⬇️ Download Data Referensi (Terbaru)",
                data=output,
                file_name="Template_Data_Referensi_TERBARU.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Gagal memuat data referensi: {e}")

    
    # ===============================
    # 🕓 TAB RIWAYAT AKTIVITAS
    # ===============================
    with tab5:
        st.subheader("🕓 Riwayat Aktivitas Sistem")

        if not st.session_state.activity_log:
            st.info("Belum ada aktivitas yang tercatat.")
        else:
            df_log = pd.DataFrame(st.session_state.activity_log)

            st.dataframe(
                df_log,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Waktu": st.column_config.TextColumn(
                        "Waktu",
                        width="medium"
                    ),
                    "Aktivitas": st.column_config.TextColumn(
                        "Aktivitas",
                        width="medium"
                    ),
                    "Detail": st.column_config.TextColumn(
                        "Detail",
                        width="large"
                    ),
                }
            )

        st.divider()

        if st.button("🧹 Bersihkan Riwayat Aktivitas"):
            st.session_state.activity_log.clear()
            st.success("Riwayat aktivitas berhasil dibersihkan.")


def show_loading_logo():
    
    st.markdown(
        """
        <div style="display:flex;justify-content:center;align-items:center;height:70vh;">
        <img src="https://raw.githubusercontent.com/ameernoor/kinerjakeuangansatkerbta/main/logo_kppn_baturaja.png" width="500">
        </div>
        """,
        unsafe_allow_html=True
    )
    

st.markdown("""
<style>

.system-status{
    padding:20px;
    border-radius:10px;
    background:#f8fafc;
    border:1px solid #e2e8f0;
    margin-bottom:20px;
}

.system-title{
    font-size:18px;
    font-weight:600;
    margin-bottom:10px;
}

.status-item{
    padding:8px 0;
    font-size:14px;
    color:#334155;
    border-bottom:1px solid #f1f5f9;
}

.status-item:last-child{
    border-bottom:none;
}

.status-ok{
    color:#16a34a;
    font-weight:600;
}

</style>
""", unsafe_allow_html=True)
    
    
# ===============================
# MAIN APP
# ===============================
def main():

    # ============================================================
    # FLAG: apakah ini pertama kali app dibuka (bukan navigasi menu)
    # ============================================================
    is_first_load = "app_fully_loaded" not in st.session_state

    # flag untuk panel status sistem (hanya tampil saat pertama buka)
    if "show_system_status" not in st.session_state:
        st.session_state.show_system_status = True

    loading_placeholder = st.empty()

    # ============================================================
    # LOGO & SEMUA PROSES LOADING — HANYA SAAT PERTAMA KALI BUKA
    # Saat pindah menu, seluruh blok ini di-skip → navigasi cepat
    # ============================================================
    if is_first_load:

        # Tampilkan logo splash screen
        with loading_placeholder.container():
            show_loading_logo()

        # --------------------------------------------------------
        # 1. LOAD REFERENCE DATA
        # --------------------------------------------------------
        if "reference_df" not in st.session_state:

            token = st.secrets.get("GITHUB_TOKEN")
            repo_name = st.secrets.get("GITHUB_REPO")

            if not token or not repo_name:
                st.session_state.reference_df = pd.DataFrame({
                    'Kode BA': [], 'K/L': [], 'Kode Satker': [],
                    'Uraian Satker-SINGKAT': [], 'Uraian Satker-LENGKAP': []
                })
            else:
                try:
                    g = Github(auth=Auth.Token(token))
                    repo = g.get_repo(repo_name)
                    ref_path = "templates/Template_Data_Referensi.xlsx"

                    ref_file = repo.get_contents(ref_path)
                    ref_data = base64.b64decode(ref_file.content)

                    ref_df = pd.read_excel(io.BytesIO(ref_data))
                    ref_df.columns = [c.strip() for c in ref_df.columns]

                    st.session_state.reference_df = ref_df

                except Exception:
                    st.session_state.reference_df = pd.DataFrame({
                        'Kode BA': [], 'K/L': [], 'Kode Satker': [],
                        'Uraian Satker-SINGKAT': [], 'Uraian Satker-LENGKAP': []
                    })

        # --------------------------------------------------------
        # 2. LOAD IKPA
        # --------------------------------------------------------
        if "data_storage" not in st.session_state:
            st.session_state.data_storage = {}

        if "ikpa_loaded" not in st.session_state:
            with st.spinner("🔄 Memuat data IKPA dari GitHub..."):
                loaded = load_data_from_github()
                if loaded:
                    st.session_state.data_storage = loaded
                    add_notification("Data IKPA berhasil dimuat dari GitHub")
                else:
                    st.warning("⚠️ Data IKPA belum tersedia")
            st.session_state.ikpa_loaded = True

        # --------------------------------------------------------
        # 3. LOAD IKPA KPPN
        # --------------------------------------------------------
        if "data_storage_kppn" not in st.session_state:
            st.session_state.data_storage_kppn = {}

        if not st.session_state.data_storage_kppn:
            result_kppn = load_data_ikpa_kppn_from_github()
            if result_kppn:
                st.session_state.data_storage_kppn = result_kppn

        if st.session_state.data_storage_kppn and not st.session_state.get("_kppn_loaded_notif"):
            add_notification("Data IKPA KPPN berhasil dimuat dari GitHub")
            st.session_state["_kppn_loaded_notif"] = True

        # --------------------------------------------------------
        # 4. LOAD DIPA
        # --------------------------------------------------------
        if "DATA_DIPA_by_year" not in st.session_state:
            st.session_state.DATA_DIPA_by_year = {}

        if not st.session_state.DATA_DIPA_by_year:
            with st.spinner("🔄 Memuat data DIPA dari GitHub..."):
                load_DATA_DIPA_from_github()

        # Finalisasi kolom RINGKAS (hanya sekali)
        if st.session_state.DATA_DIPA_by_year and not st.session_state.get("_dipa_finalized"):
            for tahun, df in st.session_state.DATA_DIPA_by_year.items():
                df = df.copy()
                if "Uraian Satker" in df.columns:
                    df["Uraian Satker-RINGKAS"] = (
                        df["Uraian Satker"]
                        .fillna("-")
                        .astype(str)
                        .str[:30]
                    )
                else:
                    df["Uraian Satker-RINGKAS"] = "-"
                st.session_state.DATA_DIPA_by_year[tahun] = df
            st.session_state["_dipa_finalized"] = True

        # --------------------------------------------------------
        # 5. AUTO MERGE IKPA + DIPA
        # --------------------------------------------------------
        if not st.session_state.get("ikpa_dipa_merged", False):
            if st.session_state.data_storage and st.session_state.DATA_DIPA_by_year:
                try:
                    merge_ikpa_dipa_auto()
                    st.session_state.ikpa_dipa_merged = True
                except Exception as e:
                    st.error(f"Gagal merge IKPA & DIPA: {e}")

        if st.session_state.get("ikpa_dipa_merged", False):
            add_notification("Data IKPA & DIPA berhasil dimuat dan siap digunakan")

        # --------------------------------------------------------
        # 6. LOAD KKP
        # --------------------------------------------------------
        if "kkp_master" not in st.session_state:
            kkp_df, kkp_loaded = load_kkp_master_from_github()
            st.session_state.kkp_master = kkp_df
            if kkp_loaded:
                add_notification("Database utama KKP berhasil dimuat dari GitHub")

        # --------------------------------------------------------
        # 7. LOAD CMS & DIGIPAY
        # --------------------------------------------------------
        cms_sudah_ada = (
            "cms_master" in st.session_state and
            not st.session_state.cms_master.empty
        )
        if not cms_sudah_ada:
            cms_df, cms_count = load_cms_from_github()
            st.session_state.cms_master = cms_df
            st.session_state.auto_loaded_cms = True
            if cms_count > 0:
                add_notification("Data CMS berhasil dimuat")
        elif "auto_loaded_cms" not in st.session_state:
            st.session_state.auto_loaded_cms = True

        digipay_sudah_ada = (
            "digipay_master" in st.session_state and
            not st.session_state.digipay_master.empty
        )
        if not digipay_sudah_ada:
            digipay_df, digipay_count = load_digipay_from_github()
            st.session_state.digipay_master = digipay_df
            st.session_state.auto_loaded_digipay = True
            if digipay_count > 0:
                add_notification("Data DIGIPAY berhasil dimuat")
        elif "auto_loaded_digipay" not in st.session_state:
            st.session_state.auto_loaded_digipay = True

        if (
            st.session_state.get("auto_loaded_cms") or
            st.session_state.get("auto_loaded_digipay")
        ):
            add_notification("Data CMS & DIGIPAY berhasil dimuat dan siap digunakan")

        # Tandai bahwa loading sudah selesai — saat pindah menu, is_first_load = False
        st.session_state.app_fully_loaded = True

        # Hapus logo splash
        loading_placeholder.empty()

    # ===============================
    # PANEL STATUS SISTEM
    # ===============================
    if st.session_state.show_system_status and "loading_notifications" in st.session_state:

        st.markdown("""
        <style>

        /* SYSTEM STATUS BOX */
        .system-status{
            background:#f1f5f9;
            border-radius:10px;
            border:1px solid #e2e8f0;

            padding:10px 16px;   /* sebelumnya lebih besar */
            margin-bottom:8px;   /* lebih rapat ke elemen bawah */
        }

        /* JUDUL */
        .system-title{
            font-size:15px;      /* sebelumnya sekitar 18-20 */
            font-weight:600;
            margin-bottom:3px;
        }

        /* TEXT STATUS */
        .status-item{
            font-size:13px;
            line-height:1.4;
        }

        /* DOT HIJAU */
        .status-ok{
            color:#16a34a;
            margin-right:6px;
        }

        </style>
        """, unsafe_allow_html=True)

        with st.expander("Detail proses loading sistem"):

            for msg in st.session_state.loading_notifications:

                st.markdown(f"""
                <div class="status-item">
                <span class="status-ok">●</span> {msg}
                </div>
                """, unsafe_allow_html=True)
        
        st.session_state.show_system_status = False
            
            
    # ============================================================
    # Sidebar + Routing halaman
    # ============================================================
    st.sidebar.title("🧭 Navigasi")
    st.sidebar.markdown("---")

    if "page" not in st.session_state:
        st.session_state.page = "Dashboard Utama"

    if st.sidebar.button("📊 Dashboard Utama", use_container_width=True):
        st.session_state.page = "Dashboard Utama"

    if st.sidebar.button("📈 Dashboard Internal", use_container_width=True):
        st.session_state.page = "Dashboard Internal"

    if st.sidebar.button("🔐 Admin", use_container_width=True):
        st.session_state.page = "Admin"

    st.sidebar.markdown("---")

    st.sidebar.info("""
    **Dashboard IKPA**  
    Indikator Kinerja Pelaksanaan Anggaran  
    KPPN Baturaja  

    📧 Support: ameer.noor@kemenkeu.go.id
    """)

    # ===============================
    # Routing Halaman
    # ===============================
    if st.session_state.page == "Dashboard Utama":
        page_dashboard()

    elif st.session_state.page == "Dashboard Internal":
        page_trend()

    elif st.session_state.page == "Admin":
        page_admin()

# ===============================
# 🔹 ENTRY POINT
# ===============================
if __name__ == "__main__":
    main()
